# -*- coding: cp860 -*-

import sagon.sagist as sg
import openpyxl as ox
import sagon.xlsage as xs
import sagon.datapi as dt
import optparse
import os

def base2xls(base_path='', filename='config.xlsx', model_wb='modelo.xlsx', **kwargs):
    wb = ox.load_workbook(model_wb)

    if xs.check_wb(wb, **kwargs):
        dt.print_msg(__name__, 'tudo OK',dt.MSG_INFO, **kwargs)
        xs.clear_wb(wb)

        if base_path == '':
            source = wb['CONF']['B1'].value
        else:
            source = base_path
            wb['CONF']['B1'].value = os.path.abspath(base_path)

        if kwargs.get('include_cmts') is None:
            include_cmts = str(wb['CONF']['B2'].value).upper() in ['S','SIM','X']
        else:
            include_cmts = kwargs.get('include_cmts')
            if include_cmts:
                wb['CONF']['B2'].value = 'SIM'
            else:
                wb['CONF']['B2'].value = None


        if include_cmts:
            dt.print_msg(__name__, 'coment rios ser„o tratados', msg_type=dt.MSG_INFO, **kwargs)
        else:
            dt.print_msg(__name__, 'coment rios ser„o ignorados', msg_type=dt.MSG_INFO, **kwargs)

        dt.print_msg(__name__,'lendo de {0}'.format(source),msg_type=dt.MSG_INFO, **kwargs)

        wsPDS = wb['PDS']
        wsCALC = wb['CALC']
        wsFILS = wb['FILS']
        wsFILC = wb['FILC']
        wsPAS = wb['PAS']
        wsPTS = wb['PTS']
        wsCGS = wb['CGS']

        calc_row = 2
        filc_row = 2
        fils_row = 3
        base = sg.load_base(source_path=source, **kwargs)
        for ws in wb:
            dat_type = str(ws.title).lower()
            #Loop principal
            if dat_type in ['pds','pas','pts','cgs']:
                dat = base[dat_type]
                if dat_type == 'pds':
                    COLS = xs.PDS_COLS
                elif dat_type == 'pas':
                    COLS = xs.PAS_COLS
                elif dat_type == 'pts':
                    COLS = xs.PTS_COLS
                elif dat_type == 'cgs':
                    COLS = xs.CGS_COLS
                row=3
                print('Lendo {0}...'.format(dat_type.upper()))
                row = 3
                key_number = 1
                total_keys = len(dat.keys())
                for key in dt.list_keys(dat):
                    print('Processando arquivo ({0} de {1}): {2}'.format(key_number, total_keys, key))
                    interation = 0
                    total = len(dat[key])
                    if total==0:
                        total=1
                    xs.printProgress(interation, total, prefix = 'Progresso:', suffix = 'Completo', barLength = 50)
                    for dat_item in dat[key]:
                        # loop de itera‡„o sobre os itens
                        if (not dt.is_comment(dat_item)):
                            if dt.is_commented_point(dat_item) and include_cmts:
                                dat_item = dt.clean_commented_point(dat_item)
                                ws[COLS[xs.cCOMENTADO]+str(row)].value = 'X'
                            if dat_item.get('ID','')=='':
                                #ponto mal formado (sem id)
                                interation +=1
                                xs.printProgress(interation, total, prefix = 'Progresso:', suffix = 'Completo', barLength = 50)
                                continue


                            dat_conf = sg.get_aconf_from_base(dat_type, item_id=dat_item.get('ID',''), base_item=base, **kwargs)

                            #checa se ‚ roteamento de controle e pula se for o caso
                            if (dat_type == 'cgs') and dat_conf.get('cgf'):
                                if (len(str(dat_conf.get('cgf').get('items')[0].get('ID')).split('-')) == 3):
                                    interation +=1
                                    xs.printProgress(interation, total, prefix = 'Progresso:', suffix = 'Completo', barLength = 50)
                                    continue

                            ws[COLS[xs.cINCLUDE]+str(row)].value = str(key).lstrip('#')
                            ws[COLS[xs.cID]+str(row)].value = dat_item.get('ID')
                            ws[COLS[xs.cNOME]+str(row)].value = dat_item.get('NOME')
                            ws[COLS[xs.cTAC]+str(row)].value = dat_item.get('TAC')

                            # extrai metacampos de cmt
                            if '|' in dat_item.get('CMT',''):
                                try:
                                    testado, vao, ied, origem = str(dat_item.get('CMT')).split('|')
                                    ws[COLS[xs.cTESTE]+str(row)].value = str(testado).upper()
                                    ws[COLS[xs.cVAO]+str(row)].value = vao
                                    ws[COLS[xs.cIED]+str(row)].value = ied
                                    ws[COLS[xs.cORIGEM]+str(row)].value = origem
                                except:
                                    pass

                            if xs.get_cell_string(ws[COLS[xs.cIED]+str(row)]) == '' and dat_conf.get('lsc'):
                                try:
                                    ws[COLS[xs.cIED]+str(row)].value = dat_conf.get('lsc').get('item').get('ID')
                                except:
                                    print('ID: ' + str(dat_item.get('ID', '')))
                                    print('TIPO : {}'.format(dat_type))
                            # campos extras comuns a todos pds, pas, pts, cgs
                            ws[COLS[xs.cIDICCP]+str(row)].value = dat_item.get('IDICCP')
                            ws[COLS[xs.cOBSRV]+str(row)].value = dat_item.get('OBSRV')



                            # demais campos gerais de pas, pds, pts
                            if dat_type !='cgs':
                                ws[COLS[xs.cCALCULO]+str(row)].value = dat_item.get('TCL')
                                print(ws[COLS[xs.cCALCULO]+str(row)].value)
                                ws[COLS[xs.cFILTRO]+str(row)].value = dat_item.get('TPFIL')
                                ws[COLS[xs.cALINT]+str(row)].value = dat_item.get('ALINT')
                                ws[COLS[xs.cATLZINV]+str(row)].value = dat_item.get('ATLZINV')
                                ws[COLS[xs.cCDINIC]+str(row)].value = dat_item.get('CDINIC')
                                ws[COLS[xs.cSELSD]+str(row)].value = dat_item.get('SELSD')
                                ws[COLS[xs.cSINCR_MAN]+str(row)].value = dat_item.get('SINCR_MAN')
                                ws[COLS[xs.cUAPL]+str(row)].value = dat_item.get('UAPL')
                                ws[COLS[xs.cOCR]+str(row)].value = dat_item.get('OCR')
                                ALRIN = dat_item.get('ALRIN')
                                if str(ALRIN).upper() == 'NAO':
                                    ws[COLS[xs.cALARME]+str(row)].value = 'X'

                            if dat_type == 'pds':
                                # campos de pds
                                SOEIN = dat_item.get('SOEIN')
                                if str(SOEIN).upper() == 'NAO':
                                    ws[COLS[xs.cSOE]+str(row)].value = 'X'

                                ws[COLS[xs.cINVRT]+str(row)].value = dat_item.get('INVRT')
                                ws[COLS[xs.cSTINI]+str(row)].value = dat_item.get('STINI')
                                ws[COLS[xs.cSTNOR]+str(row)].value = dat_item.get('STNOR')

                            elif dat_type == 'pts':
                                ws[COLS[xs.cHISTPER]+str(row)].value = dat_item.get('HISTPER')
                                ws[COLS[xs.cHISTSLC]+str(row)].value = dat_item.get('HISTSLC')
                                ws[COLS[xs.cLSA]+str(row)].value = dat_item.get('LSA')
                                ws[COLS[xs.cLSE]+str(row)].value = dat_item.get('LSE')
                                ws[COLS[xs.cLSU]+str(row)].value = dat_item.get('LSU')
                                ws[COLS[xs.cTXVAR]+str(row)].value = dat_item.get('TXVAR')
                                ws[COLS[xs.cVLINIC]+str(row)].value = dat_item.get('VLINIC')


                            elif dat_type == 'pas':
                                # campos de pas

                                ws[COLS[xs.cBNDMO]+str(row)].value = dat_item.get('BNDMO')
                                ws[COLS[xs.cBDTR]+str(row)].value = dat_item.get('BDTR')
                                ws[COLS[xs.cHISTPER]+str(row)].value = dat_item.get('HISTPER')
                                ws[COLS[xs.cHISTSLC]+str(row)].value = dat_item.get('HISTSLC')
                                ws[COLS[xs.cHTRIS]+str(row)].value = dat_item.get('HTRIS')
                                ws[COLS[xs.cLIA]+str(row)].value = dat_item.get('LIA')
                                ws[COLS[xs.cLIALE]+str(row)].value = dat_item.get('LIALE')
                                ws[COLS[xs.cLIAMA]+str(row)].value = dat_item.get('LIAMA')
                                ws[COLS[xs.cLIAME]+str(row)].value = dat_item.get('LIAME')
                                ws[COLS[xs.cLIAMI]+str(row)].value = dat_item.get('LIAMI')
                                ws[COLS[xs.cLIAPE]+str(row)].value = dat_item.get('LIAPE')
                                ws[COLS[xs.cLIE]+str(row)].value = dat_item.get('LIE')
                                ws[COLS[xs.cLIU]+str(row)].value = dat_item.get('LIU')
                                ws[COLS[xs.cLIULE]+str(row)].value = dat_item.get('LIULE')
                                ws[COLS[xs.cLIUMA]+str(row)].value = dat_item.get('LIUMA')
                                ws[COLS[xs.cLIUME]+str(row)].value = dat_item.get('LIUME')
                                ws[COLS[xs.cLIUMI]+str(row)].value = dat_item.get('LIUMI')
                                ws[COLS[xs.cLIUPE]+str(row)].value = dat_item.get('LIUPE')
                                ws[COLS[xs.cLSA]+str(row)].value = dat_item.get('LSA')
                                ws[COLS[xs.cLSALE]+str(row)].value = dat_item.get('LSALE')
                                ws[COLS[xs.cLSAMA]+str(row)].value = dat_item.get('LSAMA')
                                ws[COLS[xs.cLSAME]+str(row)].value = dat_item.get('LSAME')
                                ws[COLS[xs.cLSAMI]+str(row)].value = dat_item.get('LSAMI')
                                ws[COLS[xs.cLSAPE]+str(row)].value = dat_item.get('LSAPE')
                                ws[COLS[xs.cLSE]+str(row)].value = dat_item.get('LSE')
                                ws[COLS[xs.cLSU]+str(row)].value = dat_item.get('LSU')
                                ws[COLS[xs.cLSULE]+str(row)].value = dat_item.get('LSULE')
                                ws[COLS[xs.cLSUMA]+str(row)].value = dat_item.get('LSUMA')
                                ws[COLS[xs.cLSUME]+str(row)].value = dat_item.get('LSUME')
                                ws[COLS[xs.cLSUMI]+str(row)].value = dat_item.get('LSUMI')
                                ws[COLS[xs.cLSUPE]+str(row)].value = dat_item.get('LSUPE')
                                ws[COLS[xs.cPTC]+str(row)].value = dat_item.get('PTC')
                                ws[COLS[xs.cTEND]+str(row)].value = dat_item.get('TEND')
                                ws[COLS[xs.cTXVAR]+str(row)].value = dat_item.get('TXVAR')
                                ws[COLS[xs.cVLINIC]+str(row)].value = dat_item.get('VLINIC')

                            elif dat_type == 'cgs':
                                # campos de cgs
                                ws[COLS[xs.cCONTROLE]+str(row)].value = dat_item.get('TIPOE')
                                ws[COLS[xs.cSUPERVISAO]+str(row)].value = dat_item.get('PAC')
                                ws[COLS[xs.cTIPO]+str(row)].value = dat_item.get('TIPO')
                                ws[COLS[xs.cINTERTRAV]+str(row)].value = dat_item.get('PINT')
                                ws[COLS[xs.cINVCT]+str(row)].value = dat_item.get('INVCT')
                                ws[COLS[xs.cLMI1C]+str(row)].value = dat_item.get('LMI1C')
                                ws[COLS[xs.cLMI2C]+str(row)].value = dat_item.get('LMI2C')
                                ws[COLS[xs.cLMS1C]+str(row)].value = dat_item.get('LMS1C')
                                ws[COLS[xs.cLMS2C]+str(row)].value = dat_item.get('LMS2C')
                                ws[COLS[xs.cRSULT]+str(row)].value = dat_item.get('RSULT')
                                ws[COLS[xs.cTRRAC]+str(row)].value = dat_item.get('TRRAC')
                                ws[COLS[xs.cTPCTL]+str(row)].value = dat_item.get('TPCTL')



                            # se for um filtro composto
                            if 'rfc' in list(dat_conf.keys()):
                                i=0
                                for rfc in dat_conf.get('rfc').get('items'):
                                    wsFILC[xs.FILC_COLS[xs.cORDEM]+str(filc_row)].value = rfc.get('ORDEM','')
                                    wsFILC[xs.FILC_COLS[xs.cPNT]+str(filc_row)].value = rfc.get('PNT','')
                                    wsFILC[xs.FILC_COLS[xs.cTPPNT]+str(filc_row)].value = rfc.get('TPPNT','')
                                    wsFILC[xs.FILC_COLS[xs.cCMT]+str(filc_row)].value = rfc.get('CMT','')
                                    wsFILC[xs.FILC_COLS[xs.cPARC]+str(filc_row)].value = rfc.get('PARC','')
                                    wsFILC[xs.FILC_COLS[xs.cTPPARC]+str(filc_row)].value = rfc.get('TPPARC','')
                                    wsFILC[xs.FILC_COLS[xs.cINCLUDE]+str(filc_row)].value = str(dat_conf.get('rfc').get('locations')[i]).lstrip('#')
                                    if '|' in dat_conf['rfc']['items'][i].get('CMT',''):
                                        try:
                                            testado, vao, ied, origem = str(dat_conf['rfc']['items'][i].get('CMT','')).split('|')
                                            wsFILC[xs.FILC_COLS[xs.cTESTE]+str(filc_row)].value = str(testado).upper()
                                            wsFILC[xs.FILC_COLS[xs.cVAO]+str(filc_row)].value = vao
                                        except:
                                            pass
                                    filc_row+=1
                                    i+=1

                            # se for um filtro simples
                            if 'rfi' in list(dat_conf.keys()):
                                i=0
                                for rfi in dat_conf.get('rfi').get('items'):
                                    wsFILS[xs.FILS_COLS[xs.cORDEM]+str(fils_row)].value = rfi.get('ORDEM')
                                    wsFILS[xs.FILS_COLS[xs.cPNT]+str(fils_row)].value = rfi.get('PNT')
                                    wsFILS[xs.FILS_COLS[xs.cTIPOP]+str(fils_row)].value = rfi.get('TIPOP')
                                    #wsFILS[xs.FILS_COLS[xs.cCMT]+str(fils_row)].value = rfi.get('CMT')
                                    wsFILS[xs.FILS_COLS[xs.cINCLUDE]+str(fils_row)].value = str(dat_conf.get('rfi').get('locations')[i]).lstrip('#')

                                    if str(rfi.get('TIPOP')).upper() == 'PDF':
                                        # preenche campos pdf da planilha FILS
                                        wsFILS[xs.FILS_COLS[xs.cPDF_PNT]+str(fils_row)].value = dat_conf.get('pdf').get('items')[i].get('PNT')
                                        wsFILS[xs.FILS_COLS[xs.cPDF_TPPNT]+str(fils_row)].value = dat_conf.get('pdf').get('items')[i].get('TPPNT')
                                        wsFILS[xs.FILS_COLS[xs.cPDF_KCONV]+str(fils_row)].value = dat_conf.get('pdf').get('items')[i].get('KCONV')
                                        wsFILS[xs.FILS_COLS[xs.cPDF_NV2]+str(fils_row)].value = dat_conf.get('pdf').get('items')[i].get('NV2')
                                        wsFILS[xs.FILS_COLS[xs.cPDF_ORDEM]+str(fils_row)].value = dat_conf.get('pdf').get('items')[i].get('ORDEM')
                                        wsFILS[xs.FILS_COLS[xs.cPDF_DESC1]+str(fils_row)].value = dat_conf.get('pdf').get('items')[i].get('DESC1')
                                        wsFILS[xs.FILS_COLS[xs.cPDF_DESC2]+str(fils_row)].value = dat_conf.get('pdf').get('items')[i].get('DESC2')

                                    elif str(rfi.get('TIPOP')).upper() == 'PAF':
                                        # preenche campos pdf da planilha FILS
                                        wsFILS[xs.FILS_COLS[xs.cPAF_PNT]+str(fils_row)].value = dat_conf.get('paf').get('items')[i].get('PNT')
                                        wsFILS[xs.FILS_COLS[xs.cPAF_TPPNT]+str(fils_row)].value = dat_conf.get('paf').get('items')[i].get('TPPNT')
                                        wsFILS[xs.FILS_COLS[xs.cPAF_KCONV1]+str(fils_row)].value = dat_conf.get('paf').get('items')[i].get('KCONV1')
                                        wsFILS[xs.FILS_COLS[xs.cPAF_KCONV2]+str(fils_row)].value = dat_conf.get('paf').get('items')[i].get('KCONV2')
                                        wsFILS[xs.FILS_COLS[xs.cPAF_KCONV3]+str(fils_row)].value = dat_conf.get('paf').get('items')[i].get('KCONV3')
                                        wsFILS[xs.FILS_COLS[xs.cPAF_NV2]+str(fils_row)].value = dat_conf.get('paf').get('items')[i].get('NV2')
                                        wsFILS[xs.FILS_COLS[xs.cPAF_ORDEM]+str(fils_row)].value = dat_conf.get('paf').get('items')[i].get('ORDEM')
                                        wsFILS[xs.FILS_COLS[xs.cPAF_DESC1]+str(fils_row)].value = dat_conf.get('paf').get('items')[i].get('DESC1')
                                        wsFILS[xs.FILS_COLS[xs.cPAF_DESC2]+str(fils_row)].value = dat_conf.get('paf').get('items')[i].get('DESC2')

                                    elif str(rfi.get('TIPOP')).upper() == 'PTF':
                                        # preenche campos pdf da planilha FILS
                                        wsFILS[xs.FILS_COLS[xs.cPTF_PNT]+str(fils_row)].value = dat_conf.get('ptf').get('items')[i].get('PNT')
                                        wsFILS[xs.FILS_COLS[xs.cPTF_TPPNT]+str(fils_row)].value = dat_conf.get('ptf').get('items')[i].get('TPPNT')
                                        wsFILS[xs.FILS_COLS[xs.cPTF_KCONV1]+str(fils_row)].value = dat_conf.get('ptf').get('items')[i].get('KCONV1')
                                        wsFILS[xs.FILS_COLS[xs.cPTF_KCONV2]+str(fils_row)].value = dat_conf.get('ptf').get('items')[i].get('KCONV2')
                                        wsFILS[xs.FILS_COLS[xs.cPTF_KCONV3]+str(fils_row)].value = dat_conf.get('ptf').get('items')[i].get('KCONV3')
                                        wsFILS[xs.FILS_COLS[xs.cPTF_NV2]+str(fils_row)].value = dat_conf.get('ptf').get('items')[i].get('NV2')
                                        wsFILS[xs.FILS_COLS[xs.cPTF_ORDEM]+str(fils_row)].value = dat_conf.get('ptf').get('items')[i].get('ORDEM')
                                        wsFILS[xs.FILS_COLS[xs.cPTF_DESC1]+str(fils_row)].value = dat_conf.get('ptf').get('items')[i].get('DESC1')
                                        wsFILS[xs.FILS_COLS[xs.cPTF_DESC2]+str(fils_row)].value = dat_conf.get('ptf').get('items')[i].get('DESC2')


                                    if '|' in dat_conf['rfi']['items'][i].get('CMT',''):
                                        try:
                                            testado, vao, ied, origem = str(dat_conf['rfi']['items'][i].get('CMT','')).split('|')
                                            wsFILS[xs.FILS_COLS[xs.cTESTE]+str(fils_row)].value = str(testado).upper()
                                            wsFILS[xs.FILS_COLS[xs.cVAO]+str(fils_row)].value = vao
                                            wsFILS[xs.FILS_COLS[xs.cIED]+str(fils_row)].value = ied
                                            wsFILS[xs.FILS_COLS[xs.cORIGEM]+str(fils_row)].value = origem
                                        except:
                                            pass
                                    fils_row+=1
                                    i+=1

                            # caso seja um ponto f¡sico aquisitado, preencher conf f¡sica
                            dat_typef = dat_type[:2]+'f'
                            if dat_typef in list(dat_conf.keys()):

                                # n„o for um filtro
                                if len(dat_conf[dat_typef]['items']) == 1:
                                    if sg.is_61850(dat_type,item_id=dat_item['ID'],base_item=base):
                                        ws[COLS[xs.cENDERECO]+str(row)].value = xs.expand_address(dat_type=dat_typef, aconf=dat_conf)
                                    else:
                                        ws[COLS[xs.cENDERECO]+str(row)].value = dat_conf[dat_typef]['items'][0].get('ID')
                                    ws[COLS[xs.cDESC1]+str(row)].value = dat_conf[dat_typef]['items'][0].get('DESC1')
                                    ws[COLS[xs.cNV2]+str(row)].value = dat_conf[dat_typef]['items'][0].get('NV2')
                                    ws[COLS[xs.cORDEM]+str(row)].value = dat_conf[dat_typef]['items'][0].get('ORDEM')
                                    if dat_type == 'pds':
                                        # campos f¡sicos de pds
                                        ws[COLS[xs.cKCONV]+str(row)].value = dat_conf[dat_typef]['items'][0].get('KCONV')
                                    if dat_type in ['pas','pts']:
                                        # campos f¡sicos pas
                                        ws[COLS[xs.cKCONV1]+str(row)].value = dat_conf[dat_typef]['items'][0].get('KCONV1')
                                        ws[COLS[xs.cKCONV2]+str(row)].value = dat_conf[dat_typef]['items'][0].get('KCONV2')
                                        ws[COLS[xs.cKCONV3]+str(row)].value = dat_conf[dat_typef]['items'][0].get('KCONV3')
                                    if dat_type == 'cgs':
                                        # campos f¡sicos cgs
                                        ws[COLS[xs.cKCONV]+str(row)].value = dat_conf[dat_typef]['items'][0].get('KCONV')
                                        ws[COLS[xs.cOPCOES]+str(row)].value = dat_conf[dat_typef]['items'][0].get('OPCOES')
                                        ws[COLS[xs.cCNF]+str(row)].value = dat_conf[dat_typef]['items'][0].get('CNF')



                            # caso seja um ponto calculado, preencher planilha CALC
                            if 'rca' in list(dat_conf):
                                for i in range(0,len(dat_conf['rca']['items'])):
                                    wsCALC[xs.CALC_COLS[xs.cPNT]+str(calc_row)].value = dat_conf['rca']['items'][i].get('PNT')
                                    wsCALC[xs.CALC_COLS[xs.cTPPNT]+str(calc_row)].value = dat_conf['rca']['items'][i].get('TPPNT')
                                    wsCALC[xs.CALC_COLS[xs.cORDEM]+str(calc_row)].value = dat_conf['rca']['items'][i].get('ORDEM')
                                    wsCALC[xs.CALC_COLS[xs.cPARC]+str(calc_row)].value = dat_conf['rca']['items'][i].get('PARC')
                                    wsCALC[xs.CALC_COLS[xs.cTPPARC]+str(calc_row)].value = dat_conf['rca']['items'][i].get('TPPARC')
                                    wsCALC[xs.CALC_COLS[xs.cTIPOP]+str(calc_row)].value = dat_conf['rca']['items'][i].get('TIPOP')
                                    #wsCALC[xs.CALC_COLS[xs.cCMT]+str(calc_row)].value = dat_conf['rca']['items'][i].get('CMT')
                                    wsCALC[xs.CALC_COLS[xs.cINCLUDE]+str(calc_row)].value = str(dat_conf['rca']['locations'][i]).lstrip('#')
                                    if '|' in dat_conf['rca']['items'][i].get('CMT',''):
                                        try:
                                            testado, vao, ied, origem = str(dat_conf['rca']['items'][i].get('CMT','')).split('|')
                                            wsCALC[COLS[xs.cTESTE]+str(calc_row)].value = str(testado).upper()
                                            wsCALC[COLS[xs.cVAO]+str(calc_row)].value = vao
                                        except:
                                            pass
                                    calc_row +=1


                            row+=1
                        elif include_cmts:
                            if dt.is_comment(dat_item):
                                ws[COLS[xs.cCOMENTARIO]+str(row)].value = dat_item['comment']
                                ws[COLS[xs.cCOMENTADO]+str(row)].value = 'X'


                            row+=1
                        interation +=1
                        xs.printProgress(interation, total, prefix = 'Progresso:', suffix = 'Completo', barLength = 50)
                        # fim do loop de itera‡„o sobre os itens
                    key_number+=1
                # FIM DA LEITURA DA PLANILHA
        xs.color_wb(wb)
        wb.save(os.path.join(source,filename))
    else:
        print('MELOU')
    print(wb.sheetnames)

def main():
    parser = optparse.OptionParser()
    parser.add_option('-f','--file', dest='filename', default='config.xlsx', help='Nome do xls de destino', metavar='FILE')
    parser.add_option('-q','--quiet', dest='verbose', default=True, action='store_false',
                      help='n„o imprime mensagens de progresso do script')
    parser.add_option('-i','--ignore_cmts', action='store_false', dest='include_cmts',
                      default=True, help='ignora linhas comentadas da base')
    parser.add_option('-m','--model_file', default='modelo.xlsx', dest='model_file',
                      help='arquivo modelo xls')
    (options, args) = parser.parse_args()
    print('options:', str(options))
    print('arguments:',args)
    if len(args) !=1:
        base_path=''
    else:
        base_path=str(args[0])

    base2xls(base_path=base_path, filename=options.filename, model_wb=options.model_file, include_cmts=options.include_cmts, verbose=options.verbose)

if __name__ == '__main__':
    main()
else:
    print('base2xls carregado como m¢dulo')
