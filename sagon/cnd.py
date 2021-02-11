import objsage as obs
import openpyxl as ox


def make_cnd(file='cnd.xlsx', base='teste/cnd', out='teste/cnd_out',
             wsname='ECND', include='ecnd', row_start=7, row_end=206,
             num_gen=14, row_start2 = 213, row_end2=256):

    wb = ox.load_workbook(file)
    ws = wb[wsname]
    col_addr = 'H'
    col_dado = 'J'
    col_id = 'L'
    col_ocr = 'M'
    col_desc = 'N'
    col_tipo = 'O'
    col_com = 'P'
    col_med = 'Q'
    col_tela = 'R'
    col_alrm = 'S'
    col_soe = 'T'
    b = obs.Base(source=base)



    apfl_id = b.nv2[include].point({'TN2':'APFL'}).get_value('ID')
    aanl_id = b.nv2[include].point({'TN2': 'AANL'}).get_value('ID')
    asim_id = b.nv2[include].point({'TN2':'ASIM'}).get_value('ID')
    cstp_id = b.nv2[include].point({'TN2': 'CSTP'}).get_value('ID')

    row_end += 1

    if not b.rca.has_include(include):
        b.rca.add_include(include)

    if not b.grupo.has_include(include):
        b.grupo.add_include(include)

    if not b.grcmp.has_include(include):
        b.grcmp.add_include(include)


    for row in range(row_start, row_end):
        addr = str(ws[col_addr+str(row)].value)
        dado = str(ws[col_dado+str(row)].value)
        id = str(ws[col_id+str(row)].value)
        ocr = str(ws[col_ocr+str(row)].value)
        desc = str(ws[col_desc+str(row)].value)
        tipo = str(ws[col_tipo+str(row)].value)
        cmd = str(ws[col_com+str(row)].value)
        med = str(ws[col_med+str(row)].value)
        tela = str(ws[col_tela+str(row)].value)
        alrm = str(ws[col_alrm+str(row)].value)
        soe = str(ws[col_soe+str(row)].value)
        print('{} - {} - {}'.format(id, desc, ocr))

        if dado in ['36','35']: # PAS
            # cria paf
            paf = {}
            if dado == '36':
                nv2_id = apfl_id
            else:
                nv2_id = aanl_id
            paf['ID'] = nv2_id + '_' + addr
            paf['NV2'] = nv2_id
            paf['DESC1'] = desc[:40]
            paf['PNT'] = id
            paf['TPPNT'] = 'PAS'
            paf['KCONV1'] = '1'
            paf['KCONV2'] = '0'
            paf['ORDEM'] = addr

            # cria pas

            pas = {}
            pas['ID'] = id
            pas['IDICCP'] = id.replace(':','_')
            pas['NOME'] = desc
            pas['ALRIN'] = 'NAO'
            pas['ALINT'] = 'SIM'
            pas['TCL'] = 'NLCL'
            pas['TAC'] = 'TAC_' + include.upper()
            pas['TPFIL'] = 'NLFL'
            pas['OCR'] = 'OCR_PAS01'
            pas['LIE'] = '-99999'
            pas['LSA'] = '99999'
            pas['LIU'] = '-99999'
            pas['LSU'] = '99999'
            pas['LSE'] = '99999'
            pas['LIA'] = '-99999'

            # adiciona pontos à base

            b.paf[include].add_point(paf)
            b.pas[include].add_point(pas)

        if dado == '30': # PDS
            pdf = {}
            pdf['ID'] = asim_id + '_' + addr
            pdf['NV2'] = asim_id
            pdf['PNT'] = id
            pdf['TPPNT'] = 'PDS'
            pdf['KCONV'] = 'SQN'
            pdf['ORDEM'] = addr
            pdf['DESC1'] = desc[:40]


            pds = {}
            pds['ID'] = id
            pds['IDICCP'] = id.replace(':','_')
            pds['NOME'] = desc
            pds['ALRIN'] = 'NAO'
            pds['SOEIN'] = 'NAO'
            pds['ALINT'] = 'SIM'
            pds['TCL'] = 'NLCL'
            pds['TAC'] = 'TAC_' + include.upper()
            pds['TPFIL'] = 'NLFL'
            pds['OCR'] = ocr
            pds['STINI'] = 'A'
            pds['STNOR'] = 'A'

            b.pdf[include].add_point(pdf)
            b.pds[include].add_point(pds)

        if dado == '63':
            cgf = {}
            cgf['ID'] = cstp_id + '_' + addr
            cgf['NV2'] = cstp_id
            cgf['CGS'] = id
            cgf['ORDEM'] = addr
            cgf['DESC1'] = desc[:40]
            cgf['KCONV'] = 'NO'

            cgs = {}
            cgs['ID'] = id
            cgs['IDICCP'] = id.replace(':','_')+'C'
            cgs['NOME'] = desc
            cgs['TAC'] = 'TAC_' + include.upper()
            cgs['PAC'] = id
            cgs['TIPO'] = 'PAS'
            cgs['TPCTL'] = 'CSAC'
            cgs['TIPOE'] = 'STPT'
            cgs['TRRAC'] = '30'
            cgs['LMI1C'] = '0'
            cgs['LMI2C'] = '0'
            cgs['LMS1C'] = '0'
            cgs['LMS2C'] = '0'

            b.cgf[include].add_point(cgf)
            b.cgs[include].add_point(cgs)

        if (dado.upper() == 'CALC') and ('NGER' not in id):
            pds = {}
            pds['ALRIN'] = 'NAO'
            pds['ID'] = id
            pds['ALINT'] = 'SIM'
            pds['TAC'] = 'CALC'
            sufix = id.split(':')[-1]
            pds['TCL'] = 'C'+sufix
            pds['NOME'] = desc[:42]
            pds['STNOR'] = 'A'
            pds['TIPO'] = 'PTNI'
            pds['TPFIL'] = 'NLFL'
            pds['SOEIN'] = 'NAO'
            pds['STINI'] = 'A'
            pds['CDINIC'] = 'NORMAL'
            pds['OCR'] = ocr
            pds['IDICCP'] = id.replace(':', '_')

            rca = {}
            rca['PNT'] = id
            rca['TPPARC'] = 'PAS'
            rca['ORDEM'] = '1'
            rca['PARC'] = id.rstrip(sufix) + 'SEQP'
            rca['TPPNT'] = 'PDS'
            rca['TIPOP'] = 'VAC'

            b.pds[include].add_point(pds)
            b.rca[include].add_point(rca)


    # pontos de main status de cada UG

    row_end2 +=1

    for ug in range(1, num_gen+1):

        # cria e adiciona grupo
        grupo = {}
        grupo['ID'] = 'ECND_07G{}_ALARMES'.format(ug)
        grupo['TIPO'] = 'OUTROS'
        grupo['TPPNT'] = 'PDS'
        grupo['NOME'] = '07G{} ALARMES'.format(ug)
        grupo['APLIC'] = 'Vtelas'

        b.grupo[include].add_point(grupo)

        ordem1 = 1
        ordem2 = 1

        for row in range(row_start2, row_end2):

            # contadores para posição dos pontos em grcmp
            if ordem1 == 23:
                ordem1 = 1
                ordem2 +=1

            addr = str(ws[col_addr+str(row)].value)
            dado = str(ws[col_dado+str(row)].value)
            id = str(ws[col_id+str(row)].value)
            id = id.replace(':UG',':07G{}'.format(ug))
            ocr = str(ws[col_ocr+str(row)].value)
            desc = str(ws[col_desc+str(row)].value)
            tipo = str(ws[col_tipo+str(row)].value)
            cmd = str(ws[col_com+str(row)].value)
            med = str(ws[col_med+str(row)].value)
            tela = str(ws[col_tela+str(row)].value)
            alrm = str(ws[col_alrm+str(row)].value)
            soe = str(ws[col_soe+str(row)].value)
            print('{} - {} - {}'.format(id, desc, ocr))

            pds = {}
            pds['ALRIN'] = 'NAO'
            pds['ID'] = id
            pds['ALINT'] = 'SIM'
            pds['TAC'] = 'CALC'
            sufix = id.split(':')[-1]
            pds['TCL'] = 'C' + sufix
            pds['NOME'] = desc[:42]
            pds['STNOR'] = 'A'
            pds['TIPO'] = 'PTNI'
            pds['TPFIL'] = 'NLFL'
            pds['SOEIN'] = 'NAO'
            pds['STINI'] = 'A'
            pds['CDINIC'] = 'NORMAL'
            pds['OCR'] = ocr
            pds['IDICCP'] = id.replace(':', '_')

            rca = {}
            rca['PNT'] = id
            rca['TPPARC'] = 'PAS'
            rca['ORDEM'] = '1'
            rca['PARC'] = id.rstrip(sufix) + 'STPR'
            rca['TPPNT'] = 'PDS'
            rca['TIPOP'] = 'VAC'

            grcmp = {}
            grcmp['GRUPO'] = grupo['ID']
            grcmp['CORTXT'] = 'PRETO'
            grcmp['TPPNT'] = 'PDS'
            grcmp['ORDEM1'] = str(ordem1)
            grcmp['ORDEM2'] = str(ordem2)
            grcmp['PNT'] = id
            grcmp['TPSIMB'] = 'QUAD'
            grcmp['TPTXT'] = 'NOME'

            b.grcmp[include].add_point(grcmp)

            b.pds[include].add_point(pds)
            b.rca[include].add_point(rca)

            ordem1 +=1

    b.del_include('teste')
    b.write_dats(path=out)
    return b





