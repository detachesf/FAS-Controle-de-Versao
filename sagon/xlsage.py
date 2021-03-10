# -*- coding: cp860 -*-

import sagon.sagist as sg
import openpyxl as ox
from openpyxl.styles import colors, PatternFill
from sagon.datapi import *
import sys


SHEETS = (
    'PDS',
    'PAS',
    'CGS',
    'PTS',
    'CALC',
    'FILC',
    'FILS',
    'CONF'
)

cTESTE = 'Teste'
cCOMENTADO = 'Comentado'
cCOMENTARIO = 'Coment rio'
cVAO = 'V„o'
cIED = 'IED'
cORIGEM = 'Origem'
cINCLUDE = 'Include'
cTAC = 'TAC'
cID = 'ID'
cNOME = 'Nome'
cENDERECO = 'Endere‡o'
cOCR = 'OCR'
cTELA = 'Tela'
cALARME = 'Alarme'
cSOE = 'SOE'
cCALCULO = 'C lculo'
cFILTRO = 'Filtro'
cKCONV = 'KCONV'
cNV2 = 'NV2'
cDESC1 = 'DESC1'
cALINT = 'ALINT'
cATLZINV = 'ATLZINV'
cCDINIC = 'CDINIC'
cIDICCP = 'IDICCP'
cINVRT = 'INVRT'
cOBSRV = 'OBSRV'
cSELSD = 'SELSD'
cSINCR_MAN = 'SINCR_MAN'
cSTNOR = 'STNOR'
cUAPL = 'UAPL'
cKCONV1 = 'KCONV1'
cKCONV2 = 'KCONV2'
cKCONV3 = 'KCONV3'
cBDTR = 'BDTR'
cHISTPER = 'HISTPER'
cHISTSLC = 'HISTSLC'
cLIA = 'LIA'
cLIALE = 'LIALE'
cLIAMA = 'LIAMA'
cLIAME = 'LIAME'
cLIAMI = 'LIAMI'
cLIAPE = 'LIAPE'
cLIE = 'LIE'
cLIU = 'LIU'
cLIULE = 'LIULE'
cLIUMA = 'LIUMA'
cLIUME = 'LIUME'
cLIUMI = 'LIUMI'
cLIUPE = 'LIUPE'
cLSA = 'LSA'
cLSALE = 'LSALE'
cLSAMA = 'LSAMA'
cLSAMI = 'LSAMI'
cLSAPE = 'LSAPE'
cLSE = 'LSE'
cLSU = 'LSU'
cLSULE = 'LSULE'
cLSUMA = 'LSUMA'
cLSUME = 'LSUME'
cLSUMI = 'LSUMI'
cLSUPE = 'LSUPE'
cPTC = 'PTC'
cTEND = 'TEND'
cTXVAR = 'TXVAR'
cVLINIC = 'VLINIC'
cBNDMO = 'BNDMO'
cCONTROLE = 'Controle'
cSUPERVISAO = 'Supervis„o'
cTIPO = 'Tipo'
cINTERTRAV = 'Intertravamento'
cOPCOES = 'OPCOES'
cCNF = 'CNF'
cINVCT = 'INVCT'
cLMI1C = 'LMI1C'
cLMI2C = 'LMI2C'
cLMS1C = 'LMS1C'
cLMS2C = 'LMS2C'
cRSULT = 'RSULT'
cTRRAC = 'TRRAC'
cTPCTL = 'TPCTL'
cPNT = 'PNT'
cTPPNT = 'TPPNT'
cORDEM = 'ORDEM'
cPARC = 'PARC'
cTPPARC = 'TPPARC'
cTIPOP = 'TIPOP'
#cCMT = 'CMT'
cCMT = 'Coment rio'
cSTINI = 'STINI'
cHTRIS = 'HTRIS'
cLSAME = 'LSAME'
cPDF_PNT = 'PDF_PNT'
cPDF_TPPNT = 'PDF_TPPNT'
cPDF_KCONV = 'PDF_KCONV'
cPDF_NV2 = 'PDF_NV2'
cPDF_ORDEM = 'PDF_ORDEM'
cPDF_DESC1 = 'PDF_DESC1'
cPDF_DESC2 = 'PDF_DESC2'
cPTF_PNT = 'PTF_PNT'
cPTF_TPPNT = 'PTF_TPPNT'
cPTF_KCONV1 = 'PTF_KCONV1'
cPTF_KCONV2 = 'PTF_KCONV2'
cPTF_KCONV3 = 'PTF_KCONV3'
cPTF_NV2 = 'PTF_NV2'
cPTF_ORDEM = 'PTF_ORDEM'
cPTF_DESC1 = 'PTF_DESC1'
cPTF_DESC2 = 'PTF_DESC2'
cPAF_PNT = 'PAF_PNT'
cPAF_TPPNT = 'PAF_TPPNT'
cPAF_KCONV1 = 'PAF_KCONV1'
cPAF_KCONV2 = 'PAF_KCONV2'
cPAF_KCONV3 = 'PAF_KCONV3'
cPAF_NV2 = 'PAF_NV2'
cPAF_ORDEM = 'PAF_ORDEM'
cPAF_DESC1 = 'PAF_DESC1'
cPAF_DESC2 = 'PAF_DESC2'



PDS_COLS = {
    cTESTE:         'A',
    cCOMENTADO:     'B',
    cCOMENTARIO:    'C',
    cVAO:           'D',
    cIED:           'E',
    cORIGEM:        'F',
    cINCLUDE:       'G',
    cTAC:           'H',
    cID:            'I',
    cNOME:          'J',
    cENDERECO:      'K',
    cOCR:           'L',
    cTELA:          'M',
    cALARME:        'N',
    cSOE:           'O',
    cCALCULO:       'P',
    cFILTRO:        'Q',
    cKCONV:         'R',
    cNV2:           'S',
    cORDEM:         'T',
    cDESC1:         'U',
    cALINT:         'V',
    cATLZINV:       'W',
    cCDINIC:        'X',
    cIDICCP:        'Y',
    cINVRT:         'Z',
    cOBSRV:         'AA',
    cSELSD:         'AB',
    cSINCR_MAN:     'AC',
    cSTINI:         'AD',
    cSTNOR:         'AE',
    cUAPL:          'AF'
}

PAS_COLS = {
    cTESTE:         'A',
    cCOMENTADO:     'B',
    cCOMENTARIO:    'C',
    cVAO:           'D',
    cIED:           'E',
    cORIGEM:        'F',
    cINCLUDE:       'G',
    cTAC:           'H',
    cID:            'I',
    cNOME:          'J',
    cENDERECO:      'K',
    cOCR:           'L',
    cTELA:          'M',
    cALARME:        'N',
    cSOE:           'O',
    cCALCULO:       'P',
    cFILTRO:        'Q',
    cKCONV1:        'R',
    cKCONV2:        'S',
    cKCONV3:        'T',
    cNV2:           'U',
    cORDEM:         'V',
    cDESC1:         'W',
    cALINT:         'X',
    cBNDMO:         'Y',
    cATLZINV:       'Z',
    cBDTR:          'AA',
    cCDINIC:        'AB',
    cHISTPER:       'AC',
    cHISTSLC:       'AD',
    cHTRIS:         'AE',
    cIDICCP:        'AF',
    cOBSRV:         'AG',
    cLIA:           'AH',
    cLIALE:         'AI',
    cLIAMA:         'AJ',
    cLIAME:         'AK',
    cLIAMI:         'AL',
    cLIAPE:         'AM',
    cLIE:           'AN',
    cLIU:           'AO',
    cLIULE:         'AP',
    cLIUMA:         'AQ',
    cLIUME:         'AR',
    cLIUMI:         'AS',
    cLIUPE:         'AT',
    cLSA:           'AU',
    cLSALE:         'AV',
    cLSAMA:         'AW',
    cLSAME:         'AX',
    cLSAMI:         'AY',
    cLSAPE:         'AZ',
    cLSE:           'BA',
    cLSU:           'BB',
    cLSULE:         'BC',
    cLSUMA:         'BD',
    cLSUME:         'BE',
    cLSUMI:         'BF',
    cLSUPE:         'BG',
    cPTC:           'BH',
    cSELSD:         'BI',
    cSINCR_MAN:     'BJ',
    cTEND:          'BK',
    cTXVAR:         'BL',
    cUAPL:          'BM',
    cVLINIC:        'BN'

}


PTS_COLS = {
    cTESTE:         'A',
    cCOMENTADO:     'B',
    cCOMENTARIO:    'C',
    cVAO:           'D',
    cIED:           'E',
    cORIGEM:        'F',
    cINCLUDE:       'G',
    cTAC:           'H',
    cID:            'I',
    cNOME:          'J',
    cENDERECO:      'K',
    cOCR:           'L',
    cTELA:          'M',
    cALARME:        'N',
    cCALCULO:       'O',
    cFILTRO:        'P',
    cKCONV1:        'Q',
    cKCONV2:        'R',
    cKCONV3:        'S',
    cNV2:           'T',
    cORDEM:         'U',
    cDESC1:         'V',
    cALINT:         'W',
    cATLZINV:       'X',
    cCDINIC:        'Y',
    cHISTPER:       'Z',
    cHISTSLC:       'AA',
    cIDICCP:        'AB',
    cOBSRV:         'AC',
    cLSA:           'AD',
    cLSE:           'AE',
    cLSU:           'AF',
    cSELSD:         'AG',
    cSINCR_MAN:     'AH',
    cTXVAR:         'AI',
    cUAPL:          'AJ',
    cVLINIC:        'AK'

}

CGS_COLS = {
    cTESTE:         'A',
    cCOMENTADO:     'B',
    cCOMENTARIO:    'C',
    cVAO:           'D',
    cIED:           'E',
    cORIGEM:        'F',
    cINCLUDE:       'G',
    cTAC:           'H',
    cID:            'I',
    cNOME:          'J',
    cENDERECO:      'K',
    cCONTROLE:      'L',
    cSUPERVISAO:    'M',
    cTIPO:          'N',
    cINTERTRAV:     'O',
    cTELA:          'P',
    cKCONV:         'Q',
    cOPCOES:        'R',
    cNV2:           'S',
    cORDEM:         'T',
    cCNF:           'U',
    cDESC1:         'V',
    cIDICCP:        'W',
    cINVCT:         'X',
    cLMI1C:         'Y',
    cLMI2C:         'Z',
    cLMS1C:         'AA',
    cLMS2C:         'AB',
    cOBSRV:         'AC',
    cRSULT:         'AD',
    cTRRAC:         'AE',
    cTPCTL:         'AF'
}

CALC_COLS = {
    cTESTE:         'A',
    cCOMENTADO:     'B',
    cCOMENTARIO:    'C',
    cVAO:           'D',
    cINCLUDE:       'E',
    cPNT:           'F',
    cTPPNT:         'G',
    cORDEM:         'H',
    cPARC:          'I',
    cTPPARC:        'J',
    cTIPOP:         'K'
}

FILC_COLS = {
    cTESTE:         'A',
    cCOMENTADO:     'B',
    cCOMENTARIO:    'C',
    cVAO:           'D',
    cINCLUDE:       'E',
    cPNT:           'F',
    cTPPNT:         'G',
    cORDEM:         'H',
    cPARC:          'I',
    cTPPARC:        'J'
}

FILS_COLS = {
    cTESTE:         'A',
    cCOMENTADO:     'B',
    cCOMENTARIO:    'C',
    cVAO:           'D',
    cIED:           'E',
    cORIGEM:        'F',
    cINCLUDE:       'G',
    cORDEM:         'H',
    cPNT:           'I',
    cTIPOP:         'J',
    cPDF_PNT:       'K',
    cPDF_TPPNT:     'L',
    cPDF_KCONV:     'M',
    cPDF_NV2:       'N',
    cPDF_ORDEM:     'O',
    cPDF_DESC1:     'P',
    cPDF_DESC2:     'Q',
    cPTF_PNT:       'R',
    cPTF_TPPNT:     'S',
    cPTF_KCONV1:     'T',
    cPTF_KCONV2:     'U',
    cPTF_KCONV3:     'V',
    cPTF_NV2:       'W',
    cPTF_ORDEM:     'X',
    cPTF_DESC1:     'Y',
    cPTF_DESC2:     'Z',
    cPAF_PNT:       'AA',
    cPAF_TPPNT:     'AB',
    cPAF_KCONV1:     'AC',
    cPAF_KCONV2:     'AD',
    cPAF_KCONV3:     'AE',
    cPAF_NV2:       'AF',
    cPAF_ORDEM:     'AG',
    cPAF_DESC1:     'AH',
    cPAF_DESC2:     'AI',

}

def check_wb(wb, **kwargs):
    if set(wb.get_sheet_names()) != set(SHEETS):
        print_msg(__name__,'Arquivo possui erro de formata‡„o, planilhas n„o correspondem ao esperado - {0}'.format(SHEETS), **kwargs)
        return False
    wsPDS = wb['PDS']
    wsPAS = wb['PAS']
    wsPTS = wb['PTS']
    wsCGS = wb['CGS']
    wsCALC = wb['CALC']
    wsFILC = wb['FILC']
    wsFILS = wb['FILS']

    wsPDS_cols = [c.value for c in wsPDS['2']]
    wsPAS_cols = [c.value for c in wsPAS['2']]
    wsPTS_cols = [c.value for c in wsPTS['2']]
    wsCGS_cols = [c.value for c in wsCGS['2']]
    wsCALC_cols = [c.value for c in wsCALC['1']]
    wsFILC_cols = [c.value for c in wsFILC['1']]
    wsFILS_cols = [c.value for c in wsFILS['2']]

    if not set(PDS_COLS.keys()).issubset(set(wsPDS_cols)):
        print_msg(__name__,'Campos de PDS n„o correspondem ao esperado - {0}'.format(list(PDS_COLS.keys())),**kwargs)
        return False
    if not set(PAS_COLS.keys()).issubset(set(wsPAS_cols)):
        print_msg(__name__,'Campos de PAS n„o correspondem ao esperado - {0}'.format(list(PAS_COLS.keys())),**kwargs)
        return False
    if not set(PTS_COLS.keys()).issubset(set(wsPTS_cols)):
        print_msg(__name__,'Campos de PTS n„o correspondem ao esperado - {0}'.format(list(PTS_COLS.keys())),**kwargs)
        return False
    if not set(CGS_COLS.keys()).issubset(set(wsCGS_cols)):
        print_msg(__name__,'Campos de CGS n„o correspondem ao esperado - {0}'.format(list(CGS_COLS.keys())),**kwargs)
        return False
    if not set(CALC_COLS.keys()).issubset(set(wsCALC_cols)):
        print_msg(__name__,'Campos de CALC n„o correspondem ao esperado - {0}'.format(list(CALC_COLS.keys())),**kwargs)
        return False
    if not set(FILC_COLS.keys()).issubset(set(wsFILC_cols)):
        print_msg(__name__,'Campos de FILC n„o correspondem ao esperado - {0}'.format(list(FILC_COLS.keys())),**kwargs)
        return False
    if not set(FILS_COLS.keys()).issubset(set(wsFILS_cols)):
        print_msg(__name__,'Campos de FILS n„o correspondem ao esperado - {0}'.format(list(FILS_COLS.keys())),**kwargs)
        return False
    else:
        print_msg(__name__,'Formato do XLS v lido', msg_type=MSG_INFO, **kwargs)
        return True

def clear_wb(wb):
    for ws in wb:
        if ws.title !='CONF':
            clear_sheet(ws)


def clear_sheet(ws):
    if ws.title in ['PDS','PAS','PTS','CGS','FILS']:
        ini_row = 3
    elif ws.title in ['CALC', 'FILC']:
        ini_row = 2
    else:
        ini_row = 1
    for row in range(ini_row, ws.max_row+1):
        for col in range(1, ws.max_column+1):
            ws.cell(row=row, column=col).value=None


def color_wb(wb):
    for ws in wb:
        if ws.title != 'CONF':
            color_sheet(ws)


def color_sheet(ws):
    if ws.title in ['PDS','PAS','PTS','CGS','FILS']:
        ini_row = 3
    elif ws.title in ['CALC', 'FILC']:
        ini_row = 2
    else:
        ini_row = 1
    testedFill = PatternFill(start_color=colors.COLOR_INDEX[5],
                   end_color=colors.COLOR_INDEX[5],
                   fill_type='solid')
    issueFill = PatternFill(start_color=colors.COLOR_INDEX[2],
                   end_color=colors.COLOR_INDEX[2],
                   fill_type='solid')
    for row in range(ini_row, ws.max_row+1):
        teste = ws.cell(row=row,column=1).value
        for col in range(1, ws.max_column+1):
            if teste == 'C':
                ws.cell(row=row, column=col).fill = testedFill
            elif teste == 'P':
                ws.cell(row=row, column=col).fill = issueFill



def expand_address(dat_type, aconf):
    address = aconf[dat_type]['items'][0].get('ID',None)
    if address == None:
        return ''
    else:
        try:
            prefix, address = str(address).split('-',maxsplit=1)
        except:
            return False
        address = aconf.get('nv1',{}).get('items',[{}])[0].get('CONFIG',prefix)+'/'+address
        return address

def zip_address(dat_type, base_item, address):
    if dat_type in ['pds','pts']:
        fc = 'ST'
    elif dat_type == 'pas':
        fc = 'MX'
    elif dat_type == 'cgs':
        fc = 'CO'
    if '/' in address:
        # UCD1AY/UD1/USER1/SPS62
        prefix, rest = str(address).split('/',1)
        # prefix=UCD1AY rest=UD1/USER1/SPS62
        if '/' in rest:
            # rest=UD1/USER1/SPS62
            ld, rest2 = str(rest).split('/',1)
            # ld=UD1 rest2=USER1/SPS62
            nv1_config = prefix+ld
            if '/' in rest2:
                # rest2=USER1/SPS62
                ln, do = str(rest2).split('/',1)
                # ln=USER1, do=SPS62
                do = str(do).replace('/','$')
                address = ln + '$' + fc + '$' + do
        else:
            nv1_config = prefix
            address = rest
        location, nv1 = get_item(dat_type='nv1', generic_set=base_item['nv1'], where={'CONFIG':'=='+nv1_config})
        address = nv1.get('ID',nv1_config)+'-' + address

    return address

def get_cell_string(cell):
    if cell.value is not None:
        return cell.value
    else:
        return ''


def printProgress (iteration, total, prefix = '', suffix = '', decimals = 1, barLength = 100):
    """
    Call in a loop to create terminal progress bar
    @params:
        iteration   - Required  : current iteration (Int)
        total       - Required  : total iterations (Int)
        prefix      - Optional  : prefix string (Str)
        suffix      - Optional  : suffix string (Str)
        decimals    - Optional  : positive number of decimals in percent complete (Int)
        barLength   - Optional  : character length of bar (Int)
    """
    formatStr       = "{0:." + str(decimals) + "f}"
    percents        = formatStr.format(100 * (iteration / float(total)))
    filledLength    = int(round(barLength * iteration / float(total)))
    bar             = 'Û' * filledLength + '-' * (barLength - filledLength)
    sys.stdout.write('\r%s |%s| %s%s %s' % (prefix, bar, percents, '%', suffix)),
    if iteration == total:
        sys.stdout.write('\n')
    sys.stdout.flush()


