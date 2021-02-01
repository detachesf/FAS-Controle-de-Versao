# -*- coding: cp860 -*-

import pickle  # serve para armazenar objetos e vari†veis em arquivos
import tkinter
from tkinter import Tk
from tkinter.messagebox import showerror
from tkinter import ttk
from tkinter.filedialog import askopenfilename, askdirectory
from os import path, startfile, listdir, popen, getcwd
from sys import stdout
from traceback import print_exc
import re

try:
    from xlrd import open_workbook
except:
    showerror('Erro', 'M¢dulo xlrd nÑo instalado')
try:
    from openpyxl import load_workbook
except:
    showerror('Erro', 'M¢dulo openpyxl nÑo instalado')

try:
    from lp_lib.Gerar_LP import gerar
except:
    showerror('Erro', 'M¢dulo Gerar_LP nÑo instalado')

try:
    from lp_lib.Checar_LP import checar
except:
    showerror('Erro', 'M¢dulo Checar_LP nÑo instalado')

try:
    from lp_lib.func import processing
except:
    showerror('Erro', 'M¢dulo func nÑo instalado')

try:
    from lp_lib.func import createToolTip
except:
    showerror('Erro',
              'Arquivo "func.pyc" deve estar no diret¢rio "lp_lib"')


class Janela:
    def __init__(self, toplevel):

        # Armazena caminho dos arquivos utilizados
        try:
            self.caminhoArquivoLP_Padrao = \
                [arq for arq in listdir('.') if arq.find('Padr') > -1 and arq.find('Planilha') > -1][-1]
        except:
            self.caminhoArquivoLP_Padrao = ''
        try:
            self.PlanilhaArquivoLP_Comfig = [arq for arq in listdir('.') if arq.find('Config') > -1][-1]
        except:
            self.PlanilhaArquivoLP_Comfig = ''
        try:
            self.caminhoArquivoLP_Comfig = [arq for arq in listdir('.') if arq.find('Config') > -1][-1]
        except:
            self.caminhoArquivoLP_Comfig = ''

        self.PlanilhaArquivoLPEditado = ''
        self.caminhoArquivoLPEditado = ''
        self.pathchecar = getcwd
        self.versao = '2.0.12'
        self.data = '10/11/2020'

        # Tamanhos de frame
        frmlargura = 330
        frmaltura = 80

        # Menu
        self.menubar = tkinter.Menu(toplevel)
        self.mnArquivo = tkinter.Menu(self.menubar, tearoff=0)
        self.mnArquivo.add_command(label='Abrir pasta do programa', underline=0, command=self.fcExplorer)
        self.mnArquivo.add_command(label='Limpar relat¢rio', underline=0, command=self.fcLimparRelatorio)
        self.mnArquivo.add_separator()
        self.mnArquivo.add_command(label='Sair', underline=0, command=exit)
        self.menubar.add_cascade(label='Arquivo', underline=0, menu=self.mnArquivo)

        self.mnFerramentas = tkinter.Menu(self.menubar, tearoff=0)
        self.mnFerramentas.add_command(label='Comparar Listas de Pontos...', underline=0, command=self.fcComparar)
        self.mnFerramentas.add_command(label='Base SAGE para LP Excel...[Beta]', underline=0, command=self.fcbase2lp)
        self.mnFerramentas.add_command(label='Planilha Cepel para LP Excel...', underline=0, command=self.fccepel2lp)
        self.mnFerramentas.add_command(label='Gerar Planilha ONS...', underline=0, command=self.fcGerarONS)
        self.menubar.add_cascade(label='Ferramenta', underline=0, menu=self.mnFerramentas)

        self.mnAjuda = tkinter.Menu(self.menubar, tearoff=0)
        self.menubar.add_cascade(label='Sobre', underline=1, command=self.sobreClickButton)

        toplevel['menu'] = self.menubar

        # Frames principais
        self.frmE = ttk.Frame()  # Criar frame esquerdo
        self.frmE.pack(side=tkinter.LEFT)
        self.frmD = ttk.Frame()  # Criar frame direito
        self.frmD.pack(side=tkinter.RIGHT)

        # FRAME "Arquivo LP Padrao"-------------------------------------------------------------------
        self.frm11 = ttk.LabelFrame(self.frmE, text='Arquivo LP PadrÑo', height=int(frmaltura),
                                    width=int(frmlargura))
        self.frm11.grid(row=1, column=1, padx=3, pady=3)
        self.frm11.grid_propagate(0)

        self.nomeArquivoDOMO = ttk.Label(self.frm11, text=self.caminhoArquivoLP_Padrao)
        self.nomeArquivoDOMO.grid(row=1, column=1)

        self.botaoEscolheArquivoDOMO = ttk.Button(self.frm11, text='Selecionar',  # bg='#E0E0E0',
                                                  width=int(frmlargura / 6.5),
                                                  command=self.btArqDOMOClick)
        self.botaoEscolheArquivoDOMO.grid(row=2, column=1, sticky=tkinter.N + tkinter.S + tkinter.W, pady=5, padx=10)

        # FRAME "Arquivo LP_Config"-------------------------------------------------------------------
        self.frm21 = ttk.LabelFrame(self.frmE, text='Arquivo de ParametrizaáÑo', height=frmaltura, width=frmlargura)
        self.frm21.grid(row=2, column=1, padx=3, pady=3)
        self.frm21.grid_propagate(0)

        self.nomeArquivoLP_Comfig = ttk.Label(self.frm21, text=self.PlanilhaArquivoLP_Comfig)
        self.nomeArquivoLP_Comfig.grid(row=1, column=1, columnspan=2)
        self.nomeArquivoLP_Comfig.grid_propagate(0)

        self.botaoEscolheLPConfig = ttk.Button(self.frm21, text='Selecionar',  # bg='#E0E0E0',
                                               width=int(frmlargura / 14),
                                               command=self.botaoEscolheArquivoLPConfigClick)
        self.botaoEscolheLPConfig.grid(row=2, column=1, sticky=tkinter.E, pady=2, padx=8)
        self.botaoEscolheLPConfig.grid_propagate(0)

        self.botaoEditarLPConfig = ttk.Button(self.frm21, text='Editar',  # bg='#E0E0E0',
                                              width=int(frmlargura / 14),
                                              command=self.btEditarArqLPConfigClick)
        self.botaoEditarLPConfig.grid(row=2, column=2, pady=2, padx=8)
        self.botaoEditarLPConfig.grid_propagate(0)

        # FRAME "Arquivo LP Editado"------------------------------------------------------------------
        self.frm31 = ttk.LabelFrame(self.frmE, text='Arquivo LP a ser checado', height=1.5 * frmaltura,
                                    width=frmlargura)
        self.frm31.grid(row=3, column=1, padx=3, pady=3)

        self.nomeArquivoLPEditado = ttk.Label(self.frm31, text='Defina o arquivo...')
        self.nomeArquivoLPEditado.grid(row=1, column=1)

        self.botaoEscolheLPEditado = ttk.Button(self.frm31, text='Selecionar',  # bg='#E0E0E0',
                                                width=int(frmlargura / 6.5),
                                                command=self.botaoEscolheArquivoLPEditadoClick)
        self.botaoEscolheLPEditado.grid(row=2, column=1, sticky=tkinter.N + tkinter.E + tkinter.S + tkinter.W, pady=5,
                                        padx=10)

        self.comboplan = ttk.Combobox(self.frm31)
        self.comboplan.grid(row=3, column=1, sticky=tkinter.N + tkinter.E + tkinter.S + tkinter.W, pady=10, padx=10)

        # FRAME Botoes ------------------------------------------------------------------------------------
        self.frm41 = ttk.Frame(self.frmE, height=frmaltura, width=frmlargura)
        self.frm41.grid(row=4, column=1, padx=3, pady=3)

        self.botaoGerar = ttk.Button(self.frm41, text='\nGerar\n',  # bg='#E0E0E0',
                                     width=int(frmlargura / 23),
                                     # height=2,
                                     command=self.gerarClickButton)
        self.botaoGerar.grid(row=1, column=1, sticky=tkinter.W, pady=20, padx=5)

        self.botaoChecar = ttk.Button(self.frm41, text='\nChecar\n',
                                      # bg='#E0E0E0',
                                      width=int(frmlargura / 23),
                                      # height=2,
                                      state=tkinter.DISABLED, command=self.checarClickButton)
        self.botaoChecar.grid(row=1, column=2, sticky=tkinter.E, pady=20, padx=5)

        self.botaoArquivo = ttk.Button(self.frm41, text='\nArquivo Gerado\n',
                                       # bg='#E0E0E0',
                                       width=int(frmlargura / 23),
                                       # height=2,
                                       command=self.arquivoClickButton)
        self.botaoArquivo.grid(row=1, column=3, pady=20, padx=5)
        # createToolTip(self.botaoArquivo, 'ñltimo arquivo gerado de nova LP ou de checagem.')

        # FRAME Listbox ------------------------------------------------------------------------------------
        self.frm12 = ttk.LabelFrame(self.frmD, text=u" Relat¢rio GeraáÑo  ", height=4.7 * frmaltura,
                                    width=frmlargura)
        self.frm12.grid(row=1, column=1, columnspan=2, padx=3, pady=10)
        self.frm12.grid_propagate(0)

        self.Lb = tkinter.Listbox(self.frm12, width=int(2 * frmlargura / 13), height=int(2 * frmaltura / 7))
        self.Lb.grid(row=0, column=0, sticky=tkinter.N + tkinter.S)
        self.Lb.grid_propagate(0)

        self.scrollY = tkinter.Scrollbar(self.frm12, orient=tkinter.VERTICAL, command=self.Lb.yview)
        self.scrollY.grid(row=0, column=1, sticky=tkinter.N + tkinter.S)

    def btArqDOMOClick(self):
        temp = askopenfilename(
            filetypes=[('Arquivo do Excel', 'xls'), ('Arquivo do Excel', 'xlsx'), ('Arquivo do Excel', 'xlsm')])
        if temp:
            self.caminhoArquivoLP_Padrao = temp
            self.nomeArquivoDOMO['text'] = path.basename(temp)

    def btEditarArqLPConfigClick(self):
        #startfile(self.caminhoArquivoLP_Comfig)
        LPForm = Tk()
        LPForm.title('Formul†rio de ConfiguraáÑo')
        LPForm.geometry('500x500')
        self.Frame1 = ttk.Frame(LPForm)
        self.Frame2 = ttk.Frame(LPForm)
        self.Frame1.pack()
        self.Frame2.pack()

        barmenu = tkinter.Menu(self.Frame1)
        filemenu= tkinter.Menu(barmenu, tearoff=0)
        filemenu.add_command(label="Novo")
        filemenu.add_command(label="Abrir")
        filemenu.add_command(label="Salvar")
        filemenu.add_command(label="Salvar Como...")
        filemenu.add_command(label="Fechar",command=LPForm.destroy)
        filemenu.add_separator()

        filemenu.add_command(label="Sair", command=LPForm.quit)
        barmenu.add_cascade(label="Arquivo", menu=filemenu)

        helpmenu = tkinter.Menu(barmenu, tearoff=0)
        helpmenu.add_command(label="Sobre", command='donothing')
        barmenu.add_cascade(label="Ajuda", menu=helpmenu)

        LPForm.config(menu=barmenu)

        self.mnArquivo = tkinter.Menu(self.menubar, tearoff=0)
        self.mnArquivo.add_command(label='Abrir pasta do programa', underline=0, command=self.fcExplorer)
        self.mnArquivo.add_command(label='Limpar relat¢rio', underline=0, command=self.fcLimparRelatorio)
        self.mnArquivo.add_separator()
        self.mnArquivo.add_command(label='Sair', underline=0, command=exit)
        self.menubar.add_cascade(label='Arquivo', underline=0, menu=self.mnArquivo)

        nb= ttk.Notebook(self.Frame2) #Cria o multipage
        nb.grid(row=1, column=0, columnspan=500, rowspan=490, sticky='NESW')

        page1 = ttk.Frame(nb)
        nb.add(page1, text='LT') #Cria a aba do multipage

        FrmLt1 = ttk.Frame(page1,relief='raised')
        FrmLt1.grid(row=0, column=0)
        Label1= tkinter.Label(FrmLt1, text="Linha de TransmissÑo 1")
        Label1.grid(row=0, column=0, padx=2, pady=2, columnspan=2, sticky="W")
        R1 = tkinter.Radiobutton(FrmLt1, text="500 kV", variable='IntVar', value=1,
                                 command='sel')
        R1.grid(row=1, column=0, padx=10, pady=10)

        R2 = tkinter.Radiobutton(FrmLt1, text="230 kV", variable='IntVar', value=2,
                                 command='sel')
        R2.grid(row=1, column=1, padx=10, pady=10)
        R3 = tkinter.Radiobutton(FrmLt1, text="138 kV", variable='IntVar', value=3,
                                 command='sel')
        R3.grid(row=1, column=2, padx=10, pady=10)

        FrmLt2 = ttk.Frame(page1)
        FrmLt2.grid(row=0, column=1)

        pag2= ttk.Frame(nb)
        nb.add(pag2,  text='PAINEL SAGE E BASTIDOR DE REDE')

        page3 = ttk.Frame(nb)
        nb.add(page3, text='TRAFO')  # Cria a aba do multipage

        pag4 = ttk.Frame(nb)
        nb.add(pag4, text='VéO DE TRANSF./DISJ. CENTRAL')

        page5 = ttk.Frame(nb)
        nb.add(page5, text='REATOR')  # Cria a aba do multipage

        pag6 = ttk.Frame(nb)
        nb.add(pag6, text='TRAFO TERRA')

        page7 = ttk.Frame(nb)
        nb.add(page7, text='PAINEL DE PROTEÄéO DE BARRAS')  # Cria a aba do multipage

        pag8 = ttk.Frame(nb)
        nb.add(pag8, text='ACESSO - VéO SEGREGADO')

        page9 = ttk.Frame(nb)
        nb.add(page9, text='BANCO CAPACITOR SHUNT')  # Cria a aba do multipage

        pag10 = ttk.Frame(nb)
        nb.add(pag10, text='BANCO CAPACITOR SêRIE')

        page11 = ttk.Frame(nb)
        nb.add(page11, text='ECE')  # Cria a aba do multipage

        pag12 = ttk.Frame(nb)
        nb.add(pag12, text='SISTEMA REGULAÄéO')

        page13 = ttk.Frame(nb)
        nb.add(page13, text='PREP.REEN.')  # Cria a aba do multipage

        pag14 = ttk.Frame(nb)
        nb.add(pag14, text='COMPENSADOR SãNCRONO')

        pag15 = ttk.Frame(nb)
        nb.add(pag15, text='SERVIÄOS AUXILIARES')

    def botaoEscolheArquivoLPConfigClick(self):
        temp = askopenfilename(
            filetypes=[('Arquivo do Excel', 'xls'), ('Arquivo do Excel', 'xlsx'), ('Arquivo do Excel', 'xlsm')],
            initialdir=self.pathchecar)
        if temp:
            self.pathchecar = path.dirname(temp)
            self.caminhoArquivoLP_Comfig = temp
            self.nomeArquivoLP_Comfig['text'] = path.basename(temp)

    def botaoEscolheArquivoLPEditadoClick(self):
        temp = askopenfilename(
            filetypes=[('Arquivo do Excel', 'xls'), ('Arquivo do Excel', 'xlsx'), ('Arquivo do Excel', 'xlsm')],
            initialdir=self.pathchecar)
        if temp:
            self.caminhoArquivoLPEditado = temp
            self.nomeArquivoLPEditado['text'] = path.basename(temp)
            try:
                book = load_workbook(temp)  # Abrir arquivo de a ser checado
            except:
                aviso = 'Arquivo \"' + temp + '\" nÑo encontrado'
                showerror('Erro', aviso)
            array_combo = []
            for nome_aba in book.sheetnames:
                array_combo.append(nome_aba)
            self.comboplan['values'] = tuple(array_combo)
            self.comboplan.current(0)
            self.botaoChecar.config(state=tkinter.NORMAL)

    def gerarClickButton(self):
        self.Lb.delete(0, tkinter.END)
        try:
            arq_conf = open_workbook(self.caminhoArquivoLP_Comfig)  # Abrir arquivo de LP_Config
        except:
            print_exc(file=stdout)
            aviso = 'Arquivo \"' + self.caminhoArquivoLP_Comfig + u'\" nÑo encontrado'
            showerror('Erro', aviso)
        try:
            sheet = arq_conf.sheet_by_index(0) # Abrir planilha "Configuraáîes" do arquivo LP_config.xls
            vers = re.findall('\d+\.\d+\.\d+', str(sheet.cell(110, 0)))[0].split('.')  # Ler definiáÑo do c¢digo da SE
            vers = list(map(int, vers))  # Transformar array de string em array de inteiro
            if vers < [2, 0, 13]:
                showerror('Erro', 'Deve ser usado arquivo LP_Config.xls com versÑo igual ou maior a 2.0.13')
            else:

                try:
                    processing(gerar, {'LP_Padrao': self.caminhoArquivoLP_Padrao, 'relatorio': self.Lb,
                                       'LP_Config': self.caminhoArquivoLP_Comfig})
                except:
                    print_exc(file=stdout)
                    showerror('Erro', 'Erro inesperado ao tentar gerar lista de pontos.')

        except:
            showerror('Erro', 'Arquivo indicado nÑo corresponde a arquivo de parametrizaáÑo v†lido')

    def checarClickButton(self):

        self.PlanilhaArquivoLPEditado = self.comboplan.get()
        self.Lb.delete(0, tkinter.END)
        try:
            arq_conf = open_workbook(self.caminhoArquivoLP_Comfig)  # Abrir arquivo de LP_Config
        except:
            aviso = 'Arquivo \"' + self.caminhoArquivoLP_Comfig + u'\" nÑo encontrado'
            showerror('Erro', aviso)
        try:
            sheet = arq_conf.sheet_by_index(0)  # Abrir planilha "Configuraáîes" do arquivo LP_config.xls
            vers = re.findall('\d+\.\d+\.\d+', str(sheet.cell(110, 0)))[0].split('.')  # Ler definiáÑo do c¢digo da SE
            vers = list(map(int, vers))  # Transformar array de string em array de inteiro
            if vers < [2, 0, 13]:
                showerror('Erro', 'Deve ser usado arquivo LP_Config.xls com versÑo igual ou maior a 2.0.13')
            else:
                try:
                    processing(checar,
                               {'LP_Padrao': self.caminhoArquivoLP_Padrao, 'LP_Editado': self.caminhoArquivoLPEditado,
                                'planilha': self.PlanilhaArquivoLPEditado, 'relatorio': self.Lb,
                                'LP_Config': self.caminhoArquivoLP_Comfig})
                except:
                    print_exc(file=stdout)
                    showerror('Erro', 'Erro inesperado ao tentar checar lista de pontos.')
        except:
            showerror('Erro', 'Arquivo indicado nÑo corresponde a arquivo de parametrizaáÑo v†lido')

    def arquivoClickButton(self):
        try:
            conf = pickle.load(open('fas.p', 'r'))
            startfile(conf['arquivo'])
        except:
            showerror('Erro', 'NÑo existe arquivo definido')

    def fcExplorer(self):
        popen('explorer .')

    def fcLimparRelatorio(self):
        self.Lb.delete(0, tkinter.END)

    def fcComparar(self):
        try:
            from lp_lib.LP_Comparar import JanelaComp
        except:
            showerror('Erro', 'M¢dulo LP_Comparar nÑo instalado')
            return 0

        jncomp = tkinter.Toplevel()
        jncomp.title('Comparar Arquivos')  # T°tulo da janela
        try:
            jncomp.iconbitmap(default='lp_lib/chesf.ico')  # ãcone utilizado pela janela
        except:
            pass
        jncomp.resizable(0, 0)
        JanelaComp(jncomp, self.Lb)
        jncomp.mainloop()

    def fcbase2lp(self):
        try:
            from lp_lib.base2lp import base2lp
        except:
            showerror('Erro', 'M¢dulo base2lp nÑo instalado')
            return 0
        diretorio = askdirectory(title='Selecione o diret¢rio que estÑo os arquivos .dat')
        if diretorio:
            try:
                base2lp(diretorio)
            except:
                print_exc(file=stdout)
                showerror('Erro', 'Erro inesperado ao tentar checar lista de pontos.')

    def fccepel2lp(self):
        try:
            from lp_lib.cepel2lp import cepel2lp
        except:
            showerror('Erro', 'M¢dulo cepel2lp nÑo instalado')
            return 0
        arqcepel = askopenfilename(
            filetypes=[('Arquivo do Excel', 'xls'), ('Arquivo do Excel', 'xlsx'), ('Arquivo do Excel', 'xlsm')])
        if arqcepel:
            try:
                cepel2lp(arqcepel)
            except:
                print_exc(file=stdout)
                showerror('Erro', 'Erro inesperado ao tentar checar lista de pontos.')

    def fcGerarONS(self):
        try:
            from lp_lib.Gerar_ONS import JanelaGerarONS
        except:
            showerror('Erro', 'M¢dulo Gerar_ONS nÑo instalado')
            return 0

        jngerarons = tkinter.Toplevel()
        jngerarons.title('Gerar Planilha ONS')  # T°tulo da janela
        try:
            jngerarons.iconbitmap(default='lp_lib/chesf.ico')  # ãcone utilizado pela janela
        except:
            pass
        jngerarons.resizable(0, 0)
        JanelaGerarONS(jngerarons, self.Lb)
        jngerarons.mainloop()

    def sobreClickButton(self):
        sobre = tkinter.Tk()
        sobre.title('Sobre')

        texto = \
            '''
        VersÑo ''' + self.versao + '''
        Ferramenta de AutomatizaáÑo para Projetos de Sistemas Supervis¢rios
        Produzido e mantido pelo DETA
        AtualizaáÑo do programa: ''' + self.data

        tkinter.Label(sobre, text='\nF A S',
                      fg='blue', anchor=tkinter.CENTER, font=("Verdana", "14", "bold italic")).grid(row=1, column=1,
                                                                                                    sticky=tkinter.N + tkinter.E + tkinter.S + tkinter.W)
        tkinter.Label(sobre, text=texto,
                      anchor=tkinter.CENTER, font=("Verdana", "8")).grid(row=2, column=1,
                                                                         sticky=tkinter.N + tkinter.E + tkinter.S + tkinter.W,
                                                                         pady=10, padx=30)

        sobre.mainloop(0)


if __name__ == "__main__":
    app = tkinter.Tk()  # InstÉncia do Tk (janela principal)
    app.title('FAS - Ferramenta de AutomatizaáÑo para Projetos de Sistemas Supervis¢rios')  # T°tulo da janela
    try:
        app.iconbitmap(default='lp_lib/chesf.ico')  # ãcone utilizado pela janela
    except:
        pass

    app.resizable(0, 0)
    Janela(app)
    app.mainloop()
