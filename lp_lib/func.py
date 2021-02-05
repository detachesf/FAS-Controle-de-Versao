# -*- coding: cp860 -*-

from tkinter.messagebox import showerror
from tkinter import Toplevel, ttk, E, W, CENTER, LEFT, SOLID, Label
import threading
from time import sleep
import gi
gi.require_version("Gtk", "3.0")
from gi.repository import Gtk, GObject, GLib
GObject.threads_init()
import FASgtkui
from sys import stdout
from traceback import print_exc
import gobject

try:
    from openpyxl import load_workbook,cell
except:
    showerror('Erro', 'M¢dulo openpyxl n„o instalado')

dados= '''            
Vers„o do programa: 2.0.12
Atualiza‡„o do programa: 16/10/2020
Fun‡”es adicionais: painelLT69, linhaInicialETitulos
'''


def painelLT69(ListaPontos):
    array_ID = [col[0] for col in ListaPontos[0]] # Separar ID SAGE da lista de pontos (LP em 0 e contadores k nas outras posi‡”es)
    array_checar = [tag[-11:] for tag in array_ID] # Separar as £ltimas 11 posi‡”es (Painel e c¢digo de ponto)
    pos11dupl = [dupl11 for dupl11 in set(array_checar) if array_checar.count(dupl11)>1 and dupl11[:3]=='2UA' ] # Separar as £ltimas 11 posi‡”es duplicadas come‡adas com 2UA (referente a painel de 69kV)
    array_ID_dupl = [dupl[0] for dupl in ListaPontos[0] if dupl[0][-11:] in pos11dupl] #Separa array com IDs duplicados
    
    tratar_par = []
    for i11 in pos11dupl: #Passar array de 11 duplicados
        par = []
        for idupl in array_ID_dupl: #Passar array de IDs a serem tratados
            if idupl[-11:] == i11:
                par.append(idupl) #Separar IDs duplicados em pares para tratamento
        if len(par)>0: tratar_par.append(par) #Gravar par em array para tratamento
    for tratar in tratar_par:
        novoID = tratar[0][0:8]+'/'+tratar[1][6:8]+tratar[0][8:]
        pos0 = array_ID.index(tratar[0]) #Pegar posi‡„o do primero ID a ser Tratado
        ListaPontos[0].pop(pos0) #Excluir ponto referente ao primeiro ID tratar[0]
        array_ID.pop(pos0) #Excluir ponto referente ao primeiro ID tratar[0] para os indices da ListaPontos e array_ID continuem coerentes
        
        pos1 = array_ID.index(tratar[1]) #Pegar posi‡„o do segundo ID a ser Tratado
        ListaPontos[0][pos1][0] = novoID #Substituir ID do ponto referente ao segundo ID tratar[1]
        
        ListaPontos[1][0] -= 1 #Modificar k_lt
    
    return ListaPontos

def linhaInicialETitulos(arquivo, nomeAba):
    """
    Rotina para encontrar a primeira linha v lida de uma lista de pontos e t¡tulos do cabe‡alho da Lista de Pontos
    @param arquivo: Arquivo Excel com a Lista de Pontos
    @param nomeAba: Nome da aba que est  a Lista de Pontos    
    @return: [linhaInicial, {dicion rio de t¡tulos}] - Retorna array com a linha inicial (primeira linha ‚ a  0 - zero), 
    na primeira posi‡„o e dicion rio com t¡tulos do cabe‡alho na segunda posi‡„o. Caso n„o seja encontrada referˆncia ao 
    'D (SAGE)' o valor de retorno de linhaInicial ser  -1 (menos um).  
    """   
    
    arq_conf = load_workbook(arquivo, data_only=True)
    sheet = arq_conf[nomeAba]
    titulossuperiores = {}
    TitulosPrinc = ['CHESF - N‹VEL 2', 'CHESF - TELEASSIST‰NCIA N3', 'CHESF - N‹VEL 3', 'ONS', 'LIMITES OPERACIONAIS']
    titulos = {}
    for li in range(2, 10):                                         #Varrer as linhas de 2 a 10
        for i in range(sheet.max_column):
            texto = str(sheet.cell(row=li, column=i+1).value).upper().strip().replace("  "," ")
            if texto in TitulosPrinc:
                titulossuperiores[texto] = (li, i+1) #Pega a linha e coluna associada ao t¡tulo principal
        if 'CHESF - N‹VEL 2' in titulossuperiores: break         #Se foi passado pela linha com chave "CHESF - N‹VEL 2" parar de varrer linhas
    #Tratamento dos t¡tulos principais, se n„o estiver da mesma forma da aba PADRAO da lista de pontos, a mensagem de erro ‚ gerada
    for tit in TitulosPrinc:
        if tit not in titulossuperiores:
            showerror('Erro','T¡tulo {} n„o identificado no arquivo a ser checado, verifique se ele est  escrito desta mesma forma.'.format(tit))

    subtitulosn2 = {}
    subtitulosn3Tele = {}
    subtitulosn3 = {}
    subtitulosONS = {}
    subtitulosLO = {}
# sess„o onde os subt¡tulos s„o captados da lista a ser checada
    for coluna in range( titulossuperiores['CHESF - N‹VEL 2'][1], titulossuperiores['CHESF - TELEASSIST‰NCIA N3'][1]):
        for i in range(1,3):
            if str(sheet.cell(row = titulossuperiores['CHESF - N‹VEL 2'][0] + i, column = coluna).value).upper().strip() != '':
                subtitulosn2[str(sheet.cell(row = titulossuperiores['CHESF - N‹VEL 2'][0] + i, column = coluna).value).upper().strip()] = coluna
                if coluna == titulossuperiores['CHESF - N‹VEL 2'][1]:
                    li += 1 #guarda a £ltima linha de subt¡tulo preenchida
                break
    for coluna in range(titulossuperiores['CHESF - TELEASSIST‰NCIA N3'][1], titulossuperiores['CHESF - N‹VEL 3'][1]):
        for i in range(1, 3):
            if str(sheet.cell(row=titulossuperiores['CHESF - TELEASSIST‰NCIA N3'][0] + i, column=coluna).value).upper().strip() != '':
                subtitulosn3Tele[str(sheet.cell(row=titulossuperiores['CHESF - TELEASSIST‰NCIA N3'][0] + i, column=coluna).value).upper().strip()] = coluna
                if titulossuperiores['CHESF - TELEASSIST‰NCIA N3'][0] + i > li:
                    li = titulossuperiores['CHESF - TELEASSIST‰NCIA N3'][0] + i
                break

    for coluna in range(titulossuperiores['CHESF - N‹VEL 3'][1], titulossuperiores['ONS'][1]):
        for i in range(1, 3):
            if str(sheet.cell(row=titulossuperiores['CHESF - N‹VEL 3'][0] + i,column = coluna).value).upper().strip() != '':
                subtitulosn3[str(sheet.cell(titulossuperiores['CHESF - N‹VEL 3'][0] + i, coluna).value).upper().strip()] = coluna
                if titulossuperiores['CHESF - N‹VEL 3'][0] + i > li:
                    li = titulossuperiores['CHESF - N‹VEL 3'][0] + i
                break

    for coluna in range(titulossuperiores['ONS'][1], titulossuperiores['LIMITES OPERACIONAIS'][1]):
        if str(sheet.cell(row = titulossuperiores['ONS'][0] + 2,column = coluna).value).upper().strip() != '':
            subtitulosONS[str(sheet.cell(row = titulossuperiores['ONS'][0] + 2, column = coluna).value).upper().strip()] = coluna
            if titulossuperiores['ONS'][0] + 2 > li:
                li = titulossuperiores['ONS'][0] + 2

    for coluna in range(titulossuperiores['LIMITES OPERACIONAIS'][1], titulossuperiores['LIMITES OPERACIONAIS'][1] + 8):
        for i in range(1, 3):
            if str(sheet.cell(row = titulossuperiores['LIMITES OPERACIONAIS'][0] + i, column = coluna).value).upper().strip() != '':
                subtitulosLO[str(sheet.cell(row = titulossuperiores['LIMITES OPERACIONAIS'][0] + i, column = coluna).value).upper().strip()] = coluna
                if titulossuperiores['LIMITES OPERACIONAIS'][0] + i > li:
                    li = titulossuperiores['LIMITES OPERACIONAIS'][0] + i
                break
    #Nesse bloco s„o adicionados aos t¡tulos principais os campos com suas respectivas colunas
    titulos['CHESF - N‹VEL 2'] = subtitulosn2
    titulos['CHESF - TELEASSIST‰NCIA N3'] = subtitulosn3Tele
    titulos['CHESF - N‹VEL 3'] = subtitulosn3
    titulos['ONS'] = subtitulosONS
    titulos['LIMITES OPERACIONAIS'] = subtitulosLO

    #Campos padr„o de cada t¡tulo, nesse bloco ‚ feito o tratamento de erro, caso o arquivo a ser checado venha com os t¡tulos preenchidos erroneamente
    camposN2 = ['ID (SAGE)','OCR (SAGE)','DESCRI€ŽO','TIPO','COMANDO','MEDI€ŽO','LISTA DE ALARMES','SOE']
    camposN3Tele = ['OCR (SAGE)','COMANDO','MEDI€ŽO','LISTA DE ALARME','SOE','OBSERVA€ŽO','AGRUPAMENTO','ENDERE€O']
    camposN3 = ['OCR (SAGE)','COMANDO','MEDI€ŽO','LISTA DE ALARME','SOE','OBSERVA€ŽO','AGRUPAMENTO','ENDERE€O']
    camposONS = ['ITEM','DESCRI€ŽO']
    camposLimop = ['LIU','LIE','LIA','LSA','LSE','LSU','BNDMO','OBSERVA€™ES']
    if 'TELA' not in titulos['CHESF - N‹VEL 2'] and 'ANUNCIADOR' not in titulos['CHESF - N‹VEL 2']:
        showerror('Erro','Campo TELA ou ANUNCIADOR n„o identificado abaixo do cabe‡alho CHESF - N‹VEL 2 do arquivo a ser checado, verifique se ele est  escrito desta mesma forma.')
    for campo in camposN2:
        if campo not in titulos['CHESF - N‹VEL 2']:
            showerror('Erro','Campo {} n„o identificado abaixo do cabe‡alho CHESF - N‹VEL 2 do arquivo a ser checado, verifique se ele est  escrito desta mesma forma.'.format(campo))
    for campo in camposN3Tele:
        if campo not in titulos['CHESF - TELEASSIST‰NCIA N3']:
            showerror('Erro', 'Campo {} n„o identificado abaixo do cabe‡alho CHESF - TELEASSIST‰NCIA N3 do arquivo a ser checado, verifique se ele est  escrito desta mesma forma.'.format(campo))
    for campo in camposN3:
        if campo not in titulos['CHESF - N‹VEL 3']:
            showerror('Erro','Campo {} n„o identificado abaixo do cabe‡alho CHESF - N‹VEL 3 do arquivo a ser checado, verifique se ele est  escrito desta mesma forma.'.format(campo))
    for campo in camposONS:
        if campo not in titulos['ONS']:
            showerror('Erro','Campo {} n„o identificado abaixo do cabe‡alho ONS do arquivo a ser checado, verifique se ele est  escrito desta mesma forma.'.format(campo))
    for campo in camposLimop:
        if campo not in titulos['LIMITES OPERACIONAIS']:
            showerror('Erro','Campo {} n„o identificado abaixo do cabe‡alho LIMITES OPERACIONAIS do arquivo a ser checado, verifique se ele est  escrito desta mesma forma.'.format(campo))

    li += 1  # Seleciona linha ap¢s o t¡tulo
    if 'ID (SAGE)' in titulos['CHESF - N‹VEL 2']:                               #Verifica se foi encontrado chave "ID (SAGE)"
        while True:
            if sheet.cell(row = li, column=titulos['CHESF - N‹VEL 2']['ID (SAGE)']).value:              #Verifica se a c‚lula est  preenchida com algum valor
                break                                                   #Parar de procurar linha preenchida
            else:
                li += 1                                                 #Selecionar linha seguinte
    else:
        li = -1
    return [li,titulos]

def processing(function, args):

    def check():
        if not thread_f.isAlive():
            janela.hide()
            FASgtkui.Manipulador.on_janela_progressbar_hide(FASgtkui.Manipulador)
            return False
        return True
    janela: Gtk.Window = FASgtkui.builder.get_object('janela_progressbar')
    Spinner: Gtk.Spinner = FASgtkui.builder.get_object('spinner')
    janela.show()
    Spinner.activate()
    thread_f = threading.Thread(target=function, kwargs=args)
    thread_f.start()
    tread = GObject.timeout_add(300, check)


class ToolTip(object):
    def __init__(self, widget):
        self.widget = widget
        self.tipwindow = None
        self.id = None
        self.x = self.y = 0
    def showtip(self, text):
        "Display text in tooltip window"
        self.text = text
        if self.tipwindow or not self.text:
            return
        x, y, _cx, cy = self.widget.bbox("insert")
        x = x + self.widget.winfo_rootx() + 27
        y = y + cy + self.widget.winfo_rooty() +27
        self.tipwindow = tw = Toplevel(self.widget)
        tw.wm_overrideredirect(1)
        tw.wm_geometry("+%d+%d" % (x, y))

        label = Label(tw, text=self.text, justify=LEFT,
                         background="#ffffe0", relief=SOLID, borderwidth=1,
                         font=("tahoma", "8", "normal"))
        label.pack(ipadx=1)
    def hidetip(self):
        tw = self.tipwindow
        self.tipwindow = None
        if tw:
            tw.destroy()

def createToolTip(widget, text):
    toolTip = ToolTip(widget)
    def enter(event):
        sleep(3)
        toolTip.showtip(text)
    def leave(event):
        toolTip.hidetip()
    widget.bind('<Enter>', enter)
    widget.bind('<Leave>', leave)


