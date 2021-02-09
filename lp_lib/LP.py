# -*- coding: cp860 -*-
import re
from FASgtkui import mensagem_erro
from sys import stdout
from traceback import print_exc

try:
    from openpyxl import load_workbook,cell
except:
    mensagem_erro('Erro', 'M¢dulo openpyxl n„o instalado')

try:
    from lp_lib.func import painelLT69
except:
    mensagem_erro('Erro',
              'Arquivo "func.pyc" deve estar no diret¢rio "lp_lib"')
try:
    from bs4 import BeautifulSoup
except:
    mensagem_erro('Erro','M¢dulo BeautifulSoup n„o instalado')

dados = '''
Vers„o do programa: 2.0.13
Atualiza‡„o do programa: 29/01/2021
M¢dulo n£clea da funcionalidade de gerar planilha
'''


def gerarlp(lp_padrao, ArqConf):
    # ----------Declara‡„o de Vari veis----------#
    # Dicion rio que define de Planilha da LP padr„o vai ser lida
    evento_dic = {'LT': False, 'Trafo': False, 'T_Terra': False,
                  'B_CAP': False, 'Disjuntor': False, 'Secc': False,
                  'ECE': False, 'CS': False, 'CE': False, 'BCS': False,
                  'Reator': False, 'Prep. Reen.': False, 'SAs': False,
                  'BARRA': False, 'SD': False, 'SEP': False, 'P_Eolico': False,
                  'P_Solar': False, 'P. Interface': False}

    # Dicion rio para c¢digo de Sistema de Regula‡„o
    tensao_dic = {'230kV': '230', '138kV': '138',
                  '69kV': '069', '13,8kV': '013'}

    # Dicion rio de T¡tulo da tabela LP e index
    titulo_dic = {}

    conf_SD_array = []  # Configura‡„o de Painel SAGE e Bastidores de Rede
    conf_LT_array = []  # Configura‡„o de LT
    conf_Trafo_array = []  # Configura‡„o de TRAFO
    conf_BT_array = []  # Configura‡„o de V„o de Transferˆncia
    conf_TT_array = []  # Configura‡„o de Trafo Terra
    conf_Reator_array = []  # Configura‡„o de Reator
    conf_P87B_array = []  # Configura‡„o de Painel 87 Barra
    conf_BCap_array = []  # Configura‡„o de Banco Capacitor
    conf_BCS_array = []  # Configura‡„o de Banco Capacitor S‚rie
    conf_ECE_array = []  # Configura‡„o de Esquema Especial de Emergˆncia
    conf_CS_array = []  # Configura‡„o de Compensador S¡ncrono
    conf_PR_array = []  # Configura‡„o de Compensador S¡ncrono
    conf_SR_array = []  # Configura‡„o de Sistema Regula‡„o
    conf_PInterface_array = [] #Configura‡„o do Acesso Segregado (P.interface)

    #    conf_CE_array = []          #Configura‡„o de Compensador Est tico

    saida_array = []  # Array que ser  gravado em LP_gerada

    k_lt = 0  # |
    k_trafo = 0  # |
    k_bt = 0  # |
    k_tt = 0  # |
    k_bcap = 0  # |
    k_52 = 0  # |
    k_89 = 0  # |
    k_ece = 0  # |
    k_cs = 0  # |    Contadores de pontos de cada tipo de evento
    #    k_ce = 0      # |
    k_bcs = 0  # |
    k_reator = 0  # |
    k_pr = 0  # |
    k_sr = 0  # |
    k_sas = 0  # |
    k_barra = 0  # |
    k_sd = 0  # |
    k_Pint = 0 # |
    # ---------- Fun‡”es ----------#
    # Pega texto "A , B", coloca em  maipusculo, transforma em Array por v¡rgula e retira espa‡os de cada elemento
    def tratar_str_secc(s):
        return list(map(lambda x: x.strip(), s.upper().split(',')))

    # Fun‡„o para adicionar campos tratar e descri‡„o no saida_array
    def gravar_ponto(campo_tratar, campo_descricao):
        saida_array.append([campo_tratar, sheet.cell(row=index_linha, column=titulo_dic[u'OCR (SAGE)']), campo_descricao] +
                           [sheet.cell(row=index_linha, column=ions) for ions in
                            range(titulo_dic[u'TIPO'], titulo_dic[u'VŽOS DIGITAIS'])] +
                           [sheet.cell(row=index_linha, column=ions) for ions in
                            range(titulo_dic[u'VŽOS DIGITAIS'] + 1, titulo_dic[u'NONE'])])

    # ----------Ler Arquivo de configura‡„o----------#
    try:
        arq_conf = BeautifulSoup(open(ArqConf, 'r', encoding='utf-8'),'html.parser')  # Abrir arquivo de cofigura‡„o
    except:
        mensagem_erro('Erro', u'Arquivo de parametriza‡„o n„o encontrado')

    try:
        Codigo_SE = arq_conf.eventos['codigo_se']    #Ler defini‡„o do c¢digo da SE
    except:
        mensagem_erro('Erro', u'Arquivo indicado n„o corresponde a arquivo de parametriza‡„o v lido')

    index_linha = 9  # Linha 10 do LP_Config.xls, in¡cio de lista de Painel SAGE e Bastidores de Rede


    Eventos =  arq_conf.find_all('paisage') #Pesquisa se tem eventos referentes ao SD
    if Eventos:
        for evento in Eventos:
                               # 0 - Nome do painel Ex. 4UA7A
            conf_SD_array.append({'PNL': str(evento.string),
                                  # 1 - SAGE/BASTIDOR
                                  'SB': str(evento['sagebastidor']),
                                  # 2 - N£mero Inicial de Switch
                                  'DE_SW': int(evento['sw-de']),
                                  # 3 - N£mero Final de Switch
                                  'ATE_SW': int(evento['sw-ate']),
                                  # 4 - N£mero de portas de Switch
                                  'POR_SW': int(evento['nportas-sw']),
                                  # 5 - Firewall (Sim/N„o)
                                  'FW': 'Sim' if str(evento['fw']) == 'True' else 'N„o',
                                  # 6 - N£mero de portas do Firewall
                                  'POR_FW': int(evento['nporta-fw']),
                                  # 7 - RedBox (Sim/N„o)
                                  'RB':'Sim' if str(evento['rb']) == 'True' else 'N„o',
                                  # 8 - N£mero Inicial de RedBox
                                  'DE_RB': int(evento['rb-de']),
                                  # 9 - N£mero Final de RedBox
                                  'ATE_RB': int(evento['rb-ate']),
                                  # 10 - N£mero de portas do RedBox
                                  'POR_RB': int(evento['nporta-rb']),
                                  # A que a parametriza‡„o se refere
                                  'TIPO': 'SD'})
        evento_dic['SD'] = True  # Define que planilha SD da LP padr„o ser  lida

    #in¡cio de lista de Painel SAGE e Bastidores de Rede
    Eventos = arq_conf.rdp_central
    if Eventos:
        conf_RDP = {'DE_RDP': Eventos['rdpde'],
                    'ATE_RDP': Eventos['rdpate'] or 1,
                    # Se n„o foi definido valor, atribuir valor do campo 1
                    'TIPO': 'RDP'}
        evento_dic['SD'] = True  # Define que planilha SD da LP padr„o ser  lida

    # in¡cio de lista de LTs
    Eventos = arq_conf.find_all('lt')
    if Eventos:
        for evento in Eventos:
            # Carregar dados da LT enquanto existir c¢digo da LT na linha Excel
            # 0 - C¢digo operacional LT Ex. 04V1
            conf_LT_array.append({'COD': str(evento.string),
                                  # Nome do painel (sem -1 ou -2) Ex. 4UA2A
                                  'PNL': str(evento['codpainel']),
                                  # Tem 87L (Sim ou N„o)
                                  '87L': 'Sim' if str(evento['87l']) == 'True' else 'N„o',
                                  # Religamento
                                  '79': str(evento['religamento']),
                                  # C¢digo LT Remota Ex. NTT
                                  'LTREM': str(evento['ltremota']),
                                  # Arranjo
                                  'ARR': str(evento['arranjo']),
                                  # RDP Stand Alone
                                  'RDP': 'Sim' if str(evento['rdp'])=='True' else 'N„o',
                                  # Bay Unit do 87B
                                  'F9': 'Sim' if str(evento['f9']) == 'True' else 'N„o',
                                  # Painel Teleprot
                                  '85PNL': 'Sim' if str(evento['painelteleprot']) == 'True' else 'N„o',
                                  # Cƒmaras PASS.
                                  'PASSCam': tratar_str_secc(str(evento['camarapass'])),
                                  # Conjunto de comando das seccionadoras
                                  'PASSSecc': tratar_str_secc(str(evento['conjuntosecc'])),
                                  # A que a parametriza‡„o se refere
                                  'TIPO': 'LT'})
        evento_dic['LT'] = True  # Define que planilha LT da LP padr„o ser  lida
        evento_dic['Disjuntor'] = True  # Define que planilha Disjuntor da LP padr„o ser  lida
        evento_dic['Secc'] = True  # Define que planilha Secc da LP padr„o ser  lida

    Eventos = arq_conf.find_all('trafo')
    if Eventos:
        for evento in Eventos:
            # 0 - C¢digodigo operacional Trafo Ex. 04T1
            conf_Trafo_array.append({'COD': str(evento.string),
                                     # 1 - Nome do painel de Alta Ex. 4UA3A
                                     'PNLH': str(evento['codpainelh']),
                                     # 2 - Nome do painel de Baixa Ex. 2UA3B
                                     'PNLX': str(evento['codpainelx']),
                                     # 3 - Arranjo do setor da alta do Trafo
                                     'ARRH': str(evento['arranjoh']),
                                     # 4 - Arranjo do setor da baixa do Trafo
                                     'ARRX': str(evento['arranjox']),
                                     # 5 - Rela‡”es do Trafo Ex. 230/69/13,8
                                     'REL': str(evento['relacao']),
                                     # 6 - Prote‡„o Ex. PU/PG (Prote‡„o Unit ria/Prote‡„o Gradativa)
                                     'PUPG': str(evento['prot']),
                                     # 7 - Equipamento Ex. Banco Monof sico
                                     'BM': str(evento['equip']),
                                     # RDP Stand Alone
                                     'RDP': 'Sim' if str(evento['rdp']) == 'True' else 'N„o',
                                     # Bay Unit do 87B
                                     'F9': 'Sim' if str(evento['f9']) == 'True' else 'N„o',
                                     # Cƒmaras PASS. Pega texto "A , B", coloca em  maipusculo, transforma em Array por v¡rgula e retira espa‡o de cada elemento
                                     'PASSCam': tratar_str_secc(str(evento['camarapass'])),
                                     # Conjunto de comando das seccionadoras
                                     'PASSSecc': tratar_str_secc(str(evento['conjuntosecc'])),
                                     # Sistema de Regula‡„o via Aplica‡„o (SAGE) (Sim/N„o)...
                                     'REGAPLIC': 'Sim' if str(evento['regapp']) == 'True' else 'N„o',
                                     # A que a parametriza‡„o se refere
                                     'TIPO': 'Trafo'})
        evento_dic['Trafo'] = True  # Define que planilha Trafo da LP padr„o ser  lida
        evento_dic['Disjuntor'] = True  # Define que planilha Disjuntor da LP padr„o ser  lida
        evento_dic['Secc'] = True  # Define que planilha Secc da LP padr„o ser  lida

    #in¡cio de lista de Vao de Transferˆncia
    Eventos = arq_conf.find_all('vaotrans')
    if Eventos:
        for evento in Eventos:
            # Carregar dados Vao de Transferˆncia enquanto existir c¢digo da Vao de Transferˆncia na linha Excel
            # 0 - C¢digo operacional BT Ex. 04D1
            conf_BT_array.append({'COD': str(evento.string),
                                  # 1 - Nome do painel Ex. 4UA7A
                                  'PNL': str(evento['painel']),
                                  # 2 - 87B no painel (Sim/N„o)
                                  '87B': 'Sim' if str(evento['87b']) == 'True' else 'N„o',
                                  # 3 - Arranjo
                                  'ARR': str(evento['arranjo']),
                                  # Cƒmaras PASS. Pega texto "A , B", coloca em  maipusculo, transforma em Array por v¡rgula e retira espa‡o de cada elemento
                                  'PASSCam': tratar_str_secc(str(evento['camarapass'])),
                                  # Conjunto de comando das seccionadoras
                                  'PASSSecc': tratar_str_secc(str(evento['conjuntosecc'])),
                                  # A que a parametriza‡„o se refere
                                  'TIPO': 'BT'})
        evento_dic['BARRA'] = True  # Define que planilha BARRA da LP padr„o ser  lida
        evento_dic['Disjuntor'] = True  # Define que planilha Disjuntor da LP padr„o ser  lida
        evento_dic['Secc'] = True  # Define que planilha Secc da LP padr„o ser  lida

# in¡cio de lista de Reator
 # Carregar dados Reator na linha enquanto existir c¢digo da Reator na linha Excel
    Eventos = arq_conf.find_all('reator')
    if Eventos:
        for evento in Eventos:
            # 0 - C¢digo operacional Reator Ex. 04E1
            conf_Reator_array.append({'COD': str(evento.string),
                                      # 1 - Nome do painel Ex. 4UA4A
                                      'PNL': str(evento['painel']),
                                      # 2 - Reator Manobr vel (Sim ou N„o)
                                      'BRM': 'Sim' if str(evento['manob']) == 'True' else 'N„o',
                                      # 3 - Equipamento Ex. Banco Monof sico
                                      'EQP': str(evento['equip']),
                                      # RDP Stand Alone
                                      'RDP': 'Sim' if str(evento['rdp']) == 'True' else 'N„o',
                                      # Bay Unit do 87B
                                      'F9': 'Sim' if str(evento['f9']) == 'True' else 'N„o',
                                      # Cƒmaras PASS. Pega texto "A , B", coloca em  maipusculo, transforma em Array por v¡rgula e retira espa‡o de cada elemento
                                      'PASSCam': tratar_str_secc(str(evento['camarapass'])),
                                      # Conjunto de comando das seccionadoras
                                      'PASSSecc': tratar_str_secc(str(evento['conjuntosecc'])),
                                      # A que a parametriza‡„o se refere
                                      'TIPO': 'Reator'})
        evento_dic['Reator'] = True  # Define que planilha Reator da LP padr„o ser  lida
        evento_dic['Disjuntor'] = True  # Define que planilha Disjuntor da LP padr„o ser  lida
        evento_dic['Secc'] = True  # Define que planilha Secc da LP padr„o ser  lida
        index_linha += 1
 # in¡cio da lista de V„o Segregado

 #Carregar dados Painel de Interface enquanto existir C¢digo do v„o na linha Excel
    Eventos = arq_conf.find_all('acesso')
    if Eventos:
        for evento in Eventos:
                                   # 0 - Codigo Operacional do v„o
            conf_PInterface_array.append({'COD': str(evento.string),
                                        # 1 - Nome do Painel do ACESSANTE Ex: 4UA13
                                          'PNL': str(evento['painelacess']),
                                        # 2 - Se vai ser em um Painel j  existente
                                          'PNLEXIST':'Sim' if str(evento['painelexist']) == 'True' else 'N„o',
                                        # 3 - N£mero da UC em um painel existente
                                         'N£mero_UC_CHESF': int(evento['num-uc-chesf']),
                                        # 4 - N£mero da UC em um painel existente
                                         'N£mero_UC_ACESSANTE': int(evento['num-uc-acessante']),
                                        # 5 - Arranjo do v„o
                                          'ARR': str(evento['arranjo']),
                                        # 6 - Se tem Terminal Server
                                          'TS': 'Sim' if str(evento['ts']) == 'True' else 'N„o',
                                        # 7 - N£mero do primeiro Terminal Server
                                          'TS-DE': int(evento['ts-de']),
                                        # 8 - N£mero do £ltimo Terminal Server
                                          'TS-ATE': int(evento['ts-ate']),
                                        # 9 - Se Tem Redbox
                                          'RB': 'Sim' if str(evento['rb']) == 'True' else 'N„o',
                                        # 10 - N£mero do primeiro Redbox
                                          'RB-DE': int(evento['redbox-de']),
                                        # 11 - N£mero do £ltimo RedBox
                                          'RB-ATE': int(evento['redbox-ate']),
                                        #12 - Se Tem Multimedidor
                                          'MM': 'Sim' if str(evento['multimedidor']) == 'True' else 'N„o',
                                        #13 - N£mero do primeiro Multimedidor
                                          'MM-DE': int(evento['mm-de']),
                                        #14 - N£mero do £ltimo Multimedidor
                                          'MM-ATE': int(evento['mm-ate']),
                                        #15 - Sigla da LT Remota ao v„o segregado
                                          'LTREMOTA': str(evento['ltremota'])
                                          })
        evento_dic['P. Interface'] = True

#  in¡cio de lista de Trafo Terra
 # Carregar dados Trafo Terra enquanto existir c¢digo da Trafo Terra na linha Excel
    Eventos = arq_conf.find_all('tterra')
    if Eventos:
        for evento in Eventos:
            # 0 - C¢digo operacional TT Ex. 02A1
            conf_TT_array.append({'COD': str(evento.string),
                                  # 1 - Nome do painel Ex. 2UA4A
                                  'PNL': str(evento['painel']),
                                  # Cƒmaras PASS. Pega texto "A , B", coloca em  maipusculo, transforma em Array por v¡rgula e retira espa‡o de cada elemento
                                  'PASSCam': tratar_str_secc(str(evento['camarapass'])),
                                  # Conjunto de comando das seccionadoras
                                  'PASSSecc': tratar_str_secc(str(evento['conjuntosecc'])),
                                  # A que a parametriza‡„o se refere
                                  'TIPO': 'TT'})
        evento_dic['T_Terra'] = True  # Define que planilha T_Terra da LP padr„o ser  lida
        evento_dic['Secc'] = True  # Define que planilha Secc da LP padr„o ser  lida
        index_linha += 1

#  in¡cio de lista de Painel de Prote‡„o de Barras
# Carregar dados Painel de Prote‡„o de Barras enquanto existir Painel de Prote‡„o de Barras na linha Excel
    Eventos = arq_conf.find_all('protbarra')
    if Eventos:
        for evento in Eventos:
            # 0 - Nome do Painel Ex. 4UA8
            conf_P87B_array.append({'PNL': str(evento.string),
                                    # 1 - N£mero de Pain‚is Ex. 2
                                    'NPNL': int(evento['qtpan']),
                                    # 2 - Arranjo
                                    'ARR': str(evento['arranjo']),
                                    # 3 - B.U. Instalada no Painel (Sim/N„o)
                                    'BU': 'Sim' if str(evento['bu-no-painel']) == 'True' else 'N„o',
                                    # 4 - V„os com B.U. Ex. 04T1/04V1/04D1
                                    'VAOS': str(evento['vaos']),
                                    # A que a parametriza‡„o se refere
                                    'TIPO': '87B'})
        evento_dic['BARRA'] = True  # Define que planilha BARRA da LP padr„o ser  lida

    # in¡cio de lista de Banco Capacitor
  # Carregar dados Banco Capacitor enquanto existir Banco Capacitor na linha Excel
    Eventos = arq_conf.find_all('bcapshunt')
    if Eventos:
        for evento in Eventos:
            # 0 - C¢digo operacional Ex. 04H1
            conf_BCap_array.append({'COD': str(evento.string),
                                    # 1 - - Nome do Painel Ex. 4UA6H
                                    'PNL': str(evento['painel']),
                                    # 2 - Arranjo
                                    'ARR': str(evento['arranjo']),
                                    # RDP Stand Alone
                                    'RDP': 'Sim' if str(evento['rdp']) == 'True' else 'N„o',
                                    # Bay Unit do 87B
                                    'F9': 'Sim' if str(evento['f9']) == 'True' else 'N„o',
                                    # A que a parametriza‡„o se refere
                                    'TIPO': 'BCap'})
        evento_dic['B_CAP'] = True  # Define que planilha B_CAP da LP padr„o ser  lida
        evento_dic['Disjuntor'] = True  # Define que planilha Disjuntor da LP padr„o ser  lida
        evento_dic['Secc'] = True  # Define que planilha Secc da LP padr„o ser  lida

#in¡cio de lista de Banco Capacitor S‚rie
    # Carregar dados Banco Capacitor enquanto existir Banco Capacitor na linha Excel
    Eventos = arq_conf.find_all('bcapserie')
    if Eventos:
        for evento in Eventos:
            # 0 - C¢digo operacional Ex. 04H1
            conf_BCS_array.append({'COD': str(evento.string),
                                   # 1 - - Nome do Painel Ex. 4UA6H
                                   'PNL': str(evento['painel']),
                                   # A que a parametriza‡„o se refere
                                   'TIPO': 'BCS'})
        evento_dic['BCS'] = True  # Define que planilha BCS da LP padr„o ser  lida
        evento_dic['Disjuntor'] = True  # Define que planilha Disjuntor da LP padr„o ser  lida
        evento_dic['Secc'] = True  # Define que planilha Secc da LP padr„o ser  lida
        index_linha += 1

    # Carregar dado do ECE
    Eventos = arq_conf.find_all('ece')
    if Eventos:
        for evento in Eventos:
    # 0 - C¢digo operacional Ex. 04B1
            conf_ECE_array.append({'COD': str(evento.string),
                                   # 1 - Nome Painel UA Ex. 4UA7
                                   'PNL': str(evento['painel']),
                                   # A que a parametriza‡„o se refere
                                   'TIPO': 'ECE'})
        evento_dic['ECE'] = True  # Define que planilha ECE da LP padr„o ser  lida


 # , Carregar dado do Compensador S¡ncrono
    Eventos = arq_conf.find_all('compsinc')
    if Eventos:
        for evento in Eventos:
            # 0 - C¢digo operacional Ex. 04K1
            conf_CS_array.append({'COD': str(evento.string),
                                  # 1 - Nome Painel UA Ex. 4UA7
                                  'PNL': str(evento['painel']),
                                  # A que a parametriza‡„o se refere
                                  'TIPO': 'CS'})
        evento_dic['CS'] = True  # Define que planilha CS da LP padr„o ser  lida
        index_linha += 1

 # Carregar dado de Prepara‡„o a Reenergiza‡„o
    Eventos = arq_conf.find_all('prepreen')
    if Eventos:
        for evento in Eventos:
            # 0 - C¢digo operacional Ex. UTR
            conf_PR_array.append({'COD': str(evento.string),
                                  # A que a parametriza‡„o se refere
                                  'TIPO': 'PR'})
        evento_dic['Prep. Reen.'] = True  # Define que planilha Prep. Reen. da LP padr„o ser  lida

    # Carregar dado de Sistema de Regula‡„o
    Eventos = arq_conf.find_all('sistreg')
    if Eventos:
        for evento in Eventos:
            # 0 - C¢digo operacional. SAGE, UTR- ou PCPG
            conf_SR_array.append({'COD': str(evento.string),
                                  # 1 - Tens„o Regula‡„o. 230kV, 138kV, 69kV ou 13,8kV
                                  'TENSAO': str(evento['tensao-reg']),
                                  # 2 - Nome Painel UA, caso exista Ex. 4UA7
                                  'PNL': str(evento['painel']),
                                  # A que a parametriza‡„o se refere
                                  'TIPO': 'SR'})
        evento_dic['Prep. Reen.'] = True  # Define que planilha Prep. Reen. da LP padr„o ser  lida
        index_linha += 1

    # index_linha = 96
    #    while sheet.cell(index_linha,6).value: #Linha 97 do LP_Config.xls, Carregar dado do Compensador Est tico
    #                        # 0 - C¢digo operacional Ex. 04Q1
    #        conf_CE_array.append({'COD':sheet.cell(index_linha,0).value.upper(),
    #                        # 1 - Nome Painel UA Ex. 4UA7
    #                        'PNL':sheet.cell(index_linha,1).value.upper()})
    #        index_linha += 1
    #        evento_dic['CE'] = True

    # Carregar dados Serv. Aux. se existir nome do painel da UA
    Eventos = arq_conf.saux
    if Eventos:
        # 0 - Nome Painel UA Ex. 4UA7
        conf_SA = {'PNL': str(Eventos.string),
                   # 1 - Tens„o CA. Ex. 220Vca
                   'VCA': str(Eventos['tensao-ca']),
                   # 2 - Tens„o CC. Ex. 125Vca
                   'VCC': str(Eventos['tensao-cc']),
                   # 3 - Nome Pain‚is Serv. Aux. Ex. PT1/PT3/PT4/PT1EA/PT1EB
                   'PNLSA': str(Eventos['nome-painel-saux']),
                   # 4 - Barras CA Ex. B1
                   'BSCA': str(Eventos['barras-sup-ca']),
                   # 5 - Barras CC Ex. B1/B2
                   'BSCC': str(Eventos['barras-sup-cc']),
                   # 6 - Disjuntores Motorizados CA Ex. T7/T8/G1/B2/R1/R2
                   'DJCA': str(Eventos['disj-sup-ca']),
                   # 7 - Disjuntores Motorizados CC Ex. A1A/A2B/A2A/A1B
                   'DJCC': str(Eventos['disj-sup-cc']),
                   # A que a parametriza‡„o se refere
                   'TIPO': 'SA'}
        evento_dic['SAs'] = True  # Define que planilha SA da LP padr„o ser  lida

    # Soma de arrays para gera‡„o de Secc e Disjuntores
    parametros = conf_LT_array + conf_Trafo_array + conf_BT_array + conf_Reator_array + conf_TT_array + conf_BCap_array + conf_BCS_array
    # ----------Ler LP Padr„o----------#
    try:
        book = load_workbook(lp_padrao, data_only=True)  # Abrir arquivo de LP Padr„o definido no arquivo de configura‡„o
    except:
        print_exc(file=stdout)
        aviso = 'Arquivo \"' + lp_padrao + u'\" n„o encontrado'
        mensagem_erro('Erro', aviso)
    abas = book.sheetnames
    if len(abas) < 4: mensagem_erro('Erro', u'O programa n„o reconheceu a LP Padr„o como v lida')
    for plan_index in range(3, 22):  # Ler Planilhas com index 3 a 22 (quarta a vig‚sima primeira), uma a uma
        sheet = book[abas[plan_index]]  # Abrir planilhas
        if abas[plan_index] not in evento_dic: mensagem_erro('Erro', u'O programa n„o reconheceu a LP Padr„o como v lida {}')
        if evento_dic[abas[plan_index]]:  # Verificar se no arquivo de configura‡„o foi solicitado ler planilha
            # Gerar dicion rio titulo_dic (dicion rio de t¡tulos das colunas)
            for li in range(2, 10):  # Varrer as linhas de 2 a 10
                for i in range(sheet.max_column):  # Varrer as colunas da linha
                    texto_coluna = str(sheet.cell(row=li, column=i+1).value).upper().strip()
                    #texto_coluna = sheet.cell_value(li,
                                                 #   i).upper().strip()  # Pegar texto da c‚lula em mai£sculo e sem espa‡o antes e depois
                    if texto_coluna == '':  # Gravar posi‡„o do valor vazio (ap¢s £ltima coluna)
                        titulo_dic[texto_coluna] = i+1
                    elif texto_coluna not in titulo_dic or texto_coluna =='NONE':  # Iserir chave se n„o existir no dicion rio (garante pegar apenas primeira ocorrˆncia do t¡tulo
                        titulo_dic[texto_coluna] = i+1
                if 'ID (SAGE)' in titulo_dic: break  # Se foi passado pela linha com chave "ID (SAGE)", sair do "for" de varrer linhas

            # Definir linha de in¡cio da LP
            li += 1  # Seleciona linha ap¢s o t¡tulo
            while True:
                if sheet.cell(row=li, column=titulo_dic[
                    u'ID (SAGE)']).value:  # Verifica se a c‚lula (li,conula de t¡tulo) est  preenchida com algum valor
                    break  # Parar de procurar linha preenchida
                else:
                    li += 1  # Selecionar linha seguinte

                if sheet.cell(row=li, column=titulo_dic[
                    u'ID (SAGE)']).value:  # Verifica se a c‚lula (li,conula de t¡tulo) est  preenchida com algum valor
                    break  # Parar de procurar linha preenchida
                else:
                    li += 1  # Selecionar linha seguinte

            for index_linha in range(li, sheet.max_row):  # Ler colulas da linha definida at‚ o final da LP
                if (sheet.cell(row=index_linha, column= titulo_dic[
                    u'VŽOS DIGITAIS']).value == 'X'):  # Ler apenas linhas do Excel que tenha campo "V„o Digital" marcado
                    # ----------In¡cio de tratamento de TAG (ID SAGE) e Descri‡„o----------#
                    tratar_id = str(sheet.cell(row=index_linha, column = titulo_dic[u'ID (SAGE)']).value)  # ID SAGE
                    descricao = str(sheet.cell(row= index_linha, column = titulo_dic[u'DESCRI€ŽO']).value).strip()  # Descri‡„o
                    observacao = str(sheet.cell(row= index_linha, column = titulo_dic[u'OBSERVA€ŽO']).value)


                    tratar_id = tratar_id.replace('ZZZ', Codigo_SE)  # Substituir c¢digo da SE no ID SAGE
                    descricao = descricao.replace('ZZZ', Codigo_SE)  # Substituir c¢digo da SE na Descri‡„o

                    tratar_IdSage = []
                    if '[' in tratar_id:  # Transforma ID SAGE de string lida no Excel para um array, usando "/" para definir separa‡„o dos campos
                        campos = tratar_id.split(
                            ':')  # Dividir os 4 ou 5 campos existentes no ID SAGE (separados por ":")
                        for ied in campos[2].strip('[]').split(
                                '/'):  # Retirar "[" e "]" e gerar array de ieds com "/" como separador
                            tratar_IdSage.append(
                                ':'.join(campos[:2] + [ied] + campos[3:]))  # Adicionar a arrey o ID SAGE a ser tratado
                    else:
                        tratar_IdSage.append(tratar_id)

                    for tratar in tratar_IdSage:  # Passar arrey tratar_IdSage com ID_SAGE
                        if abas[plan_index] == 'SD':
                            if ':RDP' not in tratar:
                                for parametros_SD in conf_SD_array:
                                    if tratar[-4:] == 'FDSD' or tratar[
                                                                -4:] == 'FCpp':  # Falha Dispositivo e Falha Porta de Comunica‡„o
                                        disp_array = []  # Lista de dispositivos para SAGE e Bastidores de Rede
                                        for nsw in range(int(parametros_SD['DE_SW']), int(parametros_SD['ATE_SW']) + 1):
                                            sw = 'SW' + str(nsw)
                                            disp_array.append([sw])
                                        if parametros_SD['FW'] == 'Sim':
                                            disp_array.append(['FW'])
                                        if parametros_SD['RB'] == 'Sim':
                                            for nrb in range(int(parametros_SD['DE_RB']),int(parametros_SD['ATE_RB']) + 1):
                                                rb = 'RB' + str(nrb)
                                                disp_array.append([rb])
                                        if tratar[-4:] == 'FDSD':  # Falha Dispositivo
                                            for disp in disp_array:
                                                tratar_1 = tratar.replace('{DISP}', disp[0])
                                                gravar_ponto(tratar_1, descricao)
                                                k_sd += 1
                                        if tratar[-4:] == 'FCpp':  # Falha porta de Comunica‡„o
                                            for disp in disp_array:
                                                if 'FW' in disp[0]:
                                                    tratar_1 = tratar.replace('{DISP}', disp[0])
                                                    for porta in range(1, int(parametros_SD['POR_FW']) + 1):
                                                        tag = 'FC{:02}'.format(porta)
                                                        tratar_2 = tratar_1.replace('FCpp', tag)
                                                        gravar_ponto(tratar_2, descricao)
                                                        k_sd += 1
                                                elif 'RB' in disp[0]:
                                                    tratar_1 = tratar.replace('{DISP}', disp[0])
                                                    for porta in range(1, int(parametros_SD['POR_RB']) + 1):
                                                        tag = 'FC{:02}'.format(porta)
                                                        tratar_2 = tratar_1.replace('FCpp', tag)
                                                        gravar_ponto(tratar_2, descricao)
                                                        k_sd += 1
                                                elif 'SW' in disp[0]:
                                                    tratar_1 = tratar.replace('{DISP}', disp[0])
                                                    for porta in range(1, int(parametros_SD['POR_SW']) + 1):
                                                        tag = 'FC{:02}'.format(porta)
                                                        tratar_2 = tratar_1.replace('FCpp', tag)
                                                        gravar_ponto(tratar_2, descricao)
                                                        k_sd += 1
                                    elif 'SELE' in descricao.upper():  # pontos de sele‡„o de comando
                                        if parametros_SD['SB'] == 'SAGE':
                                            gravar_ponto(tratar, descricao)
                                            k_sd += 1
                                    else:
                                        tratar_1 = tratar.replace('{PNL}', parametros_SD['PNL'])
                                        gravar_ponto(tratar_1, descricao)
                                        k_sd += 1
                            elif 'conf_RDP' in locals():  # Trata-se de ponto de RDP, mas checa se foi definido conf_RDP
                                for i in range(int(conf_RDP['DE_RDP']), int(conf_RDP['ATE_RDP']) + 1):
                                    tratar_1 = tratar.replace('RDPn', 'RDP{}'.format(i))
                                    gravar_ponto(tratar_1, descricao)
                                    k_sd += 1

                        elif abas[plan_index] == 'LT':
                            for parametros_LT in conf_LT_array:
                                ###Condi‡”es para processar o ponto###

                                # N„o conste observa‡„o #RELIGMONO ou conste #RELIGMONO e "Relig." definido como 'MONO/TRI'
                                cd1 = ('#RELIGMONO' not in observacao.upper()) or (
                                    '#RELIGMONO'.upper() in observacao and parametros_LT['79'] == 'MONO/TRI')
                                # N„o conste observa‡„o #DISJEMEIO ou conste #DISJEMEIO e "ARRANJO" definido como 'DISJ E MEIO'
                                cd2 = ('#DISJEMEIO' not in observacao.upper()) or (
                                    '#DISJEMEIO' in observacao.upper() and parametros_LT['ARR'] == 'DISJ E MEIO')
                                # N„o conste na descri‡„o #87 ou conste #87 e "Tem 87L" definido como 'Sim'
                                cd3 = ('#87' not in observacao) or (
                                    '#87' in observacao and parametros_LT['87L'] == 'Sim')
                                # N„o conste no TAG :RDP ou conste e "RDP" definido como 'Sim'
                                cd4 = (':RDP' not in tratar) or (':RDP' in tratar and parametros_LT['RDP'] == 'Sim')
                                # N„o conste na observa‡„o #TELEPROT ou conste #TELEPROT e c¢digo da linha inicie com 04 ou 05 (LT de 230kV ou 500kV)
                                cd5 = ('#TELEPROT' not in observacao.upper()) or (
                                    '#TELEPROT' in observacao.upper() and (int(parametros_LT['COD'][1:2]) > 3))
                                # N„o conste na observacao #21 ou conste '(21)' e c¢digo da linha inicie com 04 ou 05 (LT de 230kV ou 500kV) ou ID SAGE conste 'F1' e c¢digo da linha inicie com 02,03,04,05 (excluindo LT 13,8kV)
                                cd6 = ('#21' not in observacao) or (
                                    '#21' in observacao and (int(parametros_LT['COD'][1:2]) > 3) or (
                                        ':F1' in tratar and int(parametros_LT['COD'][1:2]) > 1))
                                # N„o conste na observacao #68 ou conste #68 e c¢digo da linha inicie com 04 ou 05 (LT de 230kV ou 500kV) ou ID SAGE conste 'F1' e c¢digo da linha inicie com 02,03,04,05 (excluindo LT 13,8kV)
                                cd7 = ('#68' not in observacao) or (
                                    '#68' in observacao and (int(parametros_LT['COD'][1:2]) > 3) or (
                                        ':F1' in tratar and int(parametros_LT['COD'][1:2]) > 1))
                                # N„o conste na observacao #78 ou conste #78 e c¢digo da linha inicie com 04 ou 05 (LT de 230kV ou 500kV) ou ID SAGE conste 'F1' e c¢digo da linha inicie com 02,03,04,05 (excluindo LT 13,8kV)
                                cd8 = ('#78' not in observacao) or (
                                    '#78' in observacao and (int(parametros_LT['COD'][1:2]) > 3) or (
                                        ':F1' in tratar and int(parametros_LT['COD'][1:2]) > 1))
                                # N„o conste na observacao #BARRADUPLA ou conste #BARRADUPLA e o Arranjo n„o ‚ Barra Principal e Transferˆncia
                                cd9 = ('#BARRADUPLA' not in observacao.upper()) or (
                                    '#BARRADUPLA' in observacao.upper() and parametros_LT['ARR'] != 'BPT')
                                # N„o contem no ID 'F9' ou contem 'F9' e "Bay Unit (F9)" definido como 'Sim'
                                cd10 = ('F9' not in tratar) or ('F9' in tratar and parametros_LT['F9'] == 'Sim')
                                # N„o contenha no ID 'F2' nem 'UC' ou cotenha 'F2' ou 'UC' e segunda casa do c¢digo da linha (ex. N£mero 1 de 01L1) for maior que 1  (maior que 13,8kV)
                                cd11 = (('F2' not in tratar) and ('UC' not in tratar)) or (
                                    (('F2' in tratar) or ('UC' in tratar)) and (int(parametros_LT['COD'][1]) > 1))
                                # N„o conste obserna‡„o #PASS ou conste #PASS e "PASSSecc" n„o vazio
                                cd12 = ('#PASS' not in observacao.upper()) or (
                                    '#PASS'.upper() in observacao and bool(parametros_LT.get('PASSSecc', [None])[0]))
                                # N„o conste na observacao #85PNL ou conste #85PNL e "PAINEL TELEPROT." definido como 'Sim'
                                cd13 = ('#85PNL' not in observacao.upper()) or (
                                    '#85PNL' in tratar and parametros_LT['85PNL'] == 'Sim')
                                cd14 = ('#ACESSANTE' not in observacao.upper())


                                if cd1 * cd2 * cd3 * cd4 * cd5 * cd6 * cd7 * cd8 * cd9 * cd10 * cd11 * cd12 * cd13 * cd14 :

                                    tratar_1 = tratar.replace('0YYY', parametros_LT['COD'])
                                    descricao_0 = descricao.replace('0YYY', parametros_LT['COD'])
                                    descricao_1 = descricao_0.replace('XXX', parametros_LT['LTREM'])

                                    cd15 = 'FPCn' not in tratar_1
                                    cd16 = 'FPTn' not in tratar_1
                                    cd17 = 'FPCn' not in tratar_1
                                    cd18 = 'FPDn' not in tratar_1
                                    cd19 = 'FDSD' not in tratar_1
                                    cd20 = cd15 * cd16 * cd17 * cd18 * cd19

                                    linha69 = tratar_1[5] == '2'

                                    if '{PNL}' in tratar_1:  # Substituir {PNL} pelo nome do painel
                                        if int(parametros_LT['PNL'][
                                               0:1]) > 3:  # Caso o painel inicie com 4 ou 5 (230kV ou 500kV) gerar painel 1 e painel 2
                                            for npnl in [1, 2]:
                                                painel = parametros_LT['PNL'] + str(npnl)
                                                tratar_2 = tratar_1.replace('{PNL}', painel)
                                                gravar_ponto(tratar_2, descricao_1)
                                                k_lt += 1
                                        else:  # Se o nome do painel n„o inicia com 4 e 5 ( 1,2 ou 3 correspondendo a 13,8kV, 69kV, ou 138kV)
                                            tratar_2 = tratar_1.replace('{PNL}', parametros_LT['PNL'])
                                            gravar_ponto(tratar_2, descricao_1)
                                            k_lt += 1

                                    elif tratar_1[
                                        5] == '2' and cd20:  # Casos de linha de 69kV, substituir F2 e UC1 por F3
                                            if ':UC1' in tratar_1:
                                                tratar_2 = tratar_1.replace(':UC1', ':F3')
                                                gravar_ponto(tratar_2, descricao_1)
                                                k_lt += 1
                                            else:
                                                continue
                                    elif 'FPCn' in tratar_1:
                                        for n_canal in range(1, 3):
                                            texto_canal = tratar_1[tratar_1.find('FPCn'):-1] + str(n_canal)
                                            tratar_2 = tratar_1.replace('FPCn', texto_canal)
                                            if linha69:
                                                if ':UC1' in tratar_1:
                                                    tratar_2 = tratar_2.replace(':UC1', ':F3')
                                            gravar_ponto(tratar_2, descricao_1)
                                            k_lt += 1
                                    elif 'FPTn' in tratar_1:
                                        for n_canal in range(1, 3):
                                            texto_canal = tratar_1[tratar_1.find('FPTn'):-1] + str(n_canal)
                                            tratar_2 = tratar_1.replace('FPTn', texto_canal)
                                            if linha69:
                                                if ':UC1' in tratar_1:
                                                    tratar_2 = tratar_2.replace(':UC1', ':F3')
                                            gravar_ponto(tratar_2, descricao_1)
                                            k_lt += 1
                                    elif 'FPDn' in tratar_1:
                                        for n_canal in range(1, 3):
                                            texto_canal = tratar_1[tratar_1.find('FPDn'):-1] + str(n_canal)
                                            tratar_2 = tratar_1.replace('FPDn', texto_canal)
                                            if linha69:
                                                if ':UC1' in tratar_1:
                                                    tratar_2 = tratar_2.replace(':UC1', ':F3')
                                            gravar_ponto(tratar_2, descricao_1)
                                            k_lt += 1
                                    elif tratar_1[-4:] == 'FDSD':  # Falha Dispositivo
                                        disp_array = []  # Lista de dispositivos para P.Interface
                                        if parametros_LT['TS'] == 'Sim':
                                            for nts in range(int(parametros_LT['DE_TS']), int(parametros_LT['ATE_TS']) + 1):
                                                ts = 'TS' + str(nts)
                                                disp_array.append([ts])
                                        if parametros_LT['RB'] == 'Sim':
                                            for nrb in range(int(parametros_LT['DE_RB']),int(parametros_LT['ATE_RB']) + 1):
                                                rb = 'RB' + str(nrb)
                                                disp_array.append([rb])
                                        if tratar[-4:] == 'FDSD' and disp_array != '':  # Falha Dispositivo
                                            for disp in disp_array:
                                                tratar_1 = tratar.replace('{DISP}', disp[0])
                                                if linha69:
                                                    if ':UC1' in tratar_1:
                                                        tratar_2 = tratar_2.replace(':UC1', ':F3')
                                                gravar_ponto(tratar_1, descricao)
                                                k_lt += 1
                                    else:
                                        gravar_ponto(tratar_1, descricao_1)
                                        k_lt += 1

                        elif abas[plan_index] == 'P. Interface':
                            for parametros_PINT in conf_PInterface_array:

                                cd1 = ('#ACESSANTE' in observacao.upper() or ('#ACESSANTE' in observacao.upper() and '#ACESSADA' in observacao.upper()) or '#CHESF' in observacao.upper())
                                #                               # N„o conste na observa‡„o #ACESSADA ou conste #ACESSADA e o P.Interface definido como ACESSADA
                                # N„o conste na obesrva‡„o #PAINEL ou conste #PAINEL e o PAINELINT definido como Sim
                                cd2 = ('#PAINEL' not in observacao.upper()) or (
                                            '#PAINEL' in observacao.upper() and parametros_PINT['PNLEXIST'] == 'N„o')
                                cd3 = ('Medida Inexistente para os casos de linhas e alimentadores de 69kV e 13,8kV' not in observacao or ('Medida Inexistente para os casos de linhas e alimentadores de 69kV e 13,8kV' in observacao and parametros_PINT['COD'][1] != '2'))

                                if cd1 * cd2 * cd3:

                                    tratar_1 = tratar.replace('YYY', parametros_PINT['COD'][-3:])
                                    descricao_0 = descricao.replace('YYY',parametros_PINT['COD'][-3:])
                                    descricao_1 = descricao_0.replace('XXX', parametros_PINT['LTREMOTA'])

                                    if '{PNL}' in tratar_1:  # Substituir {PNL} pelo nome do painel
                                        if '#CHESF' in observacao.upper():
                                            painel = parametros_PINT['PNL'][0] + 'UA12-' + parametros_PINT['PNL'][6]
                                            tratar_2 = tratar_1.replace('{PNL}', painel)
                                            if 'UC1' in tratar_1:
                                                if parametros_PINT['N£mero_UC_CHESF'] != "" and parametros_PINT['PNLEXIST'] == 'Sim':
                                                    if int(parametros_PINT['N£mero_UC_CHESF']) != 1:
                                                        tratar_3 = tratar_2.replace('UC1','UC'+ '{:.0f}'.format(parametros_PINT['N£mero_UC_CHESF']))
                                                        if 'FPCn' in tratar_3:
                                                            for n_canal in range(1, 3):
                                                                texto_canal = tratar_3[tratar_3.find('FPCn'):-1] + str(n_canal)
                                                                tratar_4 = tratar_3.replace('FPCn', texto_canal)
                                                                gravar_ponto(tratar_4, descricao_1)
                                                                k_Pint += 1
                                                        else:
                                                            gravar_ponto(tratar_3, descricao_1)
                                                            k_Pint += 1
                                                else:
                                                    if 'FPCn' in tratar_2:
                                                        for n_canal in range(1, 3):
                                                            texto_canal = tratar_2[tratar_2.find('FPCn'):-1] + str(
                                                                n_canal)
                                                            tratar_3 = tratar_2.replace('FPCn', texto_canal)
                                                            gravar_ponto(tratar_3, descricao_1)
                                                            k_Pint += 1
                                                    else:
                                                        gravar_ponto(tratar_2, descricao_1)
                                                        k_Pint += 1
                                            else:
                                                gravar_ponto(tratar_2, descricao_1)
                                                k_Pint += 1
                                        if  '#ACESSANTE' in observacao.upper():
                                            tratar_2 = tratar_1.replace('{PNL}', parametros_PINT['PNL'])
                                            if 'UC1' in tratar_1:
                                                if parametros_PINT['N£mero_UC_ACESSANTE'] != '':
                                                    tratar_3 = tratar_2.replace('UC1', 'UC' + '{:.0f}'.format(
                                                    parametros_PINT['N£mero_UC_ACESSANTE']))
                                                    if int(parametros_PINT['N£mero_UC_ACESSANTE']) != 1:
                                                        gravar_ponto(tratar_3, descricao_1)
                                                        k_Pint += 1
                                                    else:
                                                        gravar_ponto(tratar_3, descricao_1)
                                                        k_Pint += 1
                                                else:
                                                    gravar_ponto(tratar_2, descricao_1)
                                                    k_Pint += 1
                                            else:
                                                gravar_ponto(tratar_2, descricao_1)
                                                k_Pint += 1
                                    elif '-Z' in tratar_1:
                                        arranjo = parametros_PINT['ARR']
                                        for nsecc in [1, 2, 4, 5, 6, 7, 8, 9]:
                                            if nsecc == 1 and (arranjo[0:2] == 'BD' or arranjo == 'BT'):
                                                tag_secc ='-' + '1'
                                                tratar_2 = tratar_1.replace('-Z', tag_secc)
                                                descricao_1 = descricao_0.replace('-Z', tag_secc)
                                                gravar_ponto(tratar_2, descricao_1)
                                                k_Pint += 1
                                            elif nsecc == 2 and (arranjo[0:2] == 'BD' or arranjo == 'BT'):
                                                tag_secc = '-' + '2'
                                                tratar_2 = tratar_1.replace('-Z', tag_secc)
                                                descricao_1 = descricao_0.replace('-Z', tag_secc)
                                                gravar_ponto(tratar_2, descricao_1)
                                                k_Pint += 1
                                            elif nsecc == 4 and (
                                                                arranjo == 'BPT' or arranjo == 'BD5' or arranjo == 'DISJ E MEIO'):
                                                tag_secc = '-' + '4'
                                                tratar_2 = tratar_1.replace('-Z', tag_secc)
                                                descricao_1 = descricao_0.replace('-Z', tag_secc)
                                                gravar_ponto(tratar_2, descricao_1)
                                                k_Pint += 1
                                            elif nsecc == 5 and (arranjo == 'BPT' or arranjo[
                                                                                     0:2] == 'BD' or arranjo == 'DISJ E MEIO'):
                                                tag_secc = '-' + '5'
                                                tratar_2 = tratar_1.replace('-Z', tag_secc)
                                                descricao_1 = descricao_0.replace('-Z', tag_secc)
                                                gravar_ponto(tratar_2, descricao_1)
                                                k_Pint += 1
                                            elif nsecc == 6 and (
                                                                    arranjo == 'BPT' or arranjo == 'BD4' or arranjo == 'BD5' or arranjo == 'BCS'):
                                                tag_secc = '-' + '6'
                                                tratar_2 = tratar_1.replace('-Z', tag_secc)
                                                descricao_1 = descricao_0.replace('-Z', tag_secc)
                                                gravar_ponto(tratar_2, descricao_1)
                                                k_Pint += 1
                                            elif nsecc == 7 and parametros_PINT['COD'][2] != 'T' and (
                                                                    arranjo == 'BPT' or arranjo[
                                                                                        0:2] == 'BD' or arranjo == 'DISJ E MEIO' or arranjo == 'BCS'):
                                                if arranjo[0:3] != 'BD3' and arranjo != 'BCS':
                                                    tag_secc = '-' + '7'
                                                    tratar_2 = tratar_1.replace('-Z', tag_secc)
                                                    descricao_1 = descricao_0.replace('-Z', tag_secc)
                                                    gravar_ponto(tratar_2, descricao_1)
                                                    k_Pint += 1
                                                else:
                                                    for secbd3 in ['A', 'B']:
                                                        tag_secc = '-' + '7' + secbd3
                                                        tratar_2 = tratar_1.replace('-Z', tag_secc)
                                                        descricao_1 = descricao_0.replace('-Z', tag_secc)
                                                        gravar_ponto(tratar_2, descricao_1)
                                                        k_Pint += 1
                                            elif nsecc == 8 and (
                                                                    arranjo == 'DISJ E MEIO' or arranjo == 'BS' or arranjo == 'TT' or arranjo == 'BCS'):
                                                tag_secc = '-' + '8'
                                                tratar_2 = tratar_1.replace('-Z', tag_secc)
                                                descricao_1 = descricao_0.replace('-Z', tag_secc)
                                                gravar_ponto(tratar_2, descricao_1)
                                                k_Pint += 1
                                            elif nsecc == 9 and (arranjo == 'TT' or arranjo == 'BCS'):
                                                tag_secc ='-' + '9'
                                                tratar_2 = tratar_1.replace('-Z', tag_secc)
                                                descricao_1 = descricao_0.replace('-Z', tag_secc)
                                                gravar_ponto(tratar_2, descricao_1)
                                                k_Pint += 1
                                    elif tratar_1[
                                        5] == '2' and ':F2' in tratar_1:  # Caso de linha de 69kV com ID contendo F2 e ponto a ser tratato que j  n„o contenha UC1
                                        cd1 = ('FPCn' not in tratar_1)
                                        cd2 = ('FGOE' not in tratar_1)
                                        cd3 = ('FGPS' not in tratar_1)
                                        cd4 = ('FIRE' not in tratar_1)
                                        cd5 = ('FSPF' not in tratar_1)
                                        cd6 = ('RAUT' not in tratar_1)
                                        if cd1 * cd2 * cd3 * cd4 * cd5 * cd6:
                                            tratar_2 = tratar_1.replace(':F2', ':UC1')
                                            gravar_ponto(tratar_2, descricao_1)
                                            k_Pint += 1
                                    elif 'FPCn' in tratar_1:
                                        for n_canal in range(1, 3):
                                            texto_canal = tratar_1[tratar_1.find('FPCn'):-1] + str(n_canal)
                                            tratar_2 = tratar_1.replace('FPCn', texto_canal)
                                            gravar_ponto(tratar_2, descricao_1)
                                            k_Pint += 1
                                    elif tratar_1[-4:] == 'FDSD':  # Falha Dispositivo
                                        disp_array = []  # Lista de dispositivos para P.Interface
                                        if parametros_PINT['TS'] == 'Sim':
                                            for nts in range(int(parametros_PINT['TS-DE']),
                                                             int(parametros_PINT['TS-ATE']) + 1):
                                                ts = 'TS' + str(nts)
                                                disp_array.append([ts])
                                        if parametros_PINT['RB'] == 'Sim':
                                            for nrb in range(int(parametros_PINT['RB-DE']),
                                                             int(parametros_PINT['RB-ATE']) + 1):
                                                rb = 'RB' + str(nrb)
                                                disp_array.append([rb])
                                        if parametros_PINT['MM'] == 'Sim':
                                            for nMM in range(int(parametros_PINT['MM-DE']),
                                                             int(parametros_PINT['MM-ATE']) + 1):
                                                MM = 'MM' + str(nMM)
                                                disp_array.append([MM])
                                        if tratar[-4:] == 'FDSD' and disp_array != '':  # Falha Dispositivo
                                            for disp in disp_array:
                                                tratar_1 = tratar.replace('{DISP}', disp[0])
                                                gravar_ponto(tratar_1, descricao)
                                                k_Pint += 1
                                    else:
                                        gravar_ponto(tratar_1, descricao_1)
                                        k_Pint += 1

                        elif abas[plan_index] == 'Trafo':

                            for parametros_Trafo in conf_Trafo_array:
                                tensoes_trafo = parametros_Trafo['REL'].split('/')  # Array com n¡vel de Tens„o

                                ###Condi‡”es para processar o ponto###

                                # Verdadeiro se achar na descri‡„o uma das tens”es da lista "tensoes_trafo"
                                cd1 = True if re.search(
                                    '|'.join(tensoes_trafo + [i.replace(',', '.') for i in tensoes_trafo]),
                                    descricao) else False
                                # Na descri‡„o n„o consta valor de Tens„o ("500", "230", "138", "69", "6,9", "6.9", "13,8" ou "13.8")
                                cd2 = False if re.search('500|230|138|69|6[\.,]9|13[\.,]8', descricao) else True
                                # N„o conste obserna‡„o #TRIFASICO ou conste 'Apenas para Trafo Trifasico.' e "Equipamen." definido como 'Trif sico'
                                cd3 = ('#TRIFASICO' not in observacao.upper() or (
                                    '#TRIFASICO' in observacao.upper() and parametros_Trafo['BM'] == 'Trif sico'))
                                # N„o conste observa‡„o #MONOFASICO ou conste #MONOFASICO e "Equipamen." definido como 'Banco Monof.'
                                cd4 = ('#MONOFASICO' not in observacao.upper() or (
                                    '#MONOFASICO' in observacao.upper() and parametros_Trafo['BM'] == 'Banco Monof.'))
                                # N„o conste na descri‡„o #63 ou conste #63 e, "PP/PA" sendo ID SAGE F1 ou F2, ou "PU/PG" sendo ID SAGE F2 ou F3
                                cd5 = ('#63' not in observacao or ('#63' in observacao and (
                                    parametros_Trafo['PUPG'] == 'PP/PA' and ('F1' in tratar or 'F2' in tratar)) or (
                                                                       parametros_Trafo['PUPG'] == 'PU/PG' and (
                                                                           'F3' in tratar or 'F4' in tratar))))
                                # N„o conste na observacao #TERCIARIO ou conste #TERCIARIO e tenha 3 tens”es
                                cd6 = (('#TERCIARIO' not in observacao.upper()) or (
                                    ('#TERCIARIO' in observacao.upper()) and len(tensoes_trafo) == 3))
                                # ID SAGE sem F2 ou ID SAGE com F2 e prote‡„o do tipo PP/PA
                                cd7 = 'F2' not in tratar or ('F2' in tratar and (parametros_Trafo['PUPG'] == 'PP/PA'))
                                # ID SAGE sem F3 ou ID SAGE com F3 e prote‡„o do tipo PU/PG
                                cd8 = 'F3' not in tratar or ('F3' in tratar and (parametros_Trafo['PUPG'] == 'PU/PG'))
                                # ID SAGE sem F4 ou ID SAGE com F4 e prote‡„o do tipo PU/PG
                                cd9 = 'F4' not in tratar or ('F4' in tratar and (parametros_Trafo['PUPG'] == 'PU/PG'))
                                # N„o conste no TAG :RDP ou conste e "RDP" definido como 'Sim'
                                cd10 = (':RDP' not in tratar) or (':RDP' in tratar and parametros_Trafo['RDP'] == 'Sim')
                                # N„o conste obserna‡„o #PASS ou conste #PASS e "PASSSecc" n„o vazio
                                cd11 = ('#PASS' not in observacao.upper()) or (
                                    '#PASS'.upper() in observacao and bool(parametros_Trafo.get('PASSSecc', [None])[0]))
                                # N„o contem no ID 'F9' ou contem 'F9' e "Bay Unit (F9)" definido como 'Sim'
                                cd12 = ('F9' not in tratar) or ('F9' in tratar and parametros_Trafo['F9'] == 'Sim')
                                cd13 = ('#APLICACAO' not in observacao.upper()) or ('#APLICACAO' in observacao.upper() and parametros_Trafo['REGAPLIC'] == 'Sim')

                                if (cd1 + cd2) * cd3 * cd4 * cd5 * cd6 * cd7 * cd8 * cd9 * cd10 * cd11 * cd12 * cd13:
                                    if '0XTY' in tratar:
                                        tratar_1 = tratar.replace('0XTY', parametros_Trafo['COD'])
                                        descricao_1 = descricao.replace('0XTY', parametros_Trafo['COD'])
                                    elif '0YYY' in tratar:
                                        tratar_1 = tratar.replace('0YYY', parametros_Trafo['COD'])
                                        descricao_1 = descricao.replace('0YYY', parametros_Trafo['COD'])

                                    if '{PNL}' in tratar_1:  # Substituir {PNL} pelo nome do painel
                                        for npnl in [parametros_Trafo['PNLH'], parametros_Trafo['PNLX']]:
                                            tratar_2 = tratar_1.replace('{PNL}', str(npnl))
                                            gravar_ponto(tratar_2, descricao_1)
                                            k_trafo += 1
                                    elif 'F9' in tratar_1:  # ID SAGE com F9 (Unidade de Bay de Prote‡„o de Barras)
                                        if parametros_Trafo['PUPG'] == 'PU/PG':
                                            UB87B = ['F9']
                                        elif parametros_Trafo['PUPG'] == 'PP/PA':
                                            UB87B = ['F9P', 'F9S']
                                        for UBB in UB87B:
                                            tratar_2 = tratar_1.replace('F9', UBB)
                                            if 'FPCn' in tratar_2:
                                                for n_canal in range(1, 3):
                                                    texto_canal = tratar_2[tratar_2.find('FPCn'):-1] + str(n_canal)
                                                    tratar_3 = tratar_2.replace('FPCn', texto_canal)
                                                    gravar_ponto(tratar_3, descricao_1)
                                                    k_trafo += 1
                                            elif 'FPDn' in tratar_2:
                                                for n_canal in range(1, 3):
                                                    texto_canal = tratar_2[tratar_2.find('FPDn'):-1] + str(n_canal)
                                                    tratar_3 = tratar_2.replace('FPDn', texto_canal)
                                                    gravar_ponto(tratar_3, descricao_1)
                                                    k_trafo += 1
                                            else:
                                                gravar_ponto(tratar_2, descricao_1)
                                            k_trafo += 1
                                    elif 'FPCn' in tratar_1:
                                        for n_canal in range(1, 3):
                                            texto_canal = tratar_1[tratar_1.find('FPCn'):-1] + str(n_canal)
                                            tratar_2 = tratar_1.replace('FPCn', texto_canal)
                                            gravar_ponto(tratar_2, descricao_1)
                                            k_trafo += 1
                                    elif 'FPDn' in tratar_1:
                                        for n_canal in range(1, 3):
                                            texto_canal = tratar_1[tratar_1.find('FPDn'):-1] + str(n_canal)
                                            tratar_2 = tratar_1.replace('FPDn', texto_canal)
                                            gravar_ponto(tratar_2, descricao_1)
                                            k_trafo += 1
                                    else:
                                        gravar_ponto(tratar_1, descricao_1)
                                        k_trafo += 1

                        elif abas[plan_index] == 'BARRA':
                            if conf_BT_array:  # Caso exista configura‡„o de V„o de Transferˆncia
                                if tratar[5:8] == 'YDY':  # Caso de ponto de V„o de Transferˆncia
                                    for parametros_BT in conf_BT_array:
                                        cd1 = ('F9' not in tratar) or (
                                            'F9' in tratar and int(parametros_BT['COD'][1]) >= 3)
                                        if '0YDY' in tratar:
                                            tratar_1 = tratar.replace('0YDY', parametros_BT['COD'])
                                            descricao_1 = descricao.replace('0YDY', parametros_BT['COD'])
                                        elif '1YDY' in tratar:
                                            tratar_1 = tratar.replace('1YDY', parametros_BT['COD'])
                                            descricao_1 = descricao.replace('1YDY', parametros_BT['COD'])
                                        tratar_2 = tratar_1.replace('{PNL}', parametros_BT['PNL'])
                                        if cd1:
                                            if 'FCOn' in tratar_2:
                                                for n_canal in range(1, 3):
                                                    texto_canal = tratar_2[tratar_2.find('FCOn'):-1] + str(n_canal)
                                                    tratar_3 = tratar_2.replace('FCOn', texto_canal)
                                                    gravar_ponto(tratar_3, descricao_1)
                                                    k_bt += 1
                                            elif 'FPDn' in tratar_2:
                                                for n_canal in range(1, 3):
                                                    texto_canal = tratar_2[tratar_2.find('FPDn'):-1] + str(n_canal)
                                                    tratar_3 = tratar_2.replace('FPDn', texto_canal)
                                                    gravar_ponto(tratar_3, descricao_1)
                                                    k_bt += 1
                                            elif 'FPCn' in tratar_2:
                                                for n_canal in range(1, 3):
                                                    texto_canal = tratar_2[tratar_2.find('FPCn'):-1] + str(n_canal)
                                                    tratar_3 = tratar_2.replace('FPCn', texto_canal)
                                                    gravar_ponto(tratar_3, descricao_1)
                                                    k_bt += 1
                                            else:
                                                gravar_ponto(tratar_2, descricao_1)
                                                k_bt += 1

                                ###Condi‡”es para processar o ponto de Prote‡„o e Anal¢gicos###
                                # Tenha o tag '0XB' no ID SAGE
                                cd1 = (tratar[4:7] == '0XB')
                                # N„o seja ponto para painel pr¢prio de prote‡„o
                                cd2 = '#PAINEL87B' not in observacao.upper()
                                # N„o tenha 'F2' no ID SAGE (F2 de barras se aplica apenas em Disjuntor e meio, que n„o se aplica a V„o de Transferˆncia)
                                cd3 = 'F2' not in tratar

                                if cd1 * cd2 * cd3:  # Ponto de 87B interno no painel do V„o de Transferˆncia e grandezas Anal¢gicas
                                    for parametros_87B in conf_BT_array:
                                        if parametros_87B['ARR'][:2] == 'BD' or parametros_87B['ARR'] == 'DISJ E MEIO':
                                            if parametros_87B['COD'][
                                               2:4] == 'D2':  # Barras 3 e 4 para caso de Disjuntor 1XD2 (barra partida)
                                                barras = [3, 4]
                                            else:
                                                barras = [1, 2]
                                        elif parametros_87B['ARR'] == 'BPT' or parametros_87B['ARR'] == 'BS':
                                            barras = [1]
                                        if sheet.cell(row=index_linha, column=titulo_dic[
                                            u'OCR (SAGE)']).value == u'OCR_PAS01':  # Pontos Anal¢gicos
                                            for i in barras:
                                                if len(barras) == 1:
                                                    nbarra = 'P'
                                                else:
                                                    nbarra = str(i)
                                                cod = '0' + str(parametros_87B['COD'][
                                                                1:2]) + 'B' + nbarra  # Definir c¢digo da barra
                                                tratar_1 = tratar.replace('0XBY', cod)
                                                descricao_1 = descricao.replace('0XBY', cod)
                                                gravar_ponto(tratar_1, descricao_1)
                                                k_barra += 1
                                        elif parametros_87B[
                                            '87B'] == "Sim":  # Prote‡„o de Barras no V„o de Transferˆncia
                                            if parametros_87B['ARR'] != 'DISJ E MEIO':
                                                if tratar[4:10] == '0XB1/2':  # Ponto Agrupado de Barra1 e Barra 2
                                                    cod = '0' + str(parametros_87B['COD'][1:2])
                                                    tratar_1 = tratar.replace('0X', cod)

                                                    if 'FPCn' in tratar_1:
                                                        for n_canal in range(1, 3):
                                                            texto_canal = tratar_1[tratar_1.find('FPCn'):-1] + str(
                                                                n_canal)
                                                            tratar_2 = tratar_1.replace('FPCn', texto_canal)
                                                            gravar_ponto(tratar_2, descricao_1)
                                                            k_barra += 1
                                                    elif 'FPDn' in tratar_1:
                                                        for n_canal in range(1, 3):
                                                            texto_canal = tratar_1[tratar_1.find('FPDn'):-1] + str(
                                                                n_canal)
                                                            tratar_2 = tratar_1.replace('FPDn', texto_canal)
                                                            gravar_ponto(tratar_2, descricao_1)
                                                            k_barra += 1
                                                    else:
                                                        gravar_ponto(tratar_1, descricao)
                                                        k_barra += 1
                                                else:
                                                    for i in barras:
                                                        cod = '0' + str(parametros_87B['COD'][1:2]) + 'B' + str(
                                                            i)  # c¢digo da Barra
                                                        tratar_1 = tratar.replace('0XBY', cod)
                                                        descricao_1 = descricao.replace('0XBY', cod)
                                                        if 'FPCn' in tratar_1:
                                                            for n_canal in range(1, 3):
                                                                texto_canal = tratar_1[tratar_1.find('FPCn'):-1] + str(n_canal)
                                                                tratar_2 = tratar_1.replace('FPCn', texto_canal)
                                                                gravar_ponto(tratar_2, descricao_1)
                                                                k_barra += 1
                                                        elif 'FPDn' in tratar_1:
                                                            for n_canal in range(1, 3):
                                                                texto_canal = tratar_1[tratar_1.find('FPDn'):-1] + str(n_canal)
                                                                tratar_2 = tratar_1.replace('FPDn', texto_canal)
                                                                gravar_ponto(tratar_2, descricao_1)
                                                                k_barra += 1
                                                        else:
                                                            gravar_ponto(tratar_1, descricao_1)
                                                            k_barra += 1

                            ###Condi‡”es para processar o ponto de Prote‡„o de Barras em painel pr¢prio###
                            # Tenha o tag '0XB' no ID SAGE ou seja ponto referente a Bay Unit de Prote‡„o de Barras
                            cd1 = tratar[4:7] == '0XB' or ('F9' in tratar and tratar[5:8] != 'YDY')
                            # N„o seja ponto anal¢gico (esse tipo de ponto deve ser processado apenas no V„o de Transferˆncia)
                            cd2 = sheet.cell(row=index_linha, column=titulo_dic[u'OCR (SAGE)']).value != u'OCR_PAS01'
                            # Existe configura‡„o de 87B
                            cd3 = conf_P87B_array
                            cd4 = ('#APLICACAO' not in observacao.upper())
                            if cd1 * cd2 * cd3 * cd4:  # Processar pontos referentes a Painel pr¢prio de Prote‡„o de Barras
                                for parametros_87B in conf_P87B_array:
                                    if parametros_87B['ARR'][:2] == 'BD' or parametros_87B['ARR'] == 'DISJ E MEIO':
                                        barras = [1, 2]
                                    elif parametros_87B['ARR'] == 'BPT':
                                        barras = [1]
                                    if 'F9' in tratar:  # Ponto referente a Bay Unit de Prote‡„o de Barras
                                        if parametros_87B['BU'] == 'Sim' and (int(parametros_87B['COD'][
                                                                                      0]) >= 3):  # Caso sejam instaladas Bay Units no painel de Prote‡„o de Barras e
                                            vaos_bu = parametros_87B['VAOS'].split(
                                                '/')  # Gerar array com v„o que ter„o Bay Unit da prote‡„o de barras
                                            for vao in vaos_bu:  # Gerar pontos referente a IED F9 (Bay Unit de 87B) para cada v„o
                                                tratar_1 = tratar.replace('1YDY', vao)
                                                if 'FPCn' in tratar_1:
                                                    for n_canal in range(1, 3):
                                                        texto_canal = tratar_1[tratar_1.find('FPCn'):-1] + str(n_canal)
                                                        tratar_2 = tratar_1.replace('FPCn', texto_canal)
                                                        gravar_ponto(tratar_2, descricao)
                                                        k_barra += 1
                                                elif 'FPDn' in tratar_1:
                                                    for n_canal in range(1, 3):
                                                        texto_canal = tratar_1[tratar_1.find('FPDn'):-1] + str(n_canal)
                                                        tratar_2 = tratar_1.replace('FPDn', texto_canal)
                                                        gravar_ponto(tratar_2, descricao)
                                                        k_barra += 1
                                                else:
                                                    gravar_ponto(tratar_1, descricao)
                                                    k_barra += 1

                                    elif '{PNL}' in tratar:  # Ponto Agrupado de Barra1 e Barra 2 com informa‡„o do Painel
                                        cod = '0' + str(parametros_87B['PNL'][0])
                                        tratar_1 = tratar.replace('0X', cod)
                                        for npnl in range(1, int(parametros_87B[
                                                                     'NPNL']) + 1):  # Gerar ponto para todos os pain‚is de 87B. Ex. 4UA8-1 e 4UA8-2
                                            if int(parametros_87B['NPNL']) > 1:  # Se tiver mais de um painel de 87B
                                                nome_painel = parametros_87B['PNL'] + str(npnl)  # Ex. 4UA81 e 4UA82
                                            else:  # Se houver apenas um painel de 87B n„o usar £ndice de numera‡„o de painel no ID SAGE
                                                nome_painel = parametros_87B['PNL']  # Ex. 4UA8
                                            tratar_2 = tratar_1.replace('{PNL}', nome_painel)
                                            gravar_ponto(tratar_2, descricao)
                                            k_barra += 1
                                    elif tratar[4:10] == '0XB1/2':
                                        if parametros_87B['ARR'] != 'DISJ E MEIO':
                                            cod = '0' + str(parametros_87B['PNL'][0])
                                            tratar_1 = tratar.replace('0X', cod)
                                            if 'FPCn' in tratar_1:
                                                for n_canal in range(1, 3):
                                                    texto_canal = tratar_1[tratar_1.find('FPCn'):-1] + str(n_canal)
                                                    tratar_2 = tratar_1.replace('FPCn', texto_canal)
                                                    gravar_ponto(tratar_2, descricao)
                                                    k_barra += 1
                                            elif 'FPDn' in tratar_1:
                                                for n_canal in range(1, 3):
                                                    texto_canal = tratar_1[tratar_1.find('FPDn'):-1] + str(n_canal)
                                                    tratar_2 = tratar_1.replace('FPDn', texto_canal)
                                                    gravar_ponto(tratar_2, descricao)
                                                    k_barra += 1
                                            else:
                                                gravar_ponto(tratar_1, descricao)
                                                k_barra += 1
                                    else:
                                        if parametros_87B['ARR'] == 'DISJ E MEIO':  # Se for disjuntor e meio
                                            for i in [1, 2]:
                                                i = str(i)
                                                cod = '0' + str(parametros_87B['PNL'][
                                                                    0]) + 'B' + i  # "0"+c¢digo de Tens„o (ex. "4")+"B"+"1" para caso seja F1 ou "2" para caso seja F2
                                                tratar_1 = tratar.replace('0XBY', cod)
                                                tratar_2 = tratar_1.replace('F8', 'F8.%s' % i)
                                                descricao_1 = descricao.replace('0XBY', cod)
                                                if 'FPCn' in tratar_2:
                                                    for n_canal in range(1, 3):
                                                        texto_canal = tratar_2[tratar_2.find('FPCn'):-1] + str(n_canal)
                                                        tratar_3 = tratar_2.replace('FPCn', texto_canal)
                                                        gravar_ponto(tratar_3, descricao_1)
                                                        k_barra += 1
                                                elif 'FPDn' in tratar_2:
                                                    for n_canal in range(1, 3):
                                                        texto_canal = tratar_2[tratar_2.find('FPDn'):-1] + str(n_canal)
                                                        tratar_3 = tratar_2.replace('FPDn', texto_canal)
                                                        gravar_ponto(tratar_3, descricao_1)
                                                        k_barra += 1
                                                else:
                                                    gravar_ponto(tratar_2, descricao_1)
                                                    k_barra += 1
                                        else:  # Se n„o for disjuntor e meio
                                            for i in barras:
                                                cod = '0' + str(parametros_87B['PNL'][0]) + 'B' + str(i)
                                                tratar_1 = tratar.replace('0XBY', cod)
                                                descricao_1 = descricao.replace('0XBY', cod)
                                                if 'FPCn' in tratar_1:
                                                    for n_canal in range(1, 3):
                                                        texto_canal = tratar_1[tratar_1.find('FPCn'):-1] + str(n_canal)
                                                        tratar_2 = tratar_1.replace('FPCn', texto_canal)
                                                        gravar_ponto(tratar_2, descricao_1)
                                                        k_barra += 1
                                                elif 'FPDn' in tratar_1:
                                                    for n_canal in range(1, 3):
                                                        texto_canal = tratar_1[tratar_1.find('FPDn'):-1] + str(n_canal)
                                                        tratar_2 = tratar_1.replace('FPDn', texto_canal)
                                                        gravar_ponto(tratar_2, descricao_1)
                                                        k_barra += 1
                                                else:
                                                    gravar_ponto(tratar_1, descricao_1)
                                                    k_barra += 1

                        elif abas[plan_index] == 'Reator':
                            for parametros_Reator in conf_Reator_array:
                                ###Condi‡”es para processar o ponto###
                                # N„o conste obserna‡„o 'Em caso de Banco Monof cico' ou conste 'Em caso de Banco Monof cico' e "Equipamen." definido como 'Banco Monof.'
                                cd1 = ('#MONOFASICO' not in observacao.upper() or (
                                    '#MONOFASICO' in observacao.upper() and parametros_Reator['EQP'] == 'Banco Monof.'))
                                # N„o conste no TAG :RDP ou conste e "RDP" definido como 'Sim'
                                cd2 = (':RDP' not in tratar) or (':RDP' in tratar and parametros_Reator['RDP'] == 'Sim')
                                # N„o conste obserna‡„o #PASS ou conste #PASS e "PASSSecc" n„o vazio
                                cd3 = ('#PASS' not in observacao.upper()) or (
                                    '#PASS'.upper() in observacao and bool(
                                        parametros_Reator.get('PASSSecc', [None])[0]))
                                # N„o contem no ID 'F9' ou contem 'F9' e "Bay Unit (F9)" definido como 'Sim'
                                cd4 = ('F9' not in tratar) or ('F9' in tratar and parametros_Reator['F9'] == 'Sim')

                                if cd1 * cd2 * cd3 * cd4:
                                    tratar_1 = tratar.replace('0XEY', parametros_Reator['COD'])
                                    descricao_1 = descricao.replace('0XEY', parametros_Reator['COD'])

                                    if 'FPCn' in tratar_1:
                                        for n_canal in range(1, 3):
                                            texto_canal = tratar_1[tratar_1.find('FPCn'):-1] + str(n_canal)
                                            tratar_2 = tratar_1.replace('FPCn', texto_canal)
                                            gravar_ponto(tratar_2, descricao_1)
                                            k_reator += 1
                                    elif 'FPDn' in tratar_1:
                                        for n_canal in range(1, 3):
                                            texto_canal = tratar_1[tratar_1.find('FPDn'):-1] + str(n_canal)
                                            tratar_2 = tratar_1.replace('FPDn', texto_canal)
                                            gravar_ponto(tratar_2, descricao_1)
                                            k_reator += 1
                                    else:
                                        tratar_2 = tratar_1.replace('{PNL}', parametros_Reator['PNL'])
                                        gravar_ponto(tratar_2, descricao_1)
                                        k_reator += 1

                        elif abas[plan_index] == 'T_Terra':
                            for parametros_TT in conf_TT_array:
                                # N„o conste obserna‡„o #PASS ou conste #PASS e "PASSSecc" n„o vazio
                                cd1 = ('#PASS' not in observacao.upper()) or (
                                    '#PASS'.upper() in observacao and bool(parametros_TT.get('PASSSecc', [None])[0]))
                                if cd1:
                                    tratar_1 = tratar.replace('02AY', parametros_TT['COD'])
                                    tratar_1 = tratar_1.replace('0XBY', parametros_TT['COD'])
                                    descricao_1 = descricao.replace('02AY', parametros_TT['COD'])

                                    if 'FPCn' not in tratar_1:
                                        tratar_2 = tratar_1.replace('{PNL}', parametros_TT['PNL'])
                                        gravar_ponto(tratar_2, descricao_1)
                                        k_tt += 1
                                    else:
                                        for n_canal in range(1, 3):
                                            texto_canal = tratar_1[tratar_1.find('FPCn'):-1] + str(n_canal)
                                            tratar_2 = tratar_1.replace('FPCn', texto_canal)
                                            gravar_ponto(tratar_2, descricao_1)
                                            k_tt += 1

                        elif abas[plan_index] == 'Disjuntor':
                            for parametros_vao in parametros:
                                if parametros_vao[
                                    'TIPO'] == 'Trafo':  # Em caso de Trafo pegar arranjo da baixa e da alta
                                    arr = [parametros_vao['ARRH'], parametros_vao['ARRX']]
                                elif parametros_vao[
                                    'TIPO'] == 'Reator':  # Em caso de Reator gravar arranjo BS (Barra Simples)
                                    arr = ['BS'] if parametros_vao['BRM'] == 'Sim' else []
                                elif parametros_vao['TIPO'] == 'BCS':  # Em caso de Banco Capacitor S‚rie
                                    arr = ['BS']
                                elif parametros_vao['TIPO'] == 'TT':  # Em caso de Trafo Terra
                                    arr = []
                                else:
                                    arr = [parametros_vao['ARR']]

                                if parametros_vao['TIPO'] =='BT':
                                    VaoTransf = True
                                else:
                                    VaoTransf = False

                                k_arr = True
                                for arranjo in arr:
                                    # Arranjo diferente de Barra Dupla a 3 chaves e diferente de BCS, ou Arranjo igual a Barra Dupla a 3 chaves ou BCS e N„o contenha no ID ':43:'
                                    cd1 = ((arranjo != 'BD3' and arranjo != 'BS') or (
                                        (arranjo == 'BD3' or arranjo == 'BS') and (':43:' not in tratar)))
                                    # Arranjo diferente de DISJ E MEIO ou Arranjo igual DISJ E MEIO e N„o contenha no ID ':43:'
                                    cd2 = (
                                        arranjo != 'DISJ E MEIO' or (arranjo == 'DISJ E MEIO' and ':43:' not in tratar))
                                    # N„o conste observa‡„o #DISJEMEIO ou conste 'Para arranjos disjuntor e meio.' e "ARRANJO" definido como 'DISJ E MEIO'
                                    cd3 = ('#DISJEMEIO' not in observacao.upper() or (
                                        '#DISJEMEIO' in observacao.upper() and arranjo == 'DISJ E MEIO'))

                                    cd4 = ('#MONOPOLAR' not in observacao.upper() or (
                                        '#MONOPOLAR' in observacao.upper() and parametros_vao['79'] == 'MONO/TRI'))

                                    # Caso Tenha F9 em ID
                                    if 'F9' not in tratar:
                                        cd5 = True
                                    elif parametros_vao['COD'] == 'BCS':
                                        cd5 = False
                                    elif parametros_vao.get('F9', False):
                                        if parametros_vao['F9'] == 'Sim':
                                            cd5 = True
                                        else:
                                            cd5 = False
                                    else:
                                        cd5 = False
                                        # N„o se trate de ponto de Falha Sele‡„o Prote‡„o Intr¡nseca ou se trate deste ponto e seja um Trafo
                                    cd6 = ('FSPI' not in tratar) or (
                                        'FSPI' in tratar and parametros_vao['TIPO'] == 'Trafo')
                                    # N„o se trate de ponto de PASS ou se trate deste ponto e seja PASS (defini‡„o de cƒmaras na parametriza‡„o)
                                    cd7 = ('#PASS' not in observacao) or (
                                        '#PASS' in observacao and bool(parametros_vao.get('PASSSecc', [None])[0]))
                                    # N„o se trate de ponto Falha disjuntor de F3 ou se trate deste ponto e n„o seja v„o de Linha
                                    cd8 = ('F3:FLDI' not in tratar) or (
                                        'F3:FLDI' in tratar and parametros_vao['TIPO'] != 'LT')
                                    # N„o se trate de ponto Trip Discordƒncia de Polos convenciol ou se trate deste ponto e n„o seja PASS (defini‡„o de cƒmaras na parametriza‡„o)
                                    cd9 = ('00:APDP' not in tratar) or (
                                        '00:APDP' in tratar and not bool(parametros_vao.get('PASSSecc', [None])[0]))
                                    # N„o se trate de ponto Baixa Press„o Sist. Extin‡„o SF6 1.Grau convenciol ou se trate deste ponto e n„o seja PASS (defini‡„o de cƒmaras na parametriza‡„o)
                                    cd10 = (not tratar.endswith('PBSP')) or (
                                        tratar.endswith('PBSP') and not bool(parametros_vao.get('PASSSecc', [None])[0]))
                                    # N„o se trate de ponto Baixa Press„o Sist. Extin‡„o SF6 2.Grau convenciol ou se trate deste ponto e n„o seja PASS (defini‡„o de cƒmaras na parametriza‡„o)
                                    cd11 = ('00:PBSS' not in tratar) or (
                                        '00:PBSS' in tratar and not bool(parametros_vao.get('PASSSecc', [None])[0]))

                                    cd12 = ('Aplic vel a Disjuntor de transferˆncia' not in observacao or
                                            ('Aplic vel a Disjuntor de transferˆncia' in observacao and VaoTransf))
                                    linha69 = tratar_1[5] == '2' and parametros_vao['TIPO'] == 'LT'

                                    if cd1 * cd2 * cd3 * cd4 * cd5 * cd6 * cd7 * cd8 * cd9 * cd10 * cd11 * cd12:
                                        if k_arr:  # Arranjo a ser processado ‚ o da posi‡„o 0 de "arr"
                                            cod_disj = parametros_vao['COD'][1:]
                                            k_arr = False
                                        elif arranjo == arr[
                                            1]:  # Arranjo a ser processado ‚ o da posi‡„o 1 de "arr" (apenas em caso de TRAFO)
                                            cod_disj = parametros_vao['PNLX'][0] + parametros_vao['COD'][2:]
                                        tag_disj = '1' + cod_disj
                                        tratar_1 = tratar.replace('1YYY', tag_disj)
                                        if linha69 and ':UC1' in tratar_1:
                                            tratar_1 = tratar_1.replace(':UC1', ':F3')
                                        descricao_1 = descricao.replace('1YYY', tag_disj)
                                        #print(parametros_vao.get('PASSCam', [None]))
                                        if parametros_vao.get('PASSCam', [None])[0] and ':Z' in tratar_1:
                                            for cam in parametros_vao['PASSCam']:
                                                tratar_2 = tratar_1.replace(':Z', ':{}'.format(cam))
                                                gravar_ponto(tratar_2, descricao_1)
                                                k_52 += 1
                                        elif 'UC1' in tratar_1 and '(25)' in descricao_1 and len(arr)>1:
                                            if arranjo == arr[1]:
                                                tratar_2 = tratar_1.replace('UC1','UC2')
                                                gravar_ponto(tratar_2, descricao_1)
                                                k_52 += 1
                                        else:
                                            gravar_ponto(tratar_1, descricao_1)
                                            k_52 += 1

                        elif abas[plan_index] == 'Secc':
                            for parametros_vao in parametros:

                                if parametros_vao[
                                    'TIPO'] == 'Trafo':  # Em caso de Trafo pegar arranjo da baixa e da alta
                                    arr = [parametros_vao['ARRH'], parametros_vao['ARRX']]
                                elif parametros_vao['TIPO'] == 'BT':  # Em caso de V„o de Transferˆncia
                                    arr = ['BT']
                                elif parametros_vao[
                                    'TIPO'] == 'Reator':  # Em caso de Reator gravar arranjo BS (Barra Simples)
                                    arr = ['BS']
                                elif parametros_vao['TIPO'] == 'TT':  # Em caso de Trafo Terra gravar arranjo TT
                                    arr = ['TT']
                                elif parametros_vao['TIPO'] == 'BCS':  # Em caso de Trafo Terra gravar arranjo TT
                                    arr = ['BCS']
                                else:
                                    arr = [parametros_vao['ARR']]

                                karr = True
                                for arranjo in arr:
                                    if karr:  # Testa se ‚ a primeira vez que passa no for de arranjo
                                        cod_secc = parametros_vao['COD'][1:]
                                        karr = False
                                    elif arranjo == arr[
                                        1]:  # Arranjo a ser processado ‚ o da posi‡„o 1 de "arr" (apenas em caso de TRAFO)
                                        cod_secc = parametros_vao['PNLX'][0] + parametros_vao['COD'][2:]

                                    tag_raiz_secc = '3' + cod_secc
                                    # Se for m¢dulo isolado a SF6 e se tratar de Falta Tens„o Comando ou Falta Alimenta‡„o CA Motor
                                    if parametros_vao.get('PASSSecc', [None])[0] and (
                                                    'FTCO' in tratar or 'FCAM' in tratar) and (not 'Z1/Zn' in tratar):
                                        for nsecc in parametros_vao['PASSSecc']:
                                            tag_secc = tag_raiz_secc + '-' + str(nsecc)
                                            tratar_1 = tratar.replace('3YYY-Z', tag_secc)
                                            gravar_ponto(tratar_1, descricao)
                                            k_89 += 1
                                    elif not 'Z1/Zn' in tratar:
                                        for nsecc in [1, 2, 4, 5, 6, 7, 8, 9]:
                                            if nsecc == 1 and (arranjo[0:2] == 'BD' or arranjo == 'BT'):
                                                tag_secc = tag_raiz_secc + '-' + '1'
                                                tratar_1 = tratar.replace('3YYY-Z', tag_secc)
                                                descricao_1 = descricao.replace('3YYY-Z', tag_secc)
                                                gravar_ponto(tratar_1, descricao_1)
                                                k_89 += 1
                                            elif nsecc == 2 and (arranjo[0:2] == 'BD' or arranjo == 'BT'):
                                                tag_secc = tag_raiz_secc + '-' + '2'
                                                tratar_1 = tratar.replace('3YYY-Z', tag_secc)
                                                descricao_1 = descricao.replace('3YYY-Z', tag_secc)
                                                gravar_ponto(tratar_1, descricao_1)
                                                k_89 += 1
                                            elif nsecc == 4 and (
                                                                arranjo == 'BPT' or arranjo == 'BD5' or arranjo == 'DISJ E MEIO'):
                                                tag_secc = tag_raiz_secc + '-' + '4'
                                                tratar_1 = tratar.replace('3YYY-Z', tag_secc)
                                                descricao_1 = descricao.replace('3YYY-Z', tag_secc)
                                                gravar_ponto(tratar_1, descricao_1)
                                                k_89 += 1
                                            elif nsecc == 5 and (arranjo == 'BPT' or arranjo[
                                                                                     0:2] == 'BD' or arranjo == 'DISJ E MEIO'):
                                                tag_secc = tag_raiz_secc + '-' + '5'
                                                tratar_1 = tratar.replace('3YYY-Z', tag_secc)
                                                descricao_1 = descricao.replace('3YYY-Z', tag_secc)
                                                gravar_ponto(tratar_1, descricao_1)
                                                k_89 += 1
                                            elif nsecc == 6 and (
                                                                    arranjo == 'BPT' or arranjo == 'BD4' or arranjo == 'BD5' or arranjo == 'BCS'):
                                                tag_secc = tag_raiz_secc + '-' + '6'
                                                tratar_1 = tratar.replace('3YYY-Z', tag_secc)
                                                descricao_1 = descricao.replace('3YYY-Z', tag_secc)
                                                gravar_ponto(tratar_1, descricao_1)
                                                k_89 += 1
                                            elif nsecc == 7 and parametros_vao['COD'][2] != 'T' and (
                                                                    arranjo == 'BPT' or arranjo[
                                                                                        0:2] == 'BD' or arranjo == 'DISJ E MEIO' or arranjo == 'BCS'):
                                                if arranjo[0:3] != 'BD3' and arranjo != 'BCS':
                                                    tag_secc = tag_raiz_secc + '-' + '7'
                                                    tratar_1 = tratar.replace('3YYY-Z', tag_secc)
                                                    descricao_1 = descricao.replace('3YYY-Z', tag_secc)
                                                    gravar_ponto(tratar_1, descricao_1)
                                                    k_89 += 1
                                                else:
                                                    for secbd3 in ['A', 'B']:
                                                        tag_secc = tag_raiz_secc + '-' + '7' + secbd3
                                                        tratar_1 = tratar.replace('3YYY-Z', tag_secc)
                                                        descricao_1 = descricao.replace('3YYY-Z', tag_secc)
                                                        gravar_ponto(tratar_1, descricao_1)
                                                        k_89 += 1
                                            elif nsecc == 8 and (
                                                                    arranjo == 'DISJ E MEIO' or arranjo == 'BS' or arranjo == 'TT' or arranjo == 'BCS'):
                                                tag_secc = tag_raiz_secc + '-' + '8'
                                                tratar_1 = tratar.replace('3YYY-Z', tag_secc)
                                                descricao_1 = descricao.replace('3YYY-Z', tag_secc)
                                                gravar_ponto(tratar_1, descricao_1)
                                                k_89 += 1
                                            elif nsecc == 9 and (arranjo == 'TT' or arranjo == 'BCS'):
                                                tag_secc = tag_raiz_secc + '-' + '9'
                                                tratar_1 = tratar.replace('3YYY-Z', tag_secc)
                                                descricao_1 = descricao.replace('3YYY-Z', tag_secc)
                                                gravar_ponto(tratar_1, descricao_1)
                                                k_89 += 1

                        elif abas[plan_index] == 'B_CAP':
                            for parametros_Bcap in conf_BCap_array:
                                # N„o contem no ID 'F9' ou contem 'F9' e segunda casa do c¢digo da linha (ex. N£mero 4 de 04H1) for maior ou igual a 3 (maior ou igual a 138kV)
                                cd1 = ('F9' not in tratar) or ('F9' in tratar and int(parametros_Bcap['COD'][1]) >= 3)
                                # N„o conste no TAG :RDP ou conste e "RDP" definido como 'Sim'
                                cd2 = (':RDP' not in tratar) or (':RDP' in tratar and parametros_Bcap['RDP'] == 'Sim')
                                # N„o contem no ID 'F9' ou contem 'F9' e "Bay Unit (F9)" definido como 'Sim'
                                cd3 = ('F9' not in tratar) or ('F9' in tratar and parametros_Bcap['F9'] == 'Sim')

                                if cd1 * cd2 * cd3:
                                    tratar_1 = tratar.replace('0XHY', parametros_Bcap['COD'])
                                    descricao_1 = descricao.replace('0XHY', parametros_Bcap['COD'])
                                    '''if 'FPCn' not in tratar_1:
                                        tratar_2 = tratar_1.replace('{PNL}', parametros_Bcap['PNL'])
                                        gravar_ponto(tratar_2, descricao_1)
                                        k_bcap += 1
                                    else:'''
                                    if 'FPCn' in tratar_1:
                                        for n_canal in range(1, 3):
                                            texto_canal = tratar_1[tratar_1.find('FPCn'):-1] + str(n_canal)
                                            tratar_2 = tratar_1.replace('FPCn', texto_canal)
                                            gravar_ponto(tratar_2, descricao_1)
                                            k_bcap += 1
                                    elif 'FPDn' in tratar_1:
                                        for n_canal in range(1, 3):
                                            texto_canal = tratar_1[tratar_1.find('FPDn'):-1] + str(n_canal)
                                            tratar_2 = tratar_1.replace('FPDn', texto_canal)
                                            gravar_ponto(tratar_2, descricao_1)
                                            k_bcap += 1
                                    else:
                                        tratar_2 = tratar_1.replace('{PNL}', parametros_Bcap['PNL'])
                                        gravar_ponto(tratar_2, descricao_1)
                                        k_bcap += 1

                        elif abas[plan_index] == 'BCS':
                            for parametros_BCS in conf_BCS_array:
                                tratar_1 = tratar.replace('0XHY', parametros_BCS['COD'])
                                '''if 'FPCn' not in tratar_1:
                                    tratar_2 = tratar_1.replace('{PNL}', parametros_BCS['PNL'])
                                    gravar_ponto(tratar_2, descricao)
                                    k_bcs += 1
                                else:'''
                                if 'FPCn' in tratar_1:
                                    for n_canal in range(1, 3):
                                        texto_canal = tratar_1[tratar_1.find('FPCn'):-1] + str(n_canal)
                                        tratar_2 = tratar_1.replace('FPCn', texto_canal)
                                        gravar_ponto(tratar_2, descricao)
                                        k_bcs += 1
                                elif 'FPDn' in tratar_1:
                                    for n_canal in range(1, 3):
                                        texto_canal = tratar_1[tratar_1.find('FPDn'):-1] + str(n_canal)
                                        tratar_2 = tratar_1.replace('FPDn', texto_canal)
                                        gravar_ponto(tratar_2, descricao)
                                        k_bcs += 1
                                else:
                                    tratar_2 = tratar_1.replace('{PNL}', parametros_BCS['PNL'])
                                    gravar_ponto(tratar_2, descricao)
                                    k_bcs += 1

                        elif abas[plan_index] == 'ECE':
                            for parametros_ECE in conf_ECE_array:
                                tratar_1 = tratar.replace('0XBY', parametros_ECE['COD'])
                                '''if 'FPCn' not in tratar_1:
                                    tratar_2 = tratar_1.replace('{PNL}', parametros_ECE['PNL'])
                                    gravar_ponto(tratar_2, descricao)
                                    k_ece += 1
                                else:'''
                                if 'FPCn' in tratar_1:
                                    for n_canal in range(1, 3):
                                        texto_canal = tratar_1[tratar_1.find('FPCn'):-1] + str(n_canal)
                                        tratar_2 = tratar_1.replace('FPCn', texto_canal)
                                        gravar_ponto(tratar_2, descricao)
                                        k_ece += 1
                                elif 'FPDn' in tratar_1:
                                    for n_canal in range(1, 3):
                                        texto_canal = tratar_1[tratar_1.find('FPDn'):-1] + str(n_canal)
                                        tratar_2 = tratar_1.replace('FPDn', texto_canal)
                                        gravar_ponto(tratar_2, descricao)
                                        k_ece += 1
                                else:
                                    tratar_2 = tratar_1.replace('{PNL}', parametros_ECE['PNL'])
                                    gravar_ponto(tratar_2, descricao)
                                    k_ece += 1

                        elif abas[plan_index] == 'CS':
                            for parametros_CS in conf_CS_array:
                                tratar_1 = tratar.replace('0XKY', parametros_CS['COD'])
                                '''if 'FPCn' not in tratar_1:
                                    tratar_2 = tratar_2.replace('{PNL}', parametros_CS['PNL'])
                                    gravar_ponto(tratar_2, descricao)
                                    k_cs += 1
                                else:'''
                                if 'FPCn' in tratar_1:
                                    for n_canal in range(1, 3):
                                        texto_canal = tratar_1[tratar_1.find('FPCn'):-1] + str(n_canal)
                                        tratar_2 = tratar_1.replace('FPCn', texto_canal)
                                        gravar_ponto(tratar_2, descricao)
                                        k_cs += 1
                                elif 'FPDn' in tratar_1:
                                    for n_canal in range(1, 3):
                                        texto_canal = tratar_1[tratar_1.find('FPDn'):-1] + str(n_canal)
                                        tratar_2 = tratar_1.replace('FPDn', texto_canal)
                                        gravar_ponto(tratar_2, descricao)
                                        k_cs += 1
                                elif '{PNL}' in tratar_1:
                                    tratar_2 = tratar_1.replace('{PNL}', parametros_CS['PNL'])
                                    gravar_ponto(tratar_2, descricao)
                                    k_cs += 1
                                else:
                                    gravar_ponto(tratar_1, descricao)
                                    k_cs += 1

                        elif abas[plan_index] == 'Prep. Reen.':
                            if tratar.split(':')[1] == 'mmmmnnn':  # Sitema de Regula‡„o de Tens„o
                                for parametros_SR in conf_SR_array:
                                    tratar_1 = tratar.replace('mmmmnnn', '%s%s' % (
                                        parametros_SR['COD'], tensao_dic[parametros_SR['TENSAO']]))
                                    if 'FPCn' in tratar_1:
                                        for n_canal in range(1, 3):
                                            texto_canal = tratar_1[tratar_1.find('FPCn'):-1] + str(n_canal)
                                            tratar_2 = tratar_1.replace('FPCn', texto_canal)
                                            gravar_ponto(tratar_2, descricao)
                                            k_sr += 1
                                    else:
                                        tratar_2 = tratar_1.replace('{PNL}', parametros_SR['PNL'])
                                        gravar_ponto(tratar_2, descricao)
                                        k_sr += 1
                            else:
                                for parametros_PR in conf_PR_array:
                                    tratar_1 = tratar.replace('mmmm', parametros_PR['COD'])
                                    gravar_ponto(tratar_1, descricao)
                                    k_pr += 1

                                    #                        elif sheet.name=='CE':
                                    #                            for parametros_CE in conf_CE_array:
                                    #                                if 'FCOn' in tratar:
                                    #                                    for n_canal in range(1,3):
                                    #                                        texto_canal = tratar[tratar.find('FCOn'):-1] + str(n_canal)
                                    #                                        tratar_1 = tratar.replace(FCOn',texto_canal)
                                    #                                        gravar_ponto(tratar_1, descricao_1)
                                    #                                        k_ce+=1
                                    #                                else:
                                    #                                    tratar_1 = tratar.replace('0XBY',parametros_CE[0])
                                    #                                   tratar_2 = tratar_1.replace('{PNL}',parametros_CE[1])
                                    #                                    gravar_ponto(tratar_2, descricao)
                                    #                                    k_ce+=1

                        elif abas[plan_index] == 'SAs':
                            # Definir c¢digo Tens„o CA
                            if conf_SA['VCA'] == '220Vca':
                                TensaoCA = '2'
                            elif conf_SA['VCA'] == '380Vca':
                                TensaoCA = '3'
                            else:
                                TensaoCA = 'X'

                            # Definir c¢digo Tens„o CC
                            if conf_SA['VCC'] == '125Vcc':
                                TensaoCC = '8'
                            elif conf_SA['VCC'] == '250Vcc':
                                TensaoCC = '9'
                            else:
                                TensaoCC = 'X'

                            if tratar[4:7] == 'X52':  # Pontos de Disjuntores CA
                                tratar_1 = tratar.replace('X', TensaoCA)
                                descricao_1 = descricao.replace('X', TensaoCA)

                                for DisjCA in conf_SA['DJCA'].split('/'):
                                    tratar_2 = tratar_1.replace('YY', DisjCA)
                                    descricao_2 = descricao_1.replace('YY', DisjCA)
                                    gravar_ponto(tratar_2, descricao_2)
                                    k_sas += 1

                            elif tratar[4:7] == 'X72':  # Pontos de Disjuntores CC
                                tratar_1 = tratar.replace('X', TensaoCC)
                                descricao_1 = descricao.replace('X', TensaoCC)

                                for DisjCC in conf_SA['DJCC'].split('/'):
                                    tratar_2 = tratar_1.replace('YYY', DisjCC)
                                    tratar_2 = tratar_2.replace('YY', DisjCC)
                                    descricao_2 = descricao_1.replace('YYY', DisjCC)
                                    descricao_2 = descricao_2.replace('YY', DisjCC)
                                    gravar_ponto(tratar_2, descricao_2)
                                    k_sas += 1

                            elif tratar[7:9] == 'BY':  # Pontos de Barra
                                for BarrasCA in conf_SA['BSCA'].split('/'):
                                    if 'CC' not in descricao:
                                        tratar_1 = tratar.replace('X', TensaoCA)
                                        tratar_2 = tratar_1.replace('BY', BarrasCA)
                                        gravar_ponto(tratar_2, descricao)
                                        k_sas += 1
                                for BarrasCC in conf_SA['BSCC'].split('/'):
                                    if 'CA' not in descricao:
                                        tratar_1 = tratar.replace('X', TensaoCC)
                                        tratar_2 = tratar_1.replace('BY', BarrasCC)
                                        gravar_ponto(tratar_2, descricao)
                                        k_sas += 1

                            elif tratar[7:9] == 'GY':  # Pontos de Grupo Gerador de Emergˆncia
                                for GGE in conf_SA['DJCA'].split('/'):
                                    if GGE[0] == 'G':
                                        tratar_1 = tratar.replace('X', TensaoCA)
                                        descricao_1 = descricao.replace('X', TensaoCA)
                                        tratar_2 = tratar_1.replace('GY', GGE)
                                        descricao_2 = descricao_1.replace('GY', GGE)
                                        gravar_ponto(tratar_2, descricao_2)
                                        k_sas += 1

                            elif tratar[7:9] == 'RY':  # Pontos de Retificador
                                for RET in conf_SA['DJCA'].split('/'):
                                    if RET[0] == 'R':
                                        tratar_1 = tratar.replace('X', TensaoCA)
                                        descricao_1 = descricao.replace('X', TensaoCA)
                                        tratar_2 = tratar_1.replace('RY', RET)
                                        descricao_2 = descricao_1.replace('RY', RET)
                                        gravar_ponto(tratar_2, descricao_2)
                                        k_sas += 1

                            elif '{PNL' in tratar:  # Pontos de Pain‚is
                                if '{PNL}' in tratar:  # Pontos geral de Pain‚is
                                    for painel in conf_SA['PNL'].split('/') + conf_SA['PNLSA'].split('/'):
                                        tratar_1 = tratar.replace('{PNL}', painel)
                                        gravar_ponto(tratar_1, descricao)
                                        k_sas += 1
                                elif '{PNLUA}' in tratar:  # Pontos de Pain‚is de UA
                                    tratar_1 = tratar.replace('{PNLUA}', conf_SA['PNL'])
                                    gravar_ponto(tratar_1, descricao)
                                    k_sas += 1
                                elif '{PNLSA}' in tratar:  # Pontos de Pain‚is de Serv. Aux.
                                    for painel in conf_SA['PNLSA'].split('/'):
                                        tratar_1 = tratar.replace('{PNLSA}', painel)
                                        gravar_ponto(tratar_1, descricao)
                                        k_sas += 1
                            elif 'FPCn' in tratar:
                                for n_canal in range(1, 3):
                                    texto_canal = tratar[tratar.find('FPCn'):-1] + str(n_canal)
                                    tratar_1 = tratar.replace('FPCn', texto_canal)
                                    gravar_ponto(tratar_1, descricao)
                                    k_sas += 1
                            else:
                                gravar_ponto(tratar, descricao)
                                k_sas += 1

                                # #                        elif sheet.name=='CE':
                                # #                            tratar = tratar.replace(0XQY',conf_dic['CE'])
                                # #                            descricao = descricao.replace(0XQY',conf_dic['CE'])
                                # #                            k_ce += 1



                                # 0 - LP N2
    saida = [saida_array,
             # 1 0 - Contador LT
             [k_lt,
              # 1 1 - Contador Disjuntor
              k_52,
              # 1 2 - Contador Seccionadoras
              k_89,
              # 1 3 - Contador Trafo
              k_trafo,
              # 1 4 - Contador BT
              k_bt,
              # 1 5 - Contador Reator
              k_reator,
              # 1 6 - Contador Trato Terra
              k_tt,
              # 1 7 - Contador Painel SAGE e Bastidor de Rede
              k_sd,
              # 1 8 - Contador Barra
              k_barra,
              # 1 9 - Contador Banco Capacitor
              k_bcap,
              # 1 10 - Contador Banco Capacitor S‚rie
              k_bcs,
              # 1 11 - Contador Servi‡os Auxiliares
              k_sas,
              # 1 12 - Contador ECE
              k_ece,
              # 1 13 - Contador Compensador S¡ncrono
              k_cs,
              # 1 14 - Contador Prepara‡„o Reenergiza‡„o
              k_pr,
              # 1 15 - Contador Sistema Regula‡„o
              k_sr,
              # 1 16 - Contador Do Painel de Interface
              k_Pint]
             ]

    saida = painelLT69(saida)
    saida[0].sort(key=lambda REG: REG[0])

    return saida
