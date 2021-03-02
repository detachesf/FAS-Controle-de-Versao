# -*- coding: cp860 -*-

dados = '''
Vers„o 2.0.12
Atualiza‡„o do programa: 16/10/2020
Funcionalidade de checagem entre LPs ou de LP com LP gerada por arquivo Arq_Conf. 
'''

import FASgtkui
import pickle
import os.path
from operator import itemgetter
from difflib import get_close_matches
from traceback import print_exc
from sys import stdout
import threading
import time
import gi
gi.require_version("Gtk", "3.0")
from gi.repository import GObject

try:
    from openpyxl import load_workbook, cell
except:
    FASgtkui.mensagem_erro('Erro', 'M¢dulo openpyxl n„o instalado')
try:
    from lp_lib.LP import gerarlp
except:
    FASgtkui.mensagem_erro('Erro', 'Arquivo "LP.py" deve estar no mesmo diret¢rio "lp_lib"')
try:
    import xlsxwriter
except:
    FASgtkui.mensagem_erro('Erro', 'M¢dulo XlsxWriter n„o instalado')
try:
    from lp_lib.func import linhaInicialETitulos
except:
    FASgtkui.mensagem_erro('Erro', 'Arquivo "func.pyc" deve estar no diret¢rio "lp_lib"')


def checar(LP_Padrao='', LP_Editado='', planilha='', Arq_Conf='',
           array_base='', Diretorio_Padrao=''):  # array_base s¢ ser  preenchido em caso de compara‡„o

    """

    :type array_base: object
    """
    gerararquivo = True

    # ----------Declara‡„o de Vari veis----------#
    array_padrao = []
    array_validar = []
    diferenca_array = []
    pfalta_array = []
    array_validar_endereco = []
    endN3Teleass = []
    endN3 = []
    endduplicado_array = []
    sugestao_ID_array = []
    k_inc = 0
    k_falta = 0
    k_enddupl = 0

    # ----------Ler LP Validar----------#

    LP_Validar = LP_Editado  # Ler defini‡„o do arquivo de LP padr„o
    Nome_Planilha = planilha

    try:
        book = load_workbook(LP_Validar, data_only=True)  # Abrir arquivo de LP a ser validada
    except:
        GObject.idle_add(FASgtkui.mensagem_erro,'Erro', 'Arquivo ' + LP_Validar + ' n„o encontrado')
        time.sleep(1)
        while FASgtkui.mensagem_erro_dialog.get_visible() == True:
            time.sleep(1)

    sheet = book[Nome_Planilha]  # Abrir planilhas
    try:
        # Lˆ planilha e recebe a linha onde come‡a a LP (aqui usando linha inicial e n„o o dicion rio de t¡tulos)
        li, titulo_dic = linhaInicialETitulos(LP_Validar, Nome_Planilha)
        if li < 0:  # Se for um n£mero negativo ent„o n„o foi encontrado "ID (SAGE)" na lista
            raise NameError('Arquivo especificado n„o possui coluna com t¡tulo "ID (SAGE)".')
        for index_linha in range(li, sheet.max_row + 1):  # Ler colunas da linha selecionada ao final
            if sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 2']['ID (SAGE)']).value != '' and \
                    sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 2']['ID (SAGE)']).value != 'CGS' and \
                    sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 2']['ID (SAGE)']).value != 'PDS' and \
                    sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 2']['ID (SAGE)']).value != 'PAS':
                try:  # Caso a descri‡„o do campo 6 seja "TELA"
                    # 0 - ID SAGE
                    array_coletado = [
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 2']['ID (SAGE)']).value),
                        # N2
                        # 1 - OCR
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 2']['OCR (SAGE)']).value),
                        # 1 - DESCRI€ŽO
                        str(sheet.cell(row=index_linha,
                                       column=titulo_dic['CHESF - N‹VEL 2']['DESCRI€ŽO']).value).strip(),
                        # 2 - TIPO
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 2']['TIPO']).value),
                        # 3 - COMANDO
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 2']['COMANDO']).value),
                        # 4 - MEDI€ŽO
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 2']['MEDI€ŽO']).value),
                        # 5 - TELA
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 2']['TELA']).value),
                        # 6 - LISTA DE ALARMES
                        str(sheet.cell(row=index_linha,
                                       column=titulo_dic['CHESF - N‹VEL 2']['LISTA DE ALARMES']).value),
                        # 7 - SOE
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 2']['SOE']).value),
                        # TELEASSIST‰NCIA N3
                        # 8 - OCR
                        str(sheet.cell(row=index_linha,
                                       column=titulo_dic['CHESF - TELEASSIST‰NCIA N3']['OCR (SAGE)']).value),
                        # 9 - COMANDO
                        str(sheet.cell(row=index_linha,
                                       column=titulo_dic['CHESF - TELEASSIST‰NCIA N3']['COMANDO']).value),
                        # 10 - MEDI€ŽO
                        str(sheet.cell(row=index_linha,
                                       column=titulo_dic['CHESF - TELEASSIST‰NCIA N3']['MEDI€ŽO']).value),
                        # 11 - LISTA DE ALARMES
                        str(sheet.cell(row=index_linha,
                                       column=titulo_dic['CHESF - TELEASSIST‰NCIA N3']['LISTA DE ALARME']).value),
                        # 12 - SOE
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - TELEASSIST‰NCIA N3']['SOE']).value),
                        # 13 - OBSERVA€ŽO
                        str(sheet.cell(row=index_linha,
                                       column=titulo_dic['CHESF - TELEASSIST‰NCIA N3']['OBSERVA€ŽO']).value),
                        # 15 - AGRUPAMENTO
                        str(sheet.cell(row=index_linha,
                                       column=titulo_dic['CHESF - TELEASSIST‰NCIA N3']['AGRUPAMENTO']).value),
                        # N3
                        # 16 - OCR (SAGE)
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 3']['OCR (SAGE)']).value),
                        # 17 - COMANDO
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 3']['COMANDO']).value),
                        # 18 - MEDI€ŽO
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 3']['MEDI€ŽO']).value),
                        # 19 - LISTA DE ALARMES
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 3']['LISTA DE ALARME']).value),
                        # 20 - SOE
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 3']['SOE']).value),
                        # 21 - OBSERVA€ŽO
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 3']['OBSERVA€ŽO']).value),
                        # 22 - AGRUPAMETO
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 3']['AGRUPAMENTO']).value),
                        # ONS
                        # 23 - ITEM
                        str(sheet.cell(row=index_linha, column=titulo_dic['ONS']['ITEM']).value),
                        # 24 - DESCRI€ŽO
                        str(sheet.cell(row=index_linha, column=titulo_dic['ONS']['DESCRI€ŽO']).value),
                        # LIMITES OPERACIONAIS
                        # 25 - LIU
                        str(sheet.cell(row=index_linha, column=titulo_dic['LIMITES OPERACIONAIS']['LIU']).value),
                        # 26 - LIE
                        str(sheet.cell(row=index_linha, column=titulo_dic['LIMITES OPERACIONAIS']['LIE']).value),
                        # 27 - LIA
                        str(sheet.cell(row=index_linha, column=titulo_dic['LIMITES OPERACIONAIS']['LIA']).value),
                        # 28 - LSA
                        str(sheet.cell(row=index_linha, column=titulo_dic['LIMITES OPERACIONAIS']['LSA']).value),
                        # 29 - LSE
                        str(sheet.cell(row=index_linha, column=titulo_dic['LIMITES OPERACIONAIS']['LSE']).value),
                        # 30 - LSU
                        str(sheet.cell(row=index_linha, column=titulo_dic['LIMITES OPERACIONAIS']['LSU']).value),
                        # 31 - BNDMO
                        str(sheet.cell(row=index_linha, column=titulo_dic['LIMITES OPERACIONAIS']['BNDMO']).value),
                        # 32 - OBSERVA€™ES
                        str(sheet.cell(row=index_linha,
                                       column=titulo_dic['LIMITES OPERACIONAIS']['OBSERVA€™ES']).value),
                        # 33 - ENDERE€O N3 Teleassistˆncia
                        str(sheet.cell(row=index_linha,
                                       column=titulo_dic['CHESF - TELEASSIST‰NCIA N3']['ENDERE€O']).value),
                        # 34 - ENDERE€O N3
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 3']['ENDERE€O']).value)]
                except:  # Caso a descri‡„o do campo 6 seja "ANUNCIADOR"
                    # 0 - ID SAGE
                    array_coletado = [
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 2']['ID (SAGE)']).value),
                        # N2
                        # 1 - OCR
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 2']['OCR (SAGE)']).value),
                        # 2 - DESCRI€ŽO
                        str(sheet.cell(row=index_linha,
                                       column=titulo_dic['CHESF - N‹VEL 2']['DESCRI€ŽO']).value).strip(),
                        # 3 - TIPO
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 2']['TIPO']).value),
                        # 4 - COMANDO
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 2']['COMANDO']).value),
                        # 5 - MEDI€ŽO
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 2']['MEDI€ŽO']).value),
                        # 6 - TELA
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 2']['ANUNCIADOR']).value),
                        # 7 - LISTA DE ALARMES
                        str(sheet.cell(row=index_linha,
                                       column=titulo_dic['CHESF - N‹VEL 2']['LISTA DE ALARMES']).value),
                        # 8 - SOE
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 2']['SOE']).value),
                        # TELEASSIST‰NCIA N3
                        # 9 - OCR
                        str(sheet.cell(row=index_linha,
                                       column=titulo_dic['CHESF - TELEASSIST‰NCIA N3']['OCR (SAGE)']).value),
                        # 10 - COMANDO
                        str(sheet.cell(row=index_linha,
                                       column=titulo_dic['CHESF - TELEASSIST‰NCIA N3']['COMANDO']).value),
                        # 11 - MEDI€ŽO
                        str(sheet.cell(row=index_linha,
                                       column=titulo_dic['CHESF - TELEASSIST‰NCIA N3']['MEDI€ŽO']).value),
                        # 12 - LISTA DE ALARMES
                        str(sheet.cell(row=index_linha,
                                       column=titulo_dic['CHESF - TELEASSIST‰NCIA N3']['LISTA DE ALARME']).value),
                        # 13 - SOE
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - TELEASSIST‰NCIA N3']['SOE']).value),
                        # 14 - OBSERVA€ŽO
                        str(sheet.cell(row=index_linha,
                                       column=titulo_dic['CHESF - TELEASSIST‰NCIA N3']['OBSERVA€ŽO']).value),
                        # 15 - AGRUPAMENTO
                        str(sheet.cell(row=index_linha,
                                       column=titulo_dic['CHESF - TELEASSIST‰NCIA N3']['AGRUPAMENTO']).value),
                        # N3
                        # 16 - OCR (SAGE)
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 3']['OCR (SAGE)']).value),
                        # 17 - COMANDO
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 3']['COMANDO']).value),
                        # 18 - MEDI€ŽO
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 3']['MEDI€ŽO']).value),
                        # 19 - LISTA DE ALARMES
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 3']['LISTA DE ALARME']).value),
                        # 20 - SOE
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 3']['SOE']).value),
                        # 21 - OBSERVA€ŽO
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 3']['OBSERVA€ŽO']).value),
                        # 22 - AGRUPAMETO
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 3']['AGRUPAMENTO']).value),
                        # ONS
                        # 23 - ITEM
                        str(sheet.cell(row=index_linha, column=titulo_dic['ONS']['ITEM']).value),
                        # 24 - DESCRI€ŽO
                        str(sheet.cell(row=index_linha, column=titulo_dic['ONS']['DESCRI€ŽO']).value),
                        # LIMITES OPERACIONAIS
                        # 25 - LIU
                        str(sheet.cell(row=index_linha, column=titulo_dic['LIMITES OPERACIONAIS']['LIU']).value),
                        # 26 - LIE
                        str(sheet.cell(row=index_linha, column=titulo_dic['LIMITES OPERACIONAIS']['LIE']).value),
                        # 27 - LIA
                        str(sheet.cell(row=index_linha, column=titulo_dic['LIMITES OPERACIONAIS']['LIA']).value),
                        # 28 - LSA
                        str(sheet.cell(row=index_linha, column=titulo_dic['LIMITES OPERACIONAIS']['LSA']).value),
                        # 29 - LSE
                        str(sheet.cell(row=index_linha, column=titulo_dic['LIMITES OPERACIONAIS']['LSE']).value),
                        # 30 - LSU
                        str(sheet.cell(row=index_linha, column=titulo_dic['LIMITES OPERACIONAIS']['LSU']).value),
                        # 31 - BNDMO
                        str(sheet.cell(row=index_linha, column=titulo_dic['LIMITES OPERACIONAIS']['BNDMO']).value),
                        # 32 - OBSERVA€™ES
                        str(sheet.cell(row=index_linha,
                                       column=titulo_dic['LIMITES OPERACIONAIS']['OBSERVA€™ES']).value),
                        # 33 - ENDERE€O N3 Teleassistˆncia
                        str(sheet.cell(row=index_linha,
                                       column=titulo_dic['CHESF - TELEASSIST‰NCIA N3']['ENDERE€O']).value),
                        # 34 - ENDERE€O N3
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 3']['ENDERE€O']).value)]
                for i in range(0, len(array_coletado)):
                    if array_coletado[i] == 'None':
                        array_coletado[i] = ''
                array_validar.append(array_coletado)
    except:
        print_exc(file=stdout)
        GObject.idle_add(FASgtkui.mensagem_erro,'Erro', 'O programa n„o reconhece o arquivo a ser checado como v lido')
        time.sleep(1)
        while FASgtkui.mensagem_erro_dialog.get_visible() == True:
            time.sleep(1)
        gerararquivo = False
    if array_base:
        array_padrao = array_base
    else:
        for pad in gerarlp(LP_Padrao, Arq_Conf)[0]:  # usar fun‡„o gerarlp para criar array_padrao
            # ID SAGE
            array_padrao_coletado = [pad[0],
                                     # OCR
                                     pad[1].value,
                                     # DESCRI€ŽO
                                     pad[2],
                                     # TIPO
                                     pad[3].value,
                                     # COMANDO
                                     pad[4].value,
                                     # MEDI€ŽO
                                     pad[5].value,
                                     # ANUNCIADOR
                                     pad[6].value,
                                     # LISTA DE ALARMES
                                     pad[7].value,
                                     # SOE
                                     pad[8].value,
                                     # N3 -TELEASSIST‰NCIA
                                     # OCR
                                     pad[11].value,
                                     # COMANDO
                                     pad[12].value,
                                     # MEDI€ŽO
                                     pad[13].value,
                                     # LISTA DE ALARMES,
                                     pad[14].value,
                                     # SOE
                                     pad[15].value,
                                     # OBSERVA€ŽO:
                                     pad[16].value,
                                     # AGRUPAMENTO,
                                     pad[18].value,
                                     # N3
                                     # OCR(SAGE)
                                     pad[19].value,
                                     # COMANDO
                                     pad[20].value,
                                     # MEDI€ŽO
                                     pad[21].value,
                                     # LISTA DE ALARMES
                                     pad[22].value,
                                     # SOE
                                     pad[23].value,
                                     # OBSERVA€ŽO
                                     pad[24].value,
                                     # AGRUPAMENTO
                                     pad[26].value,
                                     # ONS
                                     # ITEM
                                     pad[27].value,
                                     # DESCRI€ŽO
                                     pad[28].value,
                                     # LIMITES OPERACIONAIS
                                     # LIU
                                     pad[29].value,
                                     # LIE
                                     pad[30].value,
                                     # LIA
                                     pad[31].value,
                                     # LSA
                                     pad[32].value,
                                     # LSE
                                     pad[33].value,
                                     # LSU
                                     pad[34].value,
                                     # BNDMO
                                     pad[35].value,
                                     # OBSERVA€™ES
                                     pad[36].value]
            for i in range(0, len(array_padrao_coletado)):
                if array_padrao_coletado[i] == None:
                    array_padrao_coletado[i] = ''
            array_padrao.append(array_padrao_coletado)

            # -----------------------------------------------------------------------------------------------------------------------------------------
    COD_SE = array_validar[0][0].split(':')[0]
    nome_arq_saida = 'Relatorio_{}.xlsx'.format(COD_SE)  # Nome do arquivo de sa¡da
    seq_arq = 0  # Sequˆncia do n£mero de arquivo
    while os.path.exists(
            Diretorio_Padrao + '\\' + nome_arq_saida):  # Enquanto existir na pasta um arquivo com o nome definido
        seq_arq += 1  # Adicionar um a sequˆncia do n£mero do arquivo
        nome_arq_saida = 'Relatorio_{}_{}.xlsx'.format(COD_SE, str(
            seq_arq))  # Definir novo nome de arquivo (Ex './LP_gerada.'+'_'+'1'+'.xlsx')
    nome_arq_saida = Diretorio_Padrao + '\\' + nome_arq_saida
    arq_Relatorio = xlsxwriter.Workbook(nome_arq_saida[2:])

    ### Formata‡„o da c‚lula T¡tulo ###
    formatCelTitulo = arq_Relatorio.add_format({
        'bold': True,
        'font_name': 'Arial',
        'font_size': 9,
        'rotation': 90,
        'align': 'center',
        'valign': 'vcenter',
        'bg_color': 'silver',
    })

    ### Formata‡„o da c‚lula Errada###
    formatCelErro = arq_Relatorio.add_format({
        # 'bold': True,
        # 'font_name':'Arial',
        # 'font_size':12,
        'rotation': 0,
        'align': 'left',
        'valign': 'vcenter',
        'bg_color': 'red',
    })

    ### Formata‡„o da c‚lula Sugerida###
    formatCelSur = arq_Relatorio.add_format({
        # 'bold': True,
        # 'font_name':'Arial',
        # 'font_size':12,
        'rotation': 0,
        'align': 'left',
        'valign': 'vcenter',
        'bg_color': 'yellow',
    })

    # ----------Validar pontos da LP que est  sendo verificada----------#
    dic_padrao = {'{}_{}'.format(dic[0], str(dic[4]).strip()): dic for dic in
                  array_padrao}  # ID_COMANDO : LINHA COMPLETA DE REGISTRO

    dic_validar = {'{}_{}'.format(dic[0], str(dic[4]).strip()): dic[:-1] for dic in
                   array_validar}  # ID_COMANDO : LINHA DE REGISTRO SEM N3
    dic_faltando = {x: dic_padrao[x] for x in dic_padrao if x not in dic_validar}
    try:
        sugestao1_dic = {
            '{}_{}_{}_{}'.format(str(dic[0]).split(':')[1], str(dic[0]).split(':')[2], str(dic[2]).strip(),
                                 str(dic[4]).strip()): (
                dic[0], dic[4]) for dic in
            dic_faltando.values()}  # Chave: [VŽO]_[IED]_[DESCRI€ŽO]_[COMANDO], Valor: [ID SAGE]
        sugestao2_dic = {
            '{}_{}_{}'.format(str(dic[0]).split(':')[1], str(dic[2]).strip(), str(dic[4]).strip()): (dic[0], dic[4]) for
            dic in dic_faltando.values()}  # Chave: [VŽO]_[DESCRI€ŽO]_[COMANDO], Valor: [ID SAGE]
    except:
        GObject.idle_add(FASgtkui.mensagem_aviso,'Impossibilidade de Sugest„o de ID',
                                'N„o ser  poss¡vel realizar sugest„o de ponto.\nProvavelmente existem ID de pontos fora do padr„o')
        time.sleep(1)
        while FASgtkui.message_aviso_dialog.get_visible() == True:
            time.sleep(1)
    # sugestao1_dic = {'{}_{}_{}_{}'.format(dic[0].split(':')[1], dic[0].split(':')[2], dic[2].strip(),dic[4].strip()) : (dic[0],dic[4]) for dic in array_padrao} # Chave: [VŽO]_[IED]_[DESCRI€ŽO]_[COMANDO], Valor: [ID SAGE]
    # sugestao2_dic = {'{}_{}_{}'.format(dic[0].split(':')[1], dic[2].strip(),dic[4].strip()) : (dic[0],dic[4]) for dic in array_padrao} # Chave: [VŽO]_[DESCRI€ŽO]_[COMANDO], Valor: [ID SAGE]
    # array_validar_ID_COM = [(col[0],col[4]) for col in array_validar]

    array_validar_semN3 = [arr[:-2] for arr in array_validar]
    if array_base:
        array_padrao_semN3 = [arr[:-2] for arr in array_padrao]
        for validar in array_validar_semN3:
            if validar not in array_padrao_semN3:
                diferenca_array.append(validar)
                k_inc += 1
    else:
        for validar in array_validar_semN3:
            if validar not in array_padrao:
                diferenca_array.append(validar)
                k_inc += 1

    array_padrao_ID_COM = [(col[0], col[4]) for col in array_padrao]
    for diferenca in diferenca_array:
        try:
            posicao = array_padrao_ID_COM.index((diferenca[0], diferenca[4]))
            campos_corretos = []
            for i in range(33):
                if array_padrao[posicao][i] != diferenca[i]:
                    diferenca[i] = '*' + diferenca[i]
                    if array_padrao[posicao][i].strip() != 'X' or array_padrao[posicao][i].strip() != '':
                        campos_corretos.append(array_padrao[posicao][i])
            diferenca.append(' <<>> '.join(campos_corretos))
        except:  # entra aqui se "array_padrao_ID.index(diferenca[0])" levantar exce‡„o por n„o conter "diferenca[0]" no array "array_padrao_ID"
            # Sugerir ID baseado no equipamento e descri‡„o do ponto
            try:
                vao_dif = diferenca[0].split(':')[1]  # V„o/Equipamento do ponto que n„o foi achado ID(SAGE)
                ied_dif = diferenca[0].split(':')[2]  # IED do ponto que n„o foi achado ID(SAGE)
                dsc_dif = diferenca[2].strip()  # Descri‡„o do ponto que n„o foi achado ID(SAGE)
                cmd_dif = diferenca[4].strip()  # Campo Comando do ponto que n„o foi achado ID(SAGE)

                # Tentar sugest„o usando VŽO_EQUIP_DESC_COMANDO, se n„o conseguir tentar com VŽO_DESC_COMANDO
                sugestao_ID = sugestao1_dic.get('{}_{}_{}_{}'.format(vao_dif, ied_dif, dsc_dif, cmd_dif),
                                                sugestao2_dic.get('{}_{}_{}'.format(vao_dif, dsc_dif, cmd_dif), ''))

                # Se n„o conseguiu sugest„o_ID ainda, tentar por similaridade da descri‡„o nos pontos faltantes
                if not sugestao_ID:
                    dic_vao = {}
                    for reg in dic_faltando.values():  # Passar todos os registros faltantes
                        vao = reg[0].split(':')[1]  # V„o/Equipamento
                        if vao not in dic_vao:  # Se ainda n„o existir o dicion rio do V„o/Equipamento
                            dic_vao[vao] = {reg[2]: (
                                reg[0], reg[4])}  # Criar dicion rio do V„o/Equipamento com Descri‡„o como chave
                        else:  # Se existir o dicion rio do V„o/Equipamento
                            dic_vao[vao][reg[2]] = (
                                reg[0], reg[4])  # Gravar mais um registro no dicion rio do V„o/Equipamento

                    # Procura descri‡„o semelhante no dic_vao nos registros faltantes do vao_dif
                    dsc_match = get_close_matches(dsc_dif, dic_vao[vao_dif])[0]
                    # Procura ID dic_vao nos registros faltantes do vao_dif de acordo com dsc_match
                    sugestao_ID = dic_vao[vao_dif][dsc_match]

                sugestao_ID_array.append(sugestao_ID)
            except:
                sugestao_ID = ''

            diferenca[0] = '*' + diferenca[0]  # Marcar ID como n„o encontrado
            if diferenca[4] not in ['', 'CS', 'CD', 'SP']:
                diferenca[4] = '*' + diferenca[4]  # Marcar Comando inv lido
            if sugestao_ID:
                diferenca.append('Poss¡vel ID -> {}'.format(sugestao_ID[0]))
            else:
                diferenca.append('')
    # print(diferenca_array)
    planilha_problema = arq_Relatorio.add_worksheet('Problema')  # Criar Planilha "Problema"

    largura = [22, 18, 40, 8, 5, 5, 5, 5, 5, 18, 5, 8, 5, 5, 40, 5, 18, 5, 8, 5, 5, 40, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5,
               30, 50]
    for i in range(0, 34):  # Ajuste da largura das colunas
        planilha_problema.set_column(i, i, largura[i])

    array_titulo = ['ID (SAGE)',
                    'OCR (SAGE)',
                    'DESCRI€ŽO',
                    'TIPO',
                    'COMANDO',
                    'MEDI€ŽO',
                    'TELA',
                    'LISTA DE ALARMES',
                    'SOE',
                    'OCR (SAGE)', 'COMANDO', 'MEDI€ŽO', 'LISTA DE ALARMES', 'SOE', 'OBSERVA€ŽO', 'AGRUPAMENTO',
                    'OCR (SAGE)', 'COMANDO', 'MEDI€ŽO', 'LISTA DE ALARME', 'SOE', 'OBSERVA€ŽO', 'AGRUPAMENTO',
                    'ITEM', 'DESCRI€ŽO',
                    'LIU', 'LIE', 'LIA', 'LSA', 'LSE', 'LSU', 'BNDMO',
                    'OBSERVA€™ES',
                    'SUGEST™ES']
    formato7 = arq_Relatorio.add_format({
        'bold': True,
        'font_name': 'Arial',
        'font_size': 12,
        'rotation': 0,
        'align': 'left',
        'valign': 'vcenter',
        'bg_color': 'silver',
        'border': 1,
    })
    planilha_problema.merge_range('A1:I2', 'CHESF - N‹VEL 2', formato7)
    planilha_problema.merge_range('J1:P2', 'CHESF - TELEASSIST‰NCIA N3', formato7)
    planilha_problema.merge_range('Q1:W2', 'CHESF - N‹VEL 3', formato7)
    planilha_problema.merge_range('X1:Y1', 'ONS', formato7)
    planilha_problema.merge_range('X2:Y2', 'PROC DE REDE', formato7)
    planilha_problema.merge_range('Z1:AF2', 'LIMITES OPERACIONAIS', formato7)

    for i in range(0, len(array_titulo)):  # Gravar t¡tulo
        planilha_problema.write(2, i, array_titulo[i], formatCelTitulo)

    linha = 3
    msgerroNumero = False
    for dado in diferenca_array:  # Passa por todas as linhas do array de sa¡da
        for i in range(34):
            try:
                if dado[i].startswith('*'):  # testa se o campo est  marcado como "incoerente"
                    planilha_problema.write(linha, i, dado[i][1:],
                                            formatCelErro)  # se est  "incoerente" grava na planilha usando uma formata‡„o diferente
                else:
                    planilha_problema.write(linha, i, dado[
                        i])  # se est  "incoerente" grava na planilha usando uma formata‡„o default
            except:
                msgerroNumero = True
        linha += 1

    if msgerroNumero:
        print_exc(file=stdout)
        gerararquivo = False
        GObject.idle_add(FASgtkui.mensagem_erro,'Erro',
                               'Verifique preenchimento de campos no Arquivo LP a ser checado. Nem um dos campos deve ser preenchido apenas com n£meros')
        time.sleep(1)
        while FASgtkui.mensagem_erro_dialog.get_visible() == True:
            time.sleep(1)
    # ----------Pontos Faltantes----------#
    for pfaltando in sorted(dic_faltando.items(), key=itemgetter(0)):
        pfalta_array.append(pfaltando[1])
        k_falta += 1

    planilha_Pfaltantes = arq_Relatorio.add_worksheet('Pontos_faltantes')  # Criar Planilha "Pontos_faltantes"
    planilha_Pfaltantes.merge_range('A1:I2', 'CHESF - N‹VEL 2', formato7)
    planilha_Pfaltantes.merge_range('J1:P2', 'CHESF - TELEASSIST‰NCIA N3', formato7)
    planilha_Pfaltantes.merge_range('Q1:W2', 'CHESF - N‹VEL 3', formato7)
    planilha_Pfaltantes.merge_range('X1:Y1', 'ONS', formato7)
    planilha_Pfaltantes.merge_range('X2:Y2', 'PROC DE REDE', formato7)
    planilha_Pfaltantes.merge_range('Z1:AF2', 'LIMITES OPERACIONAIS', formato7)

    for i in range(0, 33):  # Ajuste da largura das colunas
        planilha_Pfaltantes.set_column(i, i, largura[i])

    for i in range(0, len(array_titulo) - 1):  # Gravar t¡tulo
        planilha_Pfaltantes.write(2, i, array_titulo[i], formatCelTitulo)

    linha = 3
    for dado in pfalta_array:  # Passa por todas as linhas do array de sa¡da
        for i in range(33):
            if (dado[i], dado[4]) in sugestao_ID_array:  # testa se o campo est  entre IDs sugeridos
                planilha_Pfaltantes.write(linha, i, dado[i], formatCelSur)
            else:
                planilha_Pfaltantes.write(linha, i, dado[i])
        linha += 1

    # ----------Verificar Endere‡o N3 da LP padr„o que n„o est„o na LP que est  sendo verificada----------#
    endN3TeleassPreenchido = False
    endN3Preenchido = False
    endN3Teleass = [arr[-2] for arr in array_validar]
    endN3 = [arr[-1] for arr in array_validar]
    # 34 - ENDERE€O N3
    for endere‡o in endN3Teleass:
        if endere‡o != '':
            endN3TeleassPreenchido = True
            break
    for endere‡o in endN3:
        if endere‡o != '':
            endN3Preenchido = True
            break
    if endN3TeleassPreenchido:
        array_validar_endereco = endN3Teleass
    elif endN3Preenchido:
        array_validar_endereco = endN3

    for endereco in array_validar_endereco:
        if array_validar_endereco.count(endereco) > 1:
            if endereco not in endduplicado_array:
                endduplicado_array.append(endereco)
                k_enddupl += array_validar_endereco.count(endereco)

    planilha_EndDupl = arq_Relatorio.add_worksheet('End. N3 duplicados')  # Criar Planilha "End. N3 duplicados"

    array_titulo = ['Endere‡o',
                    'Ocorrˆncia']

    coluna = 0
    for titulo in array_titulo:  # Gravar t¡tulo
        planilha_EndDupl.write(0, coluna, titulo)
        coluna += 1

    linha = 1
    for dado in endduplicado_array:  # Passa por todas as linhas do array de sa¡da
        planilha_EndDupl.write(linha, 0, dado)
        planilha_EndDupl.write(linha, 1, str(array_validar_endereco.count(dado)))
        linha += 1

    # ----------Planilha Resumo----------#
    planilha_resumo = arq_Relatorio.add_worksheet('Resumo')  # Criar Planilha "Resumo"

    texto_resumo = ['-----Pontos com problemas-----',
                    '',
                    'Quantidade: {} pontos'.format(k_inc),
                    'Percentual: {:2.2f}%'.format(float(len(diferenca_array)) * 100 / len(array_validar)),
                    '',
                    '-----Pontos faltantes-----',
                    '',
                    'Quantidade: {} pontos'.format(k_falta),
                    'Percentual: {:2.2f}%'.format(float(len(pfalta_array)) * 100 / len(array_validar)),
                    '',
                    '-----Endere‡o para N3 Duplicado-----',
                    '',
                    'Quantidade: {} pontos'.format(k_enddupl),
                    'Percentual: {:2.2f}%'.format(k_enddupl * 100 / len(array_validar))]

    planilha_resumo.set_column(0, 0, 35)
    for linha, texto in enumerate(texto_resumo):
        planilha_resumo.write(linha, 0, texto)

    # ----------Gravar arquivo Excel----------#
    if gerararquivo:
        arq_Relatorio.close()
        GObject.idle_add(FASgtkui.dialogo_abrir_arquivo_gerado, nome_arq_saida.rsplit('\\', 1)[1], Diretorio_Padrao)

        nomearquivo = nome_arq_saida[2:]
        conf = {'arquivo': nomearquivo}
        pickle.dump(conf, open('fas.p', 'wb'), -1)  # -1 para gravar em Bin rio
