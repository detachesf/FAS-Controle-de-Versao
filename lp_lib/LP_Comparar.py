# -*- coding: cp860 -*-
from FASgtkui import mensagem_erro
from traceback import print_exc
from sys import stdout

dados = '''
Vers„o 2.0.6
Atualiza‡„o do programa: 27/07/2015
Monta janela Comparar
'''
try:
    from openpyxl import load_workbook, cell
except:
    mensagem_erro('Erro', 'M¢dulo openpyxl n„o instalado')

try:
    from xlrd import open_workbook
except:
    mensagem_erro('Erro', 'Modulo xlrd n„o instalado')

try:
    from lp_lib.Checar_LP import checar
except:
    mensagem_erro('Erro', 'M¢dulo Checar_LP n„o instalado')

try:
    from lp_lib.func import processing
except:
    mensagem_erro('Erro', 'M¢dulo func n„o instalado')
try:
    from lp_lib.func import linhaInicialETitulos
except:
    mensagem_erro('Erro', 'Arquivo "func.pyc" deve estar no diret¢rio "lp_lib"')


def Comparar(LPBase, cbBase, Checar, cbChecar, Diretorio_Padrao):
    book = load_workbook(LPBase, data_only=True)  # Abrir arquivo de LP Base
    sheet = book[cbBase.get_active_text()]  # Abrir planilha
    array_base = []
    try:
        # Lˆ planilha e recebe a linha onde come‡a a LP (aqui usando linha inicial e n„o o dicion rio de t¡tulos)
        li, titulo_dic = linhaInicialETitulos(LPBase, cbBase.get_active_text())
        if li < 0:  # Se for um n£mero negativo ent„o n„o foi encontrado "ID (SAGE)" na lista
            raise NameError('Arquivo especificado n„o possui coluna com t¡tulo "ID (SAGE)".')
        for index_linha in range(li, sheet.max_row + 1):  # Ler c‚lulas da linha 7 ao final
            if sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 2']['ID (SAGE)']).value != '' and \
                    sheet.cell(row=index_linha,
                               column=titulo_dic['CHESF - N‹VEL 2']['ID (SAGE)']).value != 'CGS' and \
                    sheet.cell(row=index_linha,
                               column=titulo_dic['CHESF - N‹VEL 2']['ID (SAGE)']).value != 'PDS' and \
                    sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 2']['ID (SAGE)']).value != 'PAS':
                try:  # Caso a descri‡„o do campo 6 seja "TELA"
                    # 0 - ID SAGE
                    array_coletado = [
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 2']['ID (SAGE)']).value),
                        # N2
                        # 1 - OCR
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 2']['OCR (SAGE)']).value),
                        # 1 - DESCRI€ŽO
                        str(sheet.cell(row=index_linha,
                                       column=titulo_dic['CHESF - N‹VEL 2']['DESCRI€ŽO']).value).strip(),
                        # 2 - TIPO
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 2']['TIPO']).value),
                        # 3 - COMANDO
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 2']['COMANDO']).value),
                        # 4 - MEDI€ŽO
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 2']['MEDI€ŽO']).value),
                        # 5 - TELA
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 2']['TELA']).value),
                        # 6 - LISTA DE ALARMES
                        str(sheet.cell(row=index_linha,
                                       column=titulo_dic['CHESF - N‹VEL 2']['LISTA DE ALARMES']).value),
                        # 7 - SOE
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 2']['SOE']).value),
                        # TELEASSIST‰NCIA N3
                        # 8 - OCR
                        str(sheet.cell(row=index_linha,
                                       column=titulo_dic['CHESF - TELEASSIST‰NCIA N3']['OCR (SAGE)']).value),
                        # 9 - COMANDO
                        str(sheet.cell(row=index_linha,
                                       column=titulo_dic['CHESF - TELEASSIST‰NCIA N3']['COMANDO']).value),
                        # 10 - MEDI€ŽO
                        str(sheet.cell(row=index_linha,
                                       column=titulo_dic['CHESF - TELEASSIST‰NCIA N3']['MEDI€ŽO']).value),
                        # 11 - LISTA DE ALARMES
                        str(sheet.cell(row=index_linha,
                                       column=titulo_dic['CHESF - TELEASSIST‰NCIA N3']['LISTA DE ALARME']).value),
                        # 12 - SOE
                        str(sheet.cell(row=index_linha,
                                       column=titulo_dic['CHESF - TELEASSIST‰NCIA N3']['SOE']).value),
                        # 13 - OBSERVA€ŽO
                        str(sheet.cell(row=index_linha,
                                       column=titulo_dic['CHESF - TELEASSIST‰NCIA N3']['OBSERVA€ŽO']).value),
                        # 15 - AGRUPAMENTO
                        str(sheet.cell(row=index_linha,
                                       column=titulo_dic['CHESF - TELEASSIST‰NCIA N3']['AGRUPAMENTO']).value),
                        # N3
                        # 16 - OCR (SAGE)
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 3']['OCR (SAGE)']).value),
                        # 17 - COMANDO
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 3']['COMANDO']).value),
                        # 18 - MEDI€ŽO
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 3']['MEDI€ŽO']).value),
                        # 19 - LISTA DE ALARMES
                        str(sheet.cell(row=index_linha,
                                       column=titulo_dic['CHESF - N‹VEL 3']['LISTA DE ALARME']).value),
                        # 20 - SOE
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 3']['SOE']).value),
                        # 21 - OBSERVA€ŽO
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 3']['OBSERVA€ŽO']).value),
                        # 22 - AGRUPAMETO
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 3']['AGRUPAMENTO']).value),
                        # ONS
                        # 23 - ITEM
                        str(sheet.cell(row=index_linha, column=titulo_dic['ONS']['ITEM']).value),
                        # 24 - DESCRI€ŽO
                        str(sheet.cell(row=index_linha, column=titulo_dic['ONS']['DESCRI€ŽO']).value),
                        # LIMITES OPERACIONAIS
                        # 25 - LIU
                        str(sheet.cell(row=index_linha, column=titulo_dic['LIMITES OPERACIONAIS']['LIU']).value),
                        # 26 - LIE
                        str(sheet.cell(row=index_linha, column=titulo_dic['LIMITES OPERACIONAIS']['LIE']).value),
                        # 27 - LIA
                        str(sheet.cell(row=index_linha, column=titulo_dic['LIMITES OPERACIONAIS']['LIA']).value),
                        # 28 - LSA
                        str(sheet.cell(row=index_linha, column=titulo_dic['LIMITES OPERACIONAIS']['LSA']).value),
                        # 29 - LSE
                        str(sheet.cell(row=index_linha, column=titulo_dic['LIMITES OPERACIONAIS']['LSE']).value),
                        # 30 - LSU
                        str(sheet.cell(row=index_linha, column=titulo_dic['LIMITES OPERACIONAIS']['LSU']).value),
                        # 31 - BNDMO
                        str(sheet.cell(row=index_linha, column=titulo_dic['LIMITES OPERACIONAIS']['BNDMO']).value),
                        # 32 - OBSERVA€™ES
                        str(sheet.cell(row=index_linha,
                                       column=titulo_dic['LIMITES OPERACIONAIS']['OBSERVA€™ES']).value),
                        # 33 - ENDERE€O N3 Teleassistˆncia
                        str(sheet.cell(row=index_linha,
                                       column=titulo_dic['CHESF - TELEASSIST‰NCIA N3']['ENDERE€O']).value),
                        # 34 - ENDERE€O N3
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 3']['ENDERE€O']).value)]
                except:  # Caso a descri‡„o do campo 6 seja "ANUNCIADOR"
                    # 0 - ID SAGE
                    array_coletado = [
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 2']['ID (SAGE)']).value),
                        # N2
                        # 1 - OCR
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 2']['OCR (SAGE)']).value),
                        # 2 - DESCRI€ŽO
                        str(sheet.cell(row=index_linha,
                                       column=titulo_dic['CHESF - N‹VEL 2']['DESCRI€ŽO']).value).strip(),
                        # 3 - TIPO
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 2']['TIPO']).value),
                        # 4 - COMANDO
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 2']['COMANDO']).value),
                        # 5 - MEDI€ŽO
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 2']['MEDI€ŽO']).value),
                        # 6 - TELA
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 2']['ANUNCIADOR']).value),
                        # 7 - LISTA DE ALARMES
                        str(sheet.cell(row=index_linha,
                                       column=titulo_dic['CHESF - N‹VEL 2']['LISTA DE ALARMES']).value),
                        # 8 - SOE
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 2']['SOE']).value),
                        # TELEASSIST‰NCIA N3
                        # 9 - OCR
                        str(sheet.cell(row=index_linha,
                                       column=titulo_dic['CHESF - TELEASSIST‰NCIA N3']['OCR (SAGE)']).value),
                        # 10 - COMANDO
                        str(sheet.cell(row=index_linha,
                                       column=titulo_dic['CHESF - TELEASSIST‰NCIA N3']['COMANDO']).value),
                        # 11 - MEDI€ŽO
                        str(sheet.cell(row=index_linha,
                                       column=titulo_dic['CHESF - TELEASSIST‰NCIA N3']['MEDI€ŽO']).value),
                        # 12 - LISTA DE ALARMES
                        str(sheet.cell(row=index_linha,
                                       column=titulo_dic['CHESF - TELEASSIST‰NCIA N3']['LISTA DE ALARME']).value),
                        # 13 - SOE
                        str(sheet.cell(row=index_linha,
                                       column=titulo_dic['CHESF - TELEASSIST‰NCIA N3']['SOE']).value),
                        # 14 - OBSERVA€ŽO
                        str(sheet.cell(row=index_linha,
                                       column=titulo_dic['CHESF - TELEASSIST‰NCIA N3']['OBSERVA€ŽO']).value),
                        # 15 - AGRUPAMENTO
                        str(sheet.cell(row=index_linha,
                                       column=titulo_dic['CHESF - TELEASSIST‰NCIA N3']['AGRUPAMENTO']).value),
                        # N3
                        # 16 - OCR (SAGE)
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 3']['OCR (SAGE)']).value),
                        # 17 - COMANDO
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 3']['COMANDO']).value),
                        # 18 - MEDI€ŽO
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 3']['MEDI€ŽO']).value),
                        # 19 - LISTA DE ALARMES
                        str(sheet.cell(row=index_linha,
                                       column=titulo_dic['CHESF - N‹VEL 3']['LISTA DE ALARME']).value),
                        # 20 - SOE
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 3']['SOE']).value),
                        # 21 - OBSERVA€ŽO
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 3']['OBSERVA€ŽO']).value),
                        # 22 - AGRUPAMETO
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 3']['AGRUPAMENTO']).value),
                        # ONS
                        # 23 - ITEM
                        str(sheet.cell(row=index_linha, column=titulo_dic['ONS']['ITEM']).value),
                        # 24 - DESCRI€ŽO
                        str(sheet.cell(row=index_linha, column=titulo_dic['ONS']['DESCRI€ŽO']).value),
                        # LIMITES OPERACIONAIS
                        # 25 - LIU
                        str(sheet.cell(row=index_linha, column=titulo_dic['LIMITES OPERACIONAIS']['LIU']).value),
                        # 26 - LIE
                        str(sheet.cell(row=index_linha, column=titulo_dic['LIMITES OPERACIONAIS']['LIE']).value),
                        # 27 - LIA
                        str(sheet.cell(row=index_linha, column=titulo_dic['LIMITES OPERACIONAIS']['LIA']).value),
                        # 28 - LSA
                        str(sheet.cell(row=index_linha, column=titulo_dic['LIMITES OPERACIONAIS']['LSA']).value),
                        # 29 - LSE
                        str(sheet.cell(row=index_linha, column=titulo_dic['LIMITES OPERACIONAIS']['LSE']).value),
                        # 30 - LSU
                        str(sheet.cell(row=index_linha, column=titulo_dic['LIMITES OPERACIONAIS']['LSU']).value),
                        # 31 - BNDMO
                        str(sheet.cell(row=index_linha, column=titulo_dic['LIMITES OPERACIONAIS']['BNDMO']).value),
                        # 32 - OBSERVA€™ES
                        str(sheet.cell(row=index_linha,
                                       column=titulo_dic['LIMITES OPERACIONAIS']['OBSERVA€™ES']).value),
                        # 33 - ENDERE€O N3 Teleassistˆncia
                        str(sheet.cell(row=index_linha,
                                       column=titulo_dic['CHESF - TELEASSIST‰NCIA N3']['ENDERE€O']).value),
                        # 34 - ENDERE€O N3
                        str(sheet.cell(row=index_linha, column=titulo_dic['CHESF - N‹VEL 3']['ENDERE€O']).value)]
                for i in range(0, len(array_coletado)):
                    if array_coletado[i] == 'None':
                        array_coletado[i] = ''
                array_base.append(array_coletado)
    except:
        print_exc(file=stdout)
        mensagem_erro('Erro', 'O programa n„o reconhece o arquivo base como v lida')

        # checar(LP_Editado=self.Checar,planilha=self.cbChecar.get(), relatorio=self.relatorioJanelaPrincipal, array_base=array_base)
    # self.toplevel.destroy() #Fechar Janela
    try:
        processing(checar, {'LP_Editado': Checar, 'planilha': cbChecar.get_active_text(),
                            'array_base': array_base, 'Diretorio_Padrao': Diretorio_Padrao})
    except:
        mensagem_erro('Erro', 'Erro inesperado ao tentar checar lista de pontos.')
