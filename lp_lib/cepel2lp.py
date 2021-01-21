# -*- coding: cp860 -*-

dados= '''
Vers„o 2.0.0
Atualiza‡„o do programa: 31/10/2014
Gera‡„o de LP no padr„o Chesf tendo como base a planilha CEPEL de gera‡„o de base de dados SAGE
'''

from tkinter.messagebox import showerror, showwarning, askyesno
from os import path, getcwd, startfile
from sys import stdout
from traceback import print_exc
try:
    from xlrd import open_workbook
except:
    showerror('Erro', 'M¢dulo xlrd n„o instalado')
try:
    from openpyxl import load_workbook,cell
except:
    showerror('Erro', 'M¢dulo openpyxl n„o instalado')
try:
    from lp_lib.gerarPlanilhaLP import gerarPlanilha
except:
    showerror('Erro','Arquivo "gerarPlanilhaLP.py" deve estar no mesmo diret¢rio "lp_lib"')

def cepel2lp(arqcepel):
    
    nome_arq_saida = './LP_da_Planilha_CEPEL.xlsx'      #Nome do arquivo de sa¡da
    seq_arq = 0                                         #Sequˆncia do n£mero de arquivo
    while path.exists(nome_arq_saida):                  #Enquanto existir na pasta um arquivo com o nome definido
        seq_arq += 1                                    #Adicionar um a sequˆncia do n£mero do arquivo
        nome_arq_saida = nome_arq_saida[:22]+'_'+str(seq_arq)+'.xlsx' #Definir novo nome de arquivo     
    
    
    arq_LP = gerarPlanilha(nome_arq_saida)              # Gera um arquivo Excel com uma planilha com formata‡„o da Lista de Pontos Padr„o
    planilha_LP = arq_LP.worksheets()[0]
    planilha_relatorio = arq_LP.add_worksheet('RELATORIO')
    
    try:
        arq_cepel = load_workbook(arqcepel, data_only=True)  # Abrir Planilha CEPEL
    except:
        showerror('Erro', 'Planilha CEPEL n„o encontrada ou formato n„o suportado, utilize arquivos .xlsm ou .xlsx')

    def Titulos(sheet):
        titulos = {}
        for li in range(2,10):                                          #Varrer as linhas de 2 a 10
            for i in range(sheet.max_column):                                #Varrer as colunas da linha
                texto_coluna = str(sheet.cell(row= li, column= i+1).value).upper().strip()   #Pegar texto da c‚lula
                if texto_coluna == '':                                  #Gravar £ltima posi‡„o com valor vazio
                    titulos[texto_coluna] = i+1
                elif texto_coluna not in titulos:                       #Iserir chave se n„o existir no dicion rio
                    titulos[texto_coluna] = i+1
            if 'ID' in titulos: break                     #Se foi passado pela linha com chave "ID (SAGE)" parar de varrer linhas 
        
        li += 1                                                         #Seleciona linha ap¢s o t¡tulo
        if 'ID' in titulos:                               #Verifica se foi encontrado chave "ID (SAGE)"
            while True:
                if sheet.max_row+1 == li: break 
                if sheet.cell(row=li,column=titulos['ID']).value:              #Verifica se a c‚lula est  preenchida com algum valor
                    break                                                   #Parar de procurar linha preenchida
                else:
                    li += 1                                                 #Selecionar linha seguinte
        else:
            li = -1            
        return [li,titulos]       
    

    ##### Ler PDS #####
    try:
        try:
            pds_sh = arq_cepel['PDS']  # Abrir planilha "PDS" da Planilha CEPEL
        except:
            pds_sh = arq_cepel['pds']  # Abrir planilha "PDS" da Planilha CEPEL
        li, titulo_dic = Titulos(pds_sh)
        pds_dic = {}
        for index_linha in range(li, pds_sh.max_row+1):
            ID ='' if str(pds_sh.cell(row = index_linha, column = titulo_dic['ID']).value) == 'None' else str(pds_sh.cell(row = index_linha, column = titulo_dic['ID']).value)
            ColProc = titulo_dic.get('','#None')  # Pegar coluna com T¡tulo vazio com informa‡„o x/y
            if ColProc != '#None':  #Se existir coluna x/y
                c3 = str(pds_sh.cell(row=index_linha, column= ColProc).value).strip().lower() =='x'
            else:
                c3 = True
            if str(ID).strip() and str(ID).strip()[0] != ';' and c3:
                OCR = '' if str(pds_sh.cell(row = index_linha, column= titulo_dic['OCR']).value) =='None' else str(pds_sh.cell(row = index_linha, column= titulo_dic['OCR']).value)
                NOME =  '' if str(pds_sh.cell(row = index_linha, column= titulo_dic['NOME']).value) == 'None' else str(pds_sh.cell(row = index_linha, column= titulo_dic['NOME']).value)
                TIPO = '' if str(pds_sh.cell(row = index_linha, column= titulo_dic['TIPO']).value) == 'None' else str(pds_sh.cell(row = index_linha, column= titulo_dic['TIPO']).value)
                ALRIN = '' if str(pds_sh.cell(row = index_linha, column= titulo_dic['ALRIN']).value).strip() == 'SIM' else  'X'
                SOE = '' if str(pds_sh.cell(row = index_linha, column= titulo_dic['SOEIN']).value).strip() == 'SIM' else  'X'
                TAC = '' if str(pds_sh.cell(row = index_linha, column= titulo_dic['TAC']).value) == 'None' else str(pds_sh.cell(row = index_linha, column= titulo_dic['TAC']).value)
                pds_dic[ID] = [OCR, NOME, TIPO, ALRIN,SOE, TAC]
    except:
        print_exc(file=stdout)
        showerror('Erro', 'N„o foi poss¡vel processar a planilha PDS')
                
    ##### Ler PDD #####
    try:
        try:
            pdd_sh = arq_cepel['PDD']  # Abrir planilha "PDD" da Planilha CEPEL
        except:
            pdd_sh = arq_cepel['pdd']  # Abrir planilha "PDD" da Planilha CEPEL
        li, titulo_dic = Titulos(pdd_sh)
        pdd_dic = {}
        for index_linha in range(li, pdd_sh.max_row+1):
            ID = '' if str(pdd_sh.cell(row = index_linha, column= titulo_dic['ID']).value) == 'None' else str(pdd_sh.cell(row = index_linha, column= titulo_dic['ID']).value)
            ColProc = titulo_dic.get('','#None')  # Pegar coluna com T¡tulo vazio com informa‡„o x/y
            if ColProc != '#None':  #Se existir coluna x/y
                c3 = str(pdd_sh.cell(row = index_linha, column= ColProc).value).strip().lower() =='x'
            else:
                c3 = True
            if str(ID).strip() and str(ID).strip()[0] != ';' and c3:
                PDS = '' if str(pdd_sh.cell(row = index_linha, column= titulo_dic['PDS']).value) == 'None' else str(pdd_sh.cell(row = index_linha, column= titulo_dic['PDS']).value)
                pdd_dic[PDS] = [ID]
    except:
        showerror('Erro', 'N„o foi poss¡vel processar a planilha PDD')
    ##### Ler PDF #####
    try:
        try:
            pdf_sh = arq_cepel['PDF']  # Abrir planilha "PDF" da Planilha CEPEL
        except:
            pdf_sh = arq_cepel['pdf']  # Abrir planilha "PDF" da Planilha CEPEL
        li, titulo_dic = Titulos(pdf_sh)
        pdf_dic = {}
        for index_linha in range(li, pdf_sh.max_row+1):
            ID = '' if str(pdf_sh.cell(row = index_linha, column= titulo_dic['ID']).value) == 'None' else str(pdf_sh.cell(row = index_linha, column= titulo_dic['ID']).value)
            ColProc = titulo_dic.get('','#None')  # Pegar coluna com T¡tulo vazio com informa‡„o x/y
            if ColProc != '#None':  #Se existir coluna x/y
                c3 = str(pdf_sh.cell(row = index_linha, column= ColProc).value).strip().lower() =='x'
            else:
                c3 = True
            if str(ID).strip() and str(ID).strip()[0] != ';':
                PNT = '' if str(pdf_sh.cell(row = index_linha, column= titulo_dic['PNT']).value) == 'None' else str(pdf_sh.cell(row = index_linha, column= titulo_dic['PNT']).value)
                ORDEM = '' if str(pdf_sh.cell(row = index_linha, column= titulo_dic['ORDEM']).value) == 'None' else str(pdf_sh.cell(row = index_linha, column= titulo_dic['ORDEM']).value)
                pdf_dic[PNT] = [ID,ORDEM]
    except:
        print_exc(file = stdout)
        showerror('Erro', 'N„o foi poss¡vel processar a planilha PDF')
        
        ##### Ler PTS #####
    try:
        try:
            pts_sh = arq_cepel['PTS']  # Abrir planilha "PTS" da Planilha CEPEL
        except:
            pts_sh = arq_cepel['pts']  # Abrir planilha "PTS" da Planilha CEPEL
        li, titulo_dic = Titulos(pts_sh)
        pts_dic = {}
        for index_linha in range(li, pts_sh.max_row+1):
            ID = '' if str(pts_sh.cell(row = index_linha, column= titulo_dic['ID']).value) == 'None' else str(pts_sh.cell(row = index_linha, column= titulo_dic['ID']).value)
            ColProc = titulo_dic.get('','#None')  # Pegar coluna com T¡tulo vazio com informa‡„o x/y
            if ColProc != '#None':  #Se existir coluna x/y
                c3 = str(pts_sh.cell(row = index_linha, column= ColProc).value).strip().lower() =='x'
            else:
                c3 = True
            if str(ID).strip() and str(ID).strip()[0] != ';':
                OCR = '' if str(pts_sh.cell(row = index_linha, column= titulo_dic['OCR']).value) == 'None' else str(pts_sh.cell(row = index_linha, column= titulo_dic['OCR']).value)
                NOME = '' if str(pts_sh.cell(row = index_linha, column= titulo_dic['NOME']).value) == 'None' else str(pts_sh.cell(row = index_linha, column= titulo_dic['NOME']).value)
                TIPO = '' if str(pts_sh.cell(row = index_linha, column= titulo_dic['TIPO']).value) == 'None' else str(pts_sh.cell(row = index_linha, column= titulo_dic['TIPO']).value)
                ALRIN = '' if str(pts_sh.cell(row = index_linha, column= titulo_dic['ALRIN']).value).strip() == 'SIM' else  'X'
                LSA = '' if str(pts_sh.cell(row = index_linha, column= titulo_dic['LSA']).value) == 'None' else str(pts_sh.cell(row = index_linha, column= titulo_dic['LSA']).value)
                LSE = '' if str(pts_sh.cell(row = index_linha, column= titulo_dic['LSE']).value) == 'None' else str(pts_sh.cell(row = index_linha, column= titulo_dic['LSE']).value)
                LSU = '' if str(pts_sh.cell(row = index_linha, column= titulo_dic['LSU']).value) =='None' else str(pts_sh.cell(row = index_linha, column= titulo_dic['LSU']).value)
                TAC = '' if str(pts_sh.cell(row = index_linha, column= titulo_dic['TAC']).value) == 'None' else str(pts_sh.cell(row = index_linha, column= titulo_dic['TAC']).value)
                pts_dic.update({ID:[OCR, NOME, TIPO, ALRIN, LSA, LSE, LSU, TAC]})
    except:
        showwarning('Aten‡„o', 'N„o foi poss¡vel processar a planilha PTS')
        
        ##### Ler PTD #####
    try:
        try:
            ptd_sh = arq_cepel['PTD']  # Abrir planilha "PTD" da Planilha CEPEL
        except:
            ptd_sh = arq_cepel['ptd']  # Abrir planilha "PTD" da Planilha CEPEL
        li, titulo_dic = Titulos(ptd_sh)
        ptd_dic = {}
        for index_linha in range(li, ptd_sh.max_row+1):
            ID = '' if str(ptd_sh.cell(row = index_linha, column= titulo_dic['ID']).value) == 'None' else str(ptd_sh.cell(row = index_linha, column= titulo_dic['ID']).value)
            ColProc = titulo_dic.get('','#None')  # Pegar coluna com T¡tulo vazio com informa‡„o x/y
            if ColProc != '#None':  #Se existir coluna x/y
                c3 = str(ptd_sh.cell(row = index_linha, column= ColProc).value).strip().lower() =='x'
            else:
                c3 = True
            if str(ID).strip() and str(ID).strip()[0] != ';':
                PTS = '' if str(ptd_sh.cell(row = index_linha, column= titulo_dic['PTS']).value) == 'None' else str(ptd_sh.cell(row = index_linha, column= titulo_dic['PTS']).value)
                ptd_dic[PTS] = [ID]
    except:
        showwarning('Aten‡„o', 'N„o foi poss¡vel processar a planilha PTD')        

        ##### Ler PTF #####
    try:
        try:
            ptf_sh = arq_cepel['PTF']  # Abrir planilha "PTF" da Planilha CEPEL
        except:
            ptf_sh = arq_cepel['ptf']  # Abrir planilha "PTF" da Planilha CEPEL
        li, titulo_dic = Titulos(ptf_sh)
        ptf_dic = {}
        for index_linha in range(li, ptf_sh.max_row+1):
            ID = '' if str(ptf_sh.cell(row = index_linha, column= titulo_dic['ID']).value) == 'None' else str(ptf_sh.cell(row = index_linha, column= titulo_dic['ID']).value)
            ColProc = titulo_dic.get('','#None')  # Pegar coluna com T¡tulo vazio com informa‡„o x/y
            if ColProc != '#None':  #Se existir coluna x/y
                c3 = str(ptf_sh.cell(row = index_linha, column= ColProc).value).strip().lower() =='x'
            else:
                c3 = True
            if str(ID).strip() and str(ID).strip()[0] != ';':
                PNT ='' if str(ptf_sh.cell(row = index_linha, column= titulo_dic['PNT']).value) == 'None' else str(ptf_sh.cell(row = index_linha, column= titulo_dic['PNT']).value)
                ptf_dic[PNT] = [ID]
    except:
        showwarning('Aten‡„o', 'N„o foi poss¡vel processar a planilha PTF')        
        
        ##### Ler PAS #####
    try:
        try:
            pas_sh = arq_cepel['PAS']  # Abrir planilha "PAS" da Planilha CEPEL
        except:
            pas_sh = arq_cepel['pas']  # Abrir planilha "PAS" da Planilha CEPEL
        li,titulo_dic = Titulos(pas_sh)
        pas_dic = {}
        for index_linha in range(li, pas_sh.max_row+1):
            ID = '' if str(pas_sh.cell(row = index_linha, column= titulo_dic['ID']).value) == 'None' else str(pas_sh.cell(row = index_linha, column= titulo_dic['ID']).value)
            ColProc = titulo_dic.get('','#None')  # Pegar coluna com T¡tulo vazio com informa‡„o x/y
            if ColProc != '#None':  #Se existir coluna x/y
                c3 = str(pas_sh.cell(row = index_linha, column= ColProc).value).strip().lower() =='x'
            else:
                c3 = True
            if str(ID).strip() and str(ID).strip()[0] != ';':
                OCR = '' if str(pas_sh.cell(row = index_linha, column= titulo_dic['OCR']).value) == 'None' else str(pas_sh.cell(row = index_linha, column= titulo_dic['OCR']).value)
                NOME ='' if str(pas_sh.cell(row = index_linha, column= titulo_dic['NOME']).value) == 'None' else str(pas_sh.cell(row = index_linha, column= titulo_dic['NOME']).value)
                TIPO = '' if str(pas_sh.cell(row = index_linha, column= titulo_dic['TIPO']).value) == 'None' else str(pas_sh.cell(row = index_linha, column= titulo_dic['TIPO']).value)
                ALRIN = '' if str(pas_sh.cell(row = index_linha, column= titulo_dic['ALRIN']).value).strip() == 'SIM' else  'X'
                LIU = '' if str(pas_sh.cell(row = index_linha, column= titulo_dic['LIU']).value) == 'None' else str(pas_sh.cell(row = index_linha, column= titulo_dic['LIU']).value)
                LIE = '' if str(pas_sh.cell(row = index_linha, column= titulo_dic['LIE']).value) == 'None' else str(pas_sh.cell(row = index_linha, column= titulo_dic['LIE']).value)
                LIA = '' if str(pas_sh.cell(row = index_linha, column= titulo_dic['LIA']).value) == 'None' else str(pas_sh.cell(row = index_linha, column= titulo_dic['LIA']).value)
                LSA = '' if str(pas_sh.cell(row = index_linha, column= titulo_dic['LSA']).value) == 'None' else str(pas_sh.cell(row = index_linha, column= titulo_dic['LSA']).value)
                LSE = '' if str(pas_sh.cell(row = index_linha, column= titulo_dic['LSE']).value) == 'None' else str(pas_sh.cell(row = index_linha, column= titulo_dic['LSE']).value)
                LSU = '' if str(pas_sh.cell(row = index_linha, column= titulo_dic['LSU']).value) == 'None' else str(pas_sh.cell(row = index_linha, column= titulo_dic['LSU']).value)
                BNDMO = '' if str(pas_sh.cell(row = index_linha, column= titulo_dic['BNDMO']).value) == 'None' else str(pas_sh.cell(row = index_linha, column= titulo_dic['BNDMO']).value)
                TAC = ''  if str(pas_sh.cell(row = index_linha, column= titulo_dic['TAC']).value) == 'None' else str(pas_sh.cell(row = index_linha, column= titulo_dic['TAC']).value)
                pas_dic[ID] = [OCR, NOME, TIPO, ALRIN, LIU, LIE, LIA, LSA, LSE, LSU, BNDMO, TAC]
    except:
        print_exc(file=stdout)
        showerror('Erro', 'N„o foi poss¡vel processar a planilha PAS')        
        
        ##### Ler PAD #####
    try:
        try:
            pad_sh = arq_cepel['PAD']  # Abrir planilha "PAD" da Planilha CEPEL
        except:
            pad_sh = arq_cepel['pad']  # Abrir planilha "PAD" da Planilha CEPEL
        li, titulo_dic = Titulos(pad_sh)
        pad_dic = {}
        for index_linha in range(li, pad_sh.max_row+1):
            ID ='' if str(pad_sh.cell(row = index_linha, column= titulo_dic['ID']).value) == 'None' else str(pad_sh.cell(row = index_linha, column= titulo_dic['ID']).value)
            ColProc = titulo_dic.get('','#None')  # Pegar coluna com T¡tulo vazio com informa‡„o x/y
            if ColProc != '#None':  #Se existir coluna x/y
                c3 = str(pad_sh.cell(row = index_linha, column= ColProc).value).strip().lower() =='x'
            else:
                c3 = True
            if str(ID).strip() and str(ID).strip()[0] != ';':
                PAS = '' if str(pad_sh.cell(row = index_linha, column= titulo_dic['PAS']).value) else str(pad_sh.cell(row = index_linha, column= titulo_dic['PAS']).value)
                pad_dic[PAS] = [ID]
    except:
        print_exc(file=stdout)
        showerror('Erro', 'N„o foi poss¡vel processar a planilha PAD')
        
        ##### Ler PAF #####
    try:
        try:
            paf_sh = arq_cepel['PAF']  # Abrir planilha "PAF" da Planilha CEPEL
        except:
            paf_sh = arq_cepel['paf']  # Abrir planilha "PAF" da Planilha CEPEL
        li, titulo_dic = Titulos(paf_sh)
        paf_dic = {}
        for index_linha in range(li, paf_sh.max_row+1):
            ID = '' if str(paf_sh.cell(row = index_linha, column= titulo_dic['ID']).value) else str(paf_sh.cell(row = index_linha, column= titulo_dic['ID']).value)
            ColProc = titulo_dic.get('','#None')  # Pegar coluna com T¡tulo vazio com informa‡„o x/y
            if ColProc != '#None':  #Se existir coluna x/y
                c3 = str(paf_sh.cell(row = index_linha, column= ColProc).value).strip().lower() == 'x'
            else:
                c3 = True
            if str(ID).strip() and str(ID).strip()[0] != ';':
                PNT = '' if str(paf_sh.cell(row = index_linha, column= titulo_dic['PNT']).value) else str(paf_sh.cell(row = index_linha, column= titulo_dic['PNT']).value)
                paf_dic[PNT] = [ID]
    except:
        print_exc(file=stdout)
        showerror('Erro', 'N„o foi poss¡vel processar a planilha PAF')

        ##### Ler CGS #####
    try:
        try:
            cgs_sh = arq_cepel['CGS']  # Abrir planilha "CGS" da Planilha CEPEL
        except:
            cgs_sh = arq_cepel['cgs']  # Abrir planilha "CGS" da Planilha CEPEL
        li, titulo_dic = Titulos(cgs_sh)
        cgs_dic = {}
        for index_linha in range(li, cgs_sh.max_row+1):
            ID = '' if str(cgs_sh.cell(row = index_linha, column= titulo_dic['ID']).value) == 'None' else str(cgs_sh.cell(row = index_linha, column= titulo_dic['ID']).value)
            ColProc = titulo_dic.get('','#None')  # Pegar coluna com T¡tulo vazio com informa‡„o x/y
            if ColProc != '#None':  #Se existir coluna x/y
                c3 = str(cgs_sh.cell(row = index_linha, column= ColProc).value).strip().lower() =='x'
            else:
                c3 = True
            if str(ID).strip() and str(ID).strip()[0] != ';':
                NOME = '' if str(cgs_sh.cell(row = index_linha, column= titulo_dic['NOME']).value) == 'None' else str(cgs_sh.cell(row = index_linha, column= titulo_dic['NOME']).value)
                PAC = '' if str(cgs_sh.cell(row = index_linha, column= titulo_dic['PAC']).value) == 'None' else str(cgs_sh.cell(row = index_linha, column= titulo_dic['PAC']).value)
                TAC = '' if str(cgs_sh.cell(row = index_linha, column= titulo_dic['TAC']).value) == 'None' else str(cgs_sh.cell(row = index_linha, column= titulo_dic['TAC']).value)
                TIPOE = '' if str(cgs_sh.cell(row = index_linha, column= titulo_dic['TIPOE']).value) == 'None' else str(cgs_sh.cell(row = index_linha, column= titulo_dic['TIPOE']).value)
                cgs_dic.update({ID:[NOME, PAC, TAC, TIPOE]})
    except:
        print_exc(file=stdout)
        showerror('Erro', 'N„o foi poss¡vel processar a planilha CGS')
        
        ##### Ler CGF #####
    try:
        try:
            cgf_sh = arq_cepel['CGF']  # Abrir planilha "CGF" da Planilha CEPEL
        except:
            cgf_sh = arq_cepel['cgf']  # Abrir planilha "CGF" da Planilha CEPEL
        li, titulo_dic = Titulos(cgf_sh)
        cgf_dic = {}
        for index_linha in range(li, cgf_sh.max_row+1):
            ID = '' if str(cgf_sh.cell(row = index_linha, column= titulo_dic['ID']).value) == 'None' else str(cgf_sh.cell(row = index_linha, column= titulo_dic['ID']).value)
            ColProc = titulo_dic.get('','#None')  # Pegar coluna com T¡tulo vazio com informa‡„o x/y
            if ColProc != '#None':  #Se existir coluna x/y
                c3 = str(cgf_sh.cell(row = index_linha, column= ColProc).value).strip().lower() =='x'
            else:
                c3 = True
            if str(ID).strip() and str(ID).strip()[0] != ';':
                CGS = '' if str(cgf_sh.cell(row = index_linha, column= titulo_dic['CGS']).value) == 'None' else str(cgf_sh.cell(row = index_linha, column= titulo_dic['CGS']).value)
                NV2 = '' if str(cgf_sh.cell(row = index_linha, column= titulo_dic['NV2']).value) == 'None' else str(cgf_sh.cell(row = index_linha, column= titulo_dic['NV2']).value)
                cgf_dic[CGS] = [ID, NV2]
    except:
        print_exc(file=stdout)
        showerror('Erro', 'N„o foi poss¡vel processar a planilha CGF')
        
        ##### Ler TAC #####
    try:
        try:
            tac_sh = arq_cepel['TAC']  # Abrir planilha "TAC" da Planilha CEPEL
        except:
            tac_sh = arq_cepel['tac']  # Abrir planilha "TAC" da Planilha CEPEL
        li, titulo_dic = Titulos(tac_sh)
        tac_dic = {}
        for index_linha in range(li, tac_sh.max_row+1):
            ID = '' if str(tac_sh.cell(row = index_linha, column= titulo_dic['ID']).value) == 'None' else str(tac_sh.cell(row = index_linha, column= titulo_dic['ID']).value)
            ColProc = titulo_dic.get('','#None')  # Pegar coluna com T¡tulo vazio com informa‡„o x/y
            if ColProc != '#None':  #Se existir coluna x/y
                c3 = str(tac_sh.cell(row = index_linha, column= ColProc).value).strip().lower() =='x'
            else:
                c3 = True
            if str(ID).strip() and str(ID).strip()[0] != ';':
                LSC = '' if  str(tac_sh.cell(row = index_linha, column= titulo_dic['LSC']).value) == 'None' else str(tac_sh.cell(row = index_linha, column= titulo_dic['LSC']).value)
                tac_dic[ID] = [LSC]
    except:
        print_exc(file=stdout)
        showerror('Erro', 'N„o foi poss¡vel processar a planilha TAC')
                
    ##### Gravar Pontos Digitais Excel #####
    pdig = []
    linha_rel = 1
    planilha_relatorio.write(0,0,'ID n„o encontrados em PDF') 
    for id_tag in pds_dic:
        try:
            pdf_id = pdf_dic[id_tag]
        except:
            pdf_id = ['']
            if 'CALC' not in pds_dic[id_tag][5] and 'LOCAL' not in pds_dic[id_tag][5]:
                planilha_relatorio.write(linha_rel,0,id_tag)
                linha_rel += 1
        pdig.append([id_tag]+pds_dic[id_tag]+pdf_id)
    
    linha = 6
    for dado in pdig: #Passa por todas as linhas do array de pontos digitais gravando pontos no Excel
        tac = dado[6]
        lsc = tac_dic.get(tac,['?'])[0]     
        id_protocolo  = dado[7]
        tag = dado[0]
        ocr = dado[1]
        descr = dado[2]
        tipo = dado[3]
        anunciador = ''
        alarme = dado[4]
        soe = dado[5]
        obs = ''
        
        id_pdd = pdd_dic.get(tag,['',''])
        end = pdf_dic.get(id_pdd[0],['',''])[1]
    
        planilha_LP.write(linha,0,linha-5)                # escreve na coluna "ITEM"
        planilha_LP.write(linha,2,tac)                    # escreve na coluna "TAC" 
        planilha_LP.write(linha,3,lsc)                    # escreve na coluna "IED"           
        planilha_LP.write(linha,7,id_protocolo)           # escreve na coluna "ID PROTOCOLO"
        planilha_LP.write(linha,9,tag)                    # escreve na coluna "ID (SAGE)"
        planilha_LP.write(linha,10,ocr)                   # escreve na coluna "OCR"
        planilha_LP.write(linha,11,descr)                 # escreve na coluna "DESCRI€ŽO"
        planilha_LP.write(linha,12,tipo)                  # escreve na coluna "TIPO"
        planilha_LP.write(linha,15,anunciador)            # escreve na coluna "ANUNCIADOR"
        planilha_LP.write(linha,16,alarme)                # escreve na coluna "LISTA DE ALARMES"
        planilha_LP.write(linha,17,soe)                   # escreve na coluna "SOE"
        planilha_LP.write(linha,18,obs)                   # escreve na coluna "OBSERVA€ŽO"
        planilha_LP.write(linha,34,end)                   # escreve na coluna "ENDERECO"
        linha += 1                                          # incrementa a linha
    
    ##### Gravar Pontos Anal¢gicos Excel #####
    pana = []
    linha_rel = 1
    planilha_relatorio.write(0,2,'ID n„o encontrados em PAF')
    for id_tag in pas_dic:
        try:
            paf_id = paf_dic[id_tag]
        except:
            paf_id = ['']
            if 'CALC' not in pas_dic[id_tag][11] and 'LOCAL' not in pas_dic[id_tag][11]:                
                planilha_relatorio.write(linha_rel,2,id_tag)
                linha_rel += 1
        pana.append([id_tag]+pas_dic[id_tag]+paf_id)
    
    med_dic = {'FR':'Hz', 'KV':'kV', 'AM':'A', 'DI':'km', 'MV':'MVAR', 'MW':'MW', 'TM':'ø C'}
    for dado in pana: #Passa por todas as linhas do array de pontos anal¢gicos gravando pontos no Excel
        tac = dado[12]
        lsc = tac_dic.get(tac,['?'])[0]
    
        id_protocolo  = dado[7]           
        id_protocolo  = dado[13]
        tag = dado[0]
        ocr = dado[1]
        descr = dado[2]
        tipo = dado[3]
        medicao = med_dic.get(tipo[:2],'')
        anunciador = ''
        alarme = dado[4]
        obs = ''
    
        id_pad = pad_dic.get(tag,['',''])
        end = paf_dic.get(id_pad[0],['',''])[0]
        
        liu = dado[5]
        lie = dado[6]
        lia = dado[7]
        lsa = dado[8]
        lse = dado[9]
        lsu = dado[10]
        bndmo = dado[11]                     
        
        planilha_LP.write(linha,0,linha-5)                # escreve na coluna "ITEM"
        planilha_LP.write(linha,2,tac)                    # escreve na coluna "TAC" 
        planilha_LP.write(linha,3,lsc)                    # escreve na coluna "IED" 
        planilha_LP.write(linha,7,id_protocolo)           # escreve na coluna "ID PROTOCOLO"
        planilha_LP.write(linha,9,tag)                    # escreve na coluna "ID (SAGE)"
        planilha_LP.write(linha,10,ocr)                   # escreve na coluna "OCR"
        planilha_LP.write(linha,11,descr)                 # escreve na coluna "DESCRI€ŽO"
        planilha_LP.write(linha,12,tipo)                  # escreve na coluna "TIPO"
        planilha_LP.write(linha,13,'')                    # escreve na coluna "COMANDO"
        planilha_LP.write(linha,14,medicao)               # escreve na coluna "MEDI€ŽO"
        planilha_LP.write(linha,15,anunciador)            # escreve na coluna "ANUNCIADOR"
        planilha_LP.write(linha,16,alarme)                # escreve na coluna "LISTA DE ALARMES"
        planilha_LP.write(linha,17,'')                    # escreve na coluna "SOE"
        planilha_LP.write(linha,18,obs)                   # escreve na coluna "OBSERVA€ŽO"
        planilha_LP.write(linha,34,end)                   # escreve na coluna "ENDERECO"
        planilha_LP.write(linha,38,liu)                   # escreve na coluna "LIU"
        planilha_LP.write(linha,39,lie)                   # escreve na coluna "LIE"
        planilha_LP.write(linha,40,lia)                   # escreve na coluna "LIA"
        planilha_LP.write(linha,41,lsa)                   # escreve na coluna "LSA"
        planilha_LP.write(linha,42,lse)                   # escreve na coluna "LSE"
        planilha_LP.write(linha,43,lsu)                   # escreve na coluna "LSU"
        planilha_LP.write(linha,44,bndmo)                 # escreve na coluna "BNDMO"            
        linha += 1                                          # incrementa a linha  
    
    ##### Gravar Pontos Totalizadores Excel #####
    ptot = []
    linha_rel = 1
    planilha_relatorio.write(0,4,'ID n„o encontrados em PTF')
    try:
        for id_tag in pts_dic:
            try:
                ptf_id = ptf_dic[id_tag]
            except:
                ptf_id = ['']
                if 'CALC' not in pts_dic[id_tag][7] and 'LOCAL' not in pts_dic[id_tag][7]:
                    planilha_relatorio.write(linha_rel,4,id_tag)
                    linha_rel += 1
            ptot.append([id_tag]+pts_dic[id_tag]+ptf_id)
    except:
        showwarning('Aten‡„o', 'N„o foram gravados pontos totalizadores')
    for dado in ptot: #Passa por todas as linhas do array de pontos anal¢gicos gravando pontos no Excel
        tac = dado[8]
        lsc = tac_dic.get(tac,['?'])[0]      
        id_protocolo  = dado[7]
        id_protocolo  = dado[9]
        tag = dado[0]
        ocr = dado[1]
        descr = dado[2]
        tipo = dado[3]
        anunciador = ''
        alarme = dado[4]
        obs = ''
        id_ptd = ptd_dic.get(tag,['',''])
        end = ptf_dic.get(id_ptd[0],['',''])[0]               
        lsa = dado[5]
        lse = dado[6]
        lsu = dado[7]              
        
        planilha_LP.write(linha,0,linha-5)                # escreve na coluna "ITEM"
        planilha_LP.write(linha,2,tac)                    # escreve na coluna "TAC"
        planilha_LP.write(linha,3,lsc)                    # escreve na coluna "IED"  
        planilha_LP.write(linha,7,id_protocolo)           # escreve na coluna "ID PROTOCOLO"
        planilha_LP.write(linha,9,tag)                    # escreve na coluna "ID (SAGE)"
        planilha_LP.write(linha,10,ocr)                   # escreve na coluna "OCR"
        planilha_LP.write(linha,11,descr)                 # escreve na coluna "DESCRI€ŽO"
        planilha_LP.write(linha,12,tipo)                  # escreve na coluna "TIPO"
        planilha_LP.write(linha,13,'')                    # escreve na coluna "COMANDO"
        planilha_LP.write(linha,15,anunciador)            # escreve na coluna "ANUNCIADOR"
        planilha_LP.write(linha,16,alarme)                # escreve na coluna "LISTA DE ALARMES"
        planilha_LP.write(linha,17,soe)                   # escreve na coluna "SOE"
        planilha_LP.write(linha,18,obs)                   # escreve na coluna "OBSERVA€ŽO"
        planilha_LP.write(linha,34,end)                   # escreve na coluna "ENDERECO"
        planilha_LP.write(linha,41,lsa)                   # escreve na coluna "LSA"
        planilha_LP.write(linha,42,lse)                   # escreve na coluna "LSE"
        planilha_LP.write(linha,43,lsu)                   # escreve na coluna "LSU"
        linha += 1                                          # incrementa a linha
    
    ##### Gravar Pontos Comandos Excel #####
    pcmd = []
    linha_rel = 1
    planilha_relatorio.write(0,6,'ID n„o encontrados em CGF')
    for id_tag in cgs_dic:
        try:
            cgf_id = [cgf_dic[id_tag][0]]
        except:
            cgf_id = ['']
            planilha_relatorio.write(linha_rel,6,id_tag)
            linha_rel += 1
    
        try:
            cgf_com = cgf_dic[id_tag][1]
        except:
            cgf_com = ''
        if 'CSIM' in cgf_com:
            cgf_com = ['CS']
        elif 'CDUP' in cgf_com:
            cgf_com = ['CD']
        else:
            cgf_com = ['CD']
            
        pcmd.append([id_tag]+cgs_dic[id_tag]+cgf_id+cgf_com)
    
    for dado in pcmd: #Passa por todas as linhas do array de pontos anal¢gicos gravando pontos no Excel
        tac = dado[3]
        lsc = tac_dic.get(tac,['?'])[0]
        id_protocolo  = dado[5]
        tag = dado[0]
        descr = dado[1]
        pac = dado[2]
        tipo = dado[4]
        comando = dado[6]          
        
        if (tag == 'LOCAL') or ('COR' not in tag) or (len(tag) == len(pac)): 
            planilha_LP.write(linha,0,linha-5)               # escreve na coluna "ITEM"
            planilha_LP.write(linha,2,tac)                   # escreve na coluna "TAC" 
            planilha_LP.write(linha,3,lsc)                   # escreve na coluna "IED"
            planilha_LP.write(linha,7,id_protocolo)          # escreve na coluna "ID PROTOCOLO"
            planilha_LP.write(linha,9,tag)                   # escreve na coluna "ID (SAGE)"
            planilha_LP.write(linha,11,descr)                # escreve na coluna "DESCRI€ŽO"
            planilha_LP.write(linha,12,tipo)                 # escreve na coluna "TIPO"
            planilha_LP.write(linha,13,comando)                  # escreve na coluna "COMANDO"
            linha += 1                                         # incrementa a linha        
    
    #lista_ses = [i[0:3] for i in pds_dic.keys()]
    #cod_se = lista_ses[0]
    #for i in set(lista_ses):
    #    if lista_ses.count(i) > lista_ses.count(cod_se): cod_se = i
    
    arq_LP.close()
                         
    abrirarquivo = askyesno('Aviso', 'Arquivo \"'+nome_arq_saida[2:]+'\" gerado em ' + getcwd()+'\n\n Deseja abrir o arquivo gerado agora?')
    if abrirarquivo : startfile(getcwd() + '\\' + nome_arq_saida[2:])
        
if __name__ == "__main__":
    from tkinter.filedialog import askopenfilename
    arqcepel = askopenfilename(filetypes=[('Arquivo do Excel','xls'), ('Arquivo do Excel','xlsx'), ('Arquivo do Excel','xlsm')])
    if arqcepel:
        cepel2lp(arqcepel)
