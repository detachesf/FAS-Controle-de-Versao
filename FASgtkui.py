import gi
from xml.etree.ElementTree import Element, SubElement, tostring, ElementTree
from xml.dom import minidom
from datetime import date
import os
from bs4 import BeautifulSoup

gi.require_version("Gtk", "3.0")
from gi.repository import Gtk

builder = Gtk.Builder()  # inicia o construtor Gtk
builder.add_from_file('user_interface.glade')

from os import path, getcwd
from sys import stdout
from traceback import print_exc

# Caixas de diálogo

mensagem_erro_dialog: Gtk.MessageDialog = builder.get_object('message_erro')

def mensagem_erro(titulo, msg):

    mensagem_erro_dialog.props.text = titulo
    mensagem_erro_dialog.set_title('Erro')
    mensagem_erro_dialog.props.secondary_text = msg
    mensagem_erro_dialog.props.icon_name = 'dialog-error-symbolic'
    mensagem_erro_dialog.show_all()
    mensagem_erro_dialog.run()
    mensagem_erro_dialog.hide()

message_aviso_dialog: Gtk.MessageDialog = builder.get_object('message_aviso')
def mensagem_aviso(titulo, msg):

    message_aviso_dialog.props.text = titulo
    message_aviso_dialog.props.secondary_text = msg
    message_aviso_dialog.props.icon_name = 'dialog-warning-symbolic'
    message_aviso_dialog.show_all()
    message_aviso_dialog.run()
    message_aviso_dialog.hide()


def pergunta_sim_nao(titulo, msg):
    perguntasimnao: Gtk.MessageDialog = builder.get_object('pergunta_sim_nao')
    perguntasimnao.props.text = titulo
    perguntasimnao.props.secondary_text = msg
    perguntasimnao.props.icon_name = 'dialog-question-symbolic'
    perguntasimnao.show()
    resposta = perguntasimnao.run()
    perguntasimnao.hide()
    if resposta == Gtk.ResponseType.YES:
        return True
    elif resposta == Gtk.ResponseType.NO:
        return False

try:
    import lp_lib.LP_Comparar as LP_Comparar
except:
    print_exc(file=stdout)
    mensagem_erro('Erro', 'Módulo LP_Comparar não instalado')

try:
    import lp_lib.Gerar_LP as Gerar_LP
except:
    print_exc(file=stdout)
    mensagem_erro('Erro', 'Módulo Gerar_LP não instalado')

try:
    import lp_lib.Checar_LP as Checar_LP
except:
    print_exc(file=stdout)
    mensagem_erro('Erro', 'Módulo Checar_LP não instalado')

try:
    from lp_lib.func import processing
except:
    mensagem_erro('Erro', 'Módulo func não instalado')
try:
    from openpyxl import load_workbook
except:
    mensagem_erro('Erro', 'Módulo openpyxl não instalado')
try:
    from lp_lib.cepel2lp import cepel2lp
except:
    mensagem_erro('Erro', 'Módulo cepel2lp não instalado')


class Manipulador(object):

    def __init__(self):
        # Vairáveis Gerais
        self.arqconf_novo = True
        self.pathchecar = getcwd()
        self.versao = '2.1.0'
        self.data = '09/02/2020'
        self.window: Gtk.Window = builder.get_object('janela_principal')  # Pega o Objeto da janela princial
        self.window.show_all()  # Mostra a janela principal
        self.lp_de_saida = ''
        self.gerar_lp = False
        self.checar_lp = False
        self.base_para_lp = False
        self.Diretorio_de_salvamento = os.getcwd()
        self.Lb = ""

        # Arrays com os nomes padrão dos objetos de cada linha
        self.dicionario_nomes_objetos = [
            ['selec_linha_LT_', 'LT_entry_codlinha_', 'LT_entry_codpainel_', 'LT_entry_ltremota_',
             'LT_entry_camarapass_',
             'LT_entry_conjuntosecc_', 'LT_combobox_arranjo_', 'LT_combobox_religamento_',
             'LT_checkbtt_rdp_', 'LT_checkbtt_painelteleprot_', 'LT_checkbtt_f9_', 'LT_checkbtt_87l_'],
            ['selec_linha_Trafo_', 'Trafo_entry_codtrafo_', 'Trafo_entry_codpainelH_',
             'Trafo_entry_codpainelX_',
             'Trafo_entry_camarapass_',
             'Trafo_entry_conjuntosecc_', 'Trafo_combobox_arranjoH_',
             'Trafo_combobox_arranjoX_',
             'Trafo_checkbtt_rdp_', 'Trafo_checkbtt_regapp_', 'Trafo_checkbtt_f9_',
             'Trafo_combobox_equip_', 'Trafo_combobox_relacao_', 'Trafo_combobox_prot_',
             ],
            ['selec_linha_vaotrans_', 'vaotrans_entry_cod_', 'vaotrans_entry_painel_',
             'vaotrans_checkbtt_87B_', 'vaotrans_combobox_arranjo_', 'vaotrans_entry_camarapass_',
             'vaotrans_entry_conjuntosecc_'],
            ['selec_linha_paisage_', 'paisage_entry_painel_',
             'paisage_combobox_sagebastidor_',
             'paisage_entry_sw-de_', 'paisage_entry_sw-ate_', 'paisage_entry_nportas-sw_',
             'paisage_checkbtt_fw_', 'paisage_entry_nporta-fw_', 'paisage_checkbtt_rb_',
             'paisage_entry_rb-de_',
             'paisage_entry_rb-ate_', 'paisage_entry_nporta-rb_'],
            ['selec_linha_reator_', 'reator_entry_cod_', 'reator_entry_painel_',
             'reator_checkbtt_manob_', 'reator_combobox_equip_', 'reator_checkbtt_rdp_',
             'reator_checkbtt_f9_',
             'reator_entry_camarapass_', 'reator_entry_conjuntosecc_'],
            ['selec_linha_acesso_', 'acesso_entry_codvao_', 'acesso_entry_painelacess_',
             'acesso_checkbtt_painelexist_',
             'acesso_entry_num-uc-chesf_', 'acesso_entry_num-uc-acessante_',
             'acesso_combobox_arranjo_', 'acesso_checkbtt_ts_',
             'acesso_entry_ts-de_', 'acesso_entry_ts-ate_', 'acesso_checkbtt_rb_',
             'acesso_entry_redbox-de_',
             'acesso_entry_redbox-ate_', 'acesso_checkbtt_multimedidor_', 'acesso_entry_mm-de_',
             'acesso_entry_mm-ate_',
             'acesso_entry_ltremota_'],
            ['selec_linha_tterra_', 'tterra_entry_codigo_', 'tterra_entry_painel_',
             'tterra_entry_camarapass_', 'tterra_entry_conjuntosecc_'],
            ['selec_linha_protbarra_', 'protbarra_entry_painel_', 'protbarra_entry_qtpan_',
             'protbarra_combobox_arranjo_',
             'protbarra_checkbtt_bu-no-painel_', 'protbarra_entry_vaos_'],
            ['selec_linha_bcapshunt_', 'bcapshunt_entry_codigo_', 'bcapshunt_entry_painel_',
             'bcapshunt_combobox_arranjo_',
             'bcapshunt_checkbtt_rdp_', 'bcapshunt_checkbtt_f9_'],
            ['selec_linha_bcapserie_', 'bcapserie_entry_codigo_', 'bcapserie_entry_painel_'],
            ['selec_linha_ece_', 'ece_entry_codigo_', 'ece_entry_painel_'],
            ['selec_linha_sistreg_', 'sistreg_combobox_nome_', 'sistreg_combobox_tensao-reg_',
             'sistreg_entry_painel_'],
            ['selec_linha_prepreen_', 'prepreen_entry_sistema_'],
            ['selec_linha_compsinc_', 'compsinc_entry_codigo_', 'compsinc_entry_painel_'],
            ['', 'saux_entry_nome-painel-ua_', 'saux_entry_nome-painel-saux_',
             'saux_entry_barras-sup-ca_',
             'saux_entry_barras-sup-cc_', 'saux_entry_disj-sup-ca_',
             'saux_entry_disj-sup-cc_',
             'saux_combobox_tensao-ca_', 'saux_combobox_tensao-cc_']]

        # Variáveis Auxiliares na mecânica da tela de configuração

        self.dicionario_geral_objetos_dinamicos = {}  # dicionário para armazenar os objetos adicionados dinâmicamente
        self.dicionario_gera_num_de_linhas_ativas = {}  # dicionário que armazena o número das linhas ativas de cada evento
        self.dicionario_geral_linhas_removidas = {}  # Dicionáriio que registra o índice das linhas que foram removidas de todos eventos
        self.dicionario_geral_tabelas = {}  # Dicionário que guarda as tabelas de cada evento

        for i in range(0, 14):
            self.dicionario_geral_tabelas[self.dicionario_nomes_objetos[i][1].split('_')[0]] = builder.get_object(
                'tabela_' + self.dicionario_nomes_objetos[i][1].split('_')[0])
            self.dicionario_geral_objetos_dinamicos[self.dicionario_nomes_objetos[i][1].split('_')[0]] = {}
            self.dicionario_gera_num_de_linhas_ativas[self.dicionario_nomes_objetos[i][1].split('_')[0]] = [1]
            self.dicionario_geral_linhas_removidas[self.dicionario_nomes_objetos[i][1].split('_')[0]] = []

        self.Arranjos = ['DISJ E MEIO', 'BS', 'BPT', 'BD3',
                         'BD4','BD5']  # Array com os arranjos possíveis para preencher os comboboxes

        # Carregando objetos

        self.janela_sobre: Gtk.AboutDialog = builder.get_object('janela_sobre')
        self.janela_comparar: Gtk.Window = builder.get_object('janela_comparar')
        self.janela_carregando: Gtk.Window = builder.get_object('janela_carregando')

        self.label_progressbar: Gtk.Label = builder.get_object('label_progressbar')
        self.progress_bar: Gtk.ProgressBar = builder.get_object('progress_bar')

        self.notebook: Gtk.Notebook = builder.get_object('notebook1')
        #
        self.codigo_se: Gtk.Entry = builder.get_object('entry_cod_se')
        self.fornecedor: Gtk.Entry = builder.get_object('entry_fornecedor')
        self.usuario: Gtk.Entry = builder.get_object('entry_usuario')
        self.descricao: Gtk.Entry = builder.get_object('entry_descricao_evento')
        self.nome_arqconf: Gtk.Entry = builder.get_object('arqconf_entry_nome-arquivo')
        self.diretorio_dialogo_pasta_entry: Gtk.Entry = builder.get_object('diretorio_dialogo_pasta_entry')


        self.Lppadrao: Gtk.FileChooserButton = builder.get_object('file_chooser_lppadrao')
        self.Lp_a_checar: Gtk.FileChooserButton = builder.get_object('lp_a_checar')

        self.arqconf_salvar_dialogo: Gtk.FileChooserDialog = builder.get_object('arqconf_salvar_dialogo')
        self.arqconf_abrir_dialogo: Gtk.FileChooserDialog = builder.get_object('arqconf_abrir_dialogo')
        self.dialogo_diretorio: Gtk.FileChooserDialog = builder.get_object('diretorio_dialogo')
        self.dialogo_lppadrao: Gtk.FileChooserDialog = builder.get_object('arqlppadrao_dialog')
        self.dialogo_vtelas : Gtk.FileChooserDialog = builder.get_object('vtelasbotoes_dialogo')


        self.comboplan: Gtk.ComboBoxText = builder.get_object('combobox_aba_a_checar')

        self.button_checarlp: Gtk.Button = builder.get_object('button_checarlp')
        self.button_comparar: Gtk.Button = builder.get_object('button_compararlp')

        self.toolbar_checar: Gtk.ToolButton = builder.get_object('toolbar_checar')
        self.menubar_checar: Gtk.ImageMenuItem = builder.get_object('arqconf_menubar_ferramentas_checar')

        self.menubar_checar.set_sensitive(False)
        self.toolbar_checar.set_sensitive(False)
        self.button_comparar.set_sensitive(False)
        self.button_checarlp.set_sensitive(False)


        self.nome_arq_saida = 'Arqconf-novo'  # Nome do arquivo de saída
        seq_arq = 0  # Sequência do número de arquivo
        while os.path.exists(self.nome_arq_saida + '.fas'):  # Enquanto existir na pasta um arquivo com o nome definido
            seq_arq += 1  # Adicionar um a sequência do número do arquivo
            self.nome_arq_saida = self.nome_arq_saida.split('_')[0] + '_' + str(seq_arq)  #
        self.nome_arq_saida = self.nome_arq_saida + '.fas'
        self.window.set_title(self.nome_arq_saida)

        try:
            root = BeautifulSoup(open(os.getcwd() + '\\' + 'ini.xml', 'r', encoding='utf-8'), 'html.parser')
            caminho_lp_padrao = os.getcwd() + '\\' + 'static\\Planilhas Padrao' + '\\' + root.ini['arqlppadrao']
            if os.path.exists(caminho_lp_padrao):
                self.Lppadrao.set_filename(caminho_lp_padrao)
            else:
                mensagem_erro('Erro',
                              'Arquivo {} não encontrado na pasta Planilhas Padrao'.format(root.ini['arqlppadrao']))
                self.Lppadrao.set_filename('')
            self.Diretorio_de_salvamento = str(root.ini['diretorio_padrao'])
        except:
            root = Element('ini', data='{}'.format(date.today()),
                           versão=self.versao,
                           arqlppadrao='Padrao Planilha Supervisao_rev1P.xlsm',
                           diretorio_padrao=self.Diretorio_de_salvamento)
            ElementTree(root).write('ini.xml', 'UTF-8')
            self.Lppadrao.set_filename(
                os.getcwd() + '\\' + 'static\\Planilhas Padrao' + '\\' + 'Padrao Planilha Supervisao_rev1P.xlsm')

    # Sinais de navegação entre páginas

    # Janela de Sobre
    def on_arqconf_menubar_ajuda_sobre_activate(self, window):

        self.janela_sobre.set_version(self.versao)
        self.janela_sobre.show_all()
        resposta = self.janela_sobre.run()
        if resposta == -4:
            self.janela_sobre.hide()

    # Sinais de lógica na tela

    # Ações executadas quando o botão adicionar for clicado
    def on_button_add_linha_clicked(self, button):

        Aba = self.notebook.get_current_page()  # captura a aba ativa
        Nome_Aba = self.dicionario_nomes_objetos[Aba][1].split('_')[0]  # Pega o nome do evento no dicionario de nomes
        if Aba < 14:
            self.adicionar_linha(self.dicionario_geral_linhas_removidas[Nome_Aba],
                                 self.dicionario_gera_num_de_linhas_ativas[Nome_Aba],
                                 Nome_Aba,
                                 self.dicionario_nomes_objetos[Aba],
                                 self.dicionario_geral_objetos_dinamicos[Nome_Aba],
                                 self.dicionario_geral_tabelas[Nome_Aba])

    # Ações executadas quando o botão excluir for clicado
    def on_button_Excluir_clicked(self, button):

        Aba = self.notebook.get_current_page()  # captura a aba ativa
        Nome_Aba = self.dicionario_nomes_objetos[Aba][1].split('_')[0]  # Pega o nome do evento no dicionario de nomes
        if Aba < 14:
            self.exclui_linha(self.dicionario_geral_linhas_removidas[Nome_Aba],
                              self.dicionario_gera_num_de_linhas_ativas[Nome_Aba],
                              self.dicionario_nomes_objetos[Aba],
                              self.dicionario_geral_objetos_dinamicos[Nome_Aba], Nome_Aba)

    # Ações executadas quando o botão duplicar for clicado
    def on_button_duplicar_clicked(self, button):

        Aba = self.notebook.get_current_page()  # captura a aba ativa
        Nome_Aba = self.dicionario_nomes_objetos[Aba][1].split('_')[0]  # Pega o nome do evento no dicionario de nomes
        if Aba < 14:
            self.prepara_para_duplicar(self.dicionario_geral_linhas_removidas[Nome_Aba],
                                       self.dicionario_gera_num_de_linhas_ativas[Nome_Aba],
                                       Nome_Aba,
                                       self.dicionario_nomes_objetos[Aba],
                                       self.dicionario_geral_objetos_dinamicos[Nome_Aba],
                                       self.dicionario_geral_tabelas[Nome_Aba])

    # Ações executadas quando o botão limpar for clicado

    def on_button_limpar_clicked(self, button):

        Aba = self.notebook.get_current_page()  # captura a aba ativa
        Nome_Aba = self.dicionario_nomes_objetos[Aba][1].split('_')[0]  # Pega o nome do evento no dicionario de nomes
        if Aba < 14:
            self.limpar_linha(self.dicionario_gera_num_de_linhas_ativas[Nome_Aba],
                              self.dicionario_nomes_objetos[Aba],
                              self.dicionario_geral_objetos_dinamicos[Nome_Aba])

    # Ações executadas quando o botão selecionar todas for clicado

    def on_selecionar_todas_clicked(self, button):
        Aba = self.notebook.get_current_page()  # captura a aba ativa
        Nome_Aba = self.dicionario_nomes_objetos[Aba][1].split('_')[0]  # Pega o nome do evento no dicionario de nomes
        if Aba < 14:
            self.selecionar_todas(self.dicionario_gera_num_de_linhas_ativas[Nome_Aba],
                                  self.dicionario_nomes_objetos[Aba],
                                  self.dicionario_geral_objetos_dinamicos[Nome_Aba])

    # Ações executadas quando os botões de novo forem clicados
    def on_arqconf_novo_activate(self, button):
        self.restaurar_tela()
        self.arqconf_novo = True
        self.notebook.set_current_page(0)

    # Função que capta os dados dos eventos e joga dentro do elemento 'evento' do arquivo xml
    def recolhe_dados(self, Numero_linhas_ativas, array_nomes_objetos, dicionario_objetos, eventos):
        for linha in Numero_linhas_ativas:  # Varre todas as linhas para achar os checkboxes selecionados
            try:  # Caso para os objetos que foram criados no botão adicionar (dinamicamente)
                objeto = dicionario_objetos[
                    array_nomes_objetos[1] + str(linha)]  # Resgatando o objeto que identifica evento (código ou painel)
                if objeto.get_name().__contains__('entry'):
                    if objeto.get_text() == '':
                        continue
                    else:
                        evento = SubElement(eventos, array_nomes_objetos[1].split('_')[0].upper())
                        evento.text = objeto.get_text().strip().upper()
                        for i in range(2, len(array_nomes_objetos)):
                            caixa = dicionario_objetos[array_nomes_objetos[i] + str(linha)]
                            if array_nomes_objetos[i].__contains__('entry'):
                                evento.set(array_nomes_objetos[i].split('_')[2], str(caixa.get_text()).strip().upper())
                            elif array_nomes_objetos[i].__contains__('combobox'):
                                evento.set(array_nomes_objetos[i].split('_')[2], str(caixa.get_active_text()))
                            elif array_nomes_objetos[i].__contains__('checkbtt'):
                                evento.set(array_nomes_objetos[i].split('_')[2], str(caixa.get_active()))
                elif objeto.get_name().__contains__('combobox'):
                    if objeto.get_active() == -1:
                        continue
                    else:
                        evento = SubElement(eventos, array_nomes_objetos[1].split('_')[0].upper())
                        evento.text = objeto.get_active_text().strip().upper()
                        for i in range(2, len(array_nomes_objetos)):
                            caixa = dicionario_objetos[array_nomes_objetos[i] + str(linha)]
                            if array_nomes_objetos[i].__contains__('entry'):
                                evento.set(array_nomes_objetos[i].split('_')[2], str(caixa.get_text()).strip().upper())
                            elif array_nomes_objetos[i].__contains__('combobox'):
                                evento.set(array_nomes_objetos[i].split('_')[2], str(caixa.get_active_text()))
                            elif array_nomes_objetos[i].__contains__('checkbtt'):
                                evento.set(array_nomes_objetos[i].split('_')[2], str(caixa.get_active()))
            except:
                objeto = builder.get_object(array_nomes_objetos[1] + str(linha))
                if objeto.get_name().__contains__('entry'):
                    if objeto.get_text() == '':
                        continue
                    else:
                        evento = SubElement(eventos, array_nomes_objetos[1].split('_')[0].upper())
                        evento.text = objeto.get_text().strip().upper()
                        for i in range(2, len(array_nomes_objetos)):
                            caixa = builder.get_object(array_nomes_objetos[i] + str(linha))
                            if array_nomes_objetos[i].__contains__('entry'):
                                evento.set(array_nomes_objetos[i].split('_')[2], str(caixa.get_text()).strip().upper())
                            elif array_nomes_objetos[i].__contains__('combobox'):
                                evento.set(array_nomes_objetos[i].split('_')[2], str(caixa.get_active_text()))
                            elif array_nomes_objetos[i].__contains__('checkbtt'):
                                evento.set(array_nomes_objetos[i].split('_')[2], str(caixa.get_active()))
                        if array_nomes_objetos == self.dicionario_nomes_objetos[3]:
                            caixa = builder.get_object('paisage_entry_rdp-central-de_1')
                            evento.set('rdp-central-de', str(caixa.get_text()))
                            caixa = builder.get_object('paisage_entry_rdp-central-ate_1')
                            evento.set('rdp-central-ate', str(caixa.get_text()))
                elif objeto.get_name().__contains__('combobox'):
                    if objeto.get_active() == -1:
                        continue
                    else:
                        evento = SubElement(eventos, array_nomes_objetos[1].split('_')[0].upper())
                        evento.text = objeto.get_active_text().strip().upper()
                        for i in range(2, len(array_nomes_objetos)):
                            caixa = builder.get_object(array_nomes_objetos[i] + str(linha))
                            if array_nomes_objetos[i].__contains__('entry'):
                                evento.set(array_nomes_objetos[i].split('_')[2], str(caixa.get_text()).strip().upper())
                            elif array_nomes_objetos[i].__contains__('combobox'):
                                evento.set(array_nomes_objetos[i].split('_')[2], str(caixa.get_active_text()))
                            elif array_nomes_objetos[i].__contains__('checkbtt'):
                                evento.set(array_nomes_objetos[i].split('_')[2], str(caixa.get_active()))
        # Funções de ação gerais

    # Função para adicionar uma linha
    def adicionar_linha(self, Linhas_Removidas, Numero_linhas_ativas, tipo_evento, array_nomes_objetos,
                        dicionario_objetos, tabela_evento):
        if tipo_evento == 'LT':
            array_objetos = [Gtk.CheckButton(),
                             Gtk.Entry(),
                             Gtk.Entry(),
                             Gtk.Entry(),
                             Gtk.Entry(),
                             Gtk.Entry(),
                             Gtk.ComboBoxText(),
                             Gtk.ComboBoxText(),
                             Gtk.CheckButton(),
                             Gtk.CheckButton(),
                             Gtk.CheckButton(),
                             Gtk.CheckButton()]  # Array que cria novos objetos do evento LT na sequência da tela
        elif tipo_evento == 'Trafo':
            array_objetos = [Gtk.CheckButton(),
                             Gtk.Entry(),
                             Gtk.Entry(),
                             Gtk.Entry(),
                             Gtk.Entry(),
                             Gtk.Entry(),
                             Gtk.ComboBoxText(),
                             Gtk.ComboBoxText(),
                             Gtk.CheckButton(),
                             Gtk.CheckButton(),
                             Gtk.CheckButton(),
                             Gtk.ComboBoxText(),
                             Gtk.ComboBoxText(),
                             Gtk.ComboBoxText()]
        elif tipo_evento == 'vaotrans':
            array_objetos = [Gtk.CheckButton(),
                             Gtk.Entry(),
                             Gtk.Entry(),
                             Gtk.CheckButton(),
                             Gtk.ComboBoxText(),
                             Gtk.Entry(),
                             Gtk.Entry(),
                             ]
        elif tipo_evento == 'paisage':
            array_objetos = [Gtk.CheckButton(),
                             Gtk.Entry(),
                             Gtk.ComboBoxText(),
                             Gtk.Entry(),
                             Gtk.Entry(),
                             Gtk.Entry(),
                             Gtk.CheckButton(),
                             Gtk.Entry(),
                             Gtk.CheckButton(),
                             Gtk.Entry(),
                             Gtk.Entry(),
                             Gtk.Entry()]
        elif tipo_evento == 'reator':
            array_objetos = [Gtk.CheckButton(),
                             Gtk.Entry(),
                             Gtk.Entry(),
                             Gtk.CheckButton(),
                             Gtk.ComboBoxText(),
                             Gtk.CheckButton(),
                             Gtk.CheckButton(),
                             Gtk.Entry(),
                             Gtk.Entry()]

        elif tipo_evento == 'acesso':
            array_objetos = [Gtk.CheckButton(),
                             Gtk.Entry(),
                             Gtk.Entry(),
                             Gtk.CheckButton(),
                             Gtk.Entry(),
                             Gtk.Entry(),
                             Gtk.ComboBoxText(),
                             Gtk.CheckButton(),
                             Gtk.Entry(),
                             Gtk.Entry(),
                             Gtk.CheckButton(),
                             Gtk.Entry(),
                             Gtk.Entry(),
                             Gtk.CheckButton(),
                             Gtk.Entry(),
                             Gtk.Entry(),
                             Gtk.Entry()]
        elif tipo_evento == 'tterra':
            array_objetos = [Gtk.CheckButton(),
                             Gtk.Entry(),
                             Gtk.Entry(),
                             Gtk.Entry(),
                             Gtk.Entry()]
        elif tipo_evento == 'protbarra':
            array_objetos = [Gtk.CheckButton(),
                             Gtk.Entry(),
                             Gtk.Entry(),
                             Gtk.ComboBoxText(),
                             Gtk.CheckButton(),
                             Gtk.Entry()]

        elif tipo_evento == 'bcapshunt':
            array_objetos = [Gtk.CheckButton(),
                             Gtk.Entry(),
                             Gtk.Entry(),
                             Gtk.ComboBoxText(),
                             Gtk.CheckButton(),
                             Gtk.CheckButton()]

        elif tipo_evento == 'bcapserie':
            array_objetos = [Gtk.CheckButton(),
                             Gtk.Entry(),
                             Gtk.Entry()]

        elif tipo_evento == 'ece':
            array_objetos = [Gtk.CheckButton(),
                             Gtk.Entry(),
                             Gtk.Entry()]

        elif tipo_evento == 'sistreg':
            array_objetos = [Gtk.CheckButton(),
                             Gtk.ComboBoxText(),
                             Gtk.ComboBoxText(),
                             Gtk.Entry()]

        elif tipo_evento == 'prepreen':
            array_objetos = [Gtk.CheckButton(),
                             Gtk.Entry()]

        elif tipo_evento == 'compsinc':
            array_objetos = [Gtk.CheckButton(),
                             Gtk.Entry(),
                             Gtk.Entry()]

        if not Linhas_Removidas:  # identifica se já foi removida alguma linha anteriormente
            # Caso nenhuma linha tenha sido removida, a nova linha a adicionar será uma a mais da ultima
            indice_a_adicionar = max(Numero_linhas_ativas, key=int) + 1
            Numero_linhas_ativas.append(indice_a_adicionar)
        else:
            # Caso já tenha sido removida alguma linha, a linha a adicionar será a de índice menor dentre as removidas
            indice_a_adicionar = min(Linhas_Removidas, key=int)
            Numero_linhas_ativas.append(indice_a_adicionar)
            if indice_a_adicionar == 1 and 0 in Numero_linhas_ativas:
                del Numero_linhas_ativas[Numero_linhas_ativas.index(0)]
            del Linhas_Removidas[Linhas_Removidas.index(indice_a_adicionar)]
        for i in range(0, len(array_nomes_objetos)):  # For para tratar cada objeto da linha
            objeto = array_objetos[i]
            objeto.set_name(array_nomes_objetos[i] + str(
                indice_a_adicionar))  # Seta o nome do objeto para obedecer o padrão, junto ao número sequencial da linha
            objeto.props.visible = True  # Faz o objeto ficar visível na tela
            dicionario_objetos[
                objeto.get_name()] = objeto  # Armazena o objeto no dicionário para que seja possível acessá-lo posteriormente
            tabela_evento.attach(objeto, i, i + 1, indice_a_adicionar + 1,
                                 indice_a_adicionar + 2)  # organiza os objetos na tela, dentro do objeto tabela
            # Configurações adicionais a objetos específicos
            if objeto.get_name().__contains__('selec_linha') or objeto.get_name().__contains__('checkbtt'):
                objeto.set_halign(Gtk.Align.CENTER)
                objeto.set_valign(Gtk.Align.CENTER)
            if objeto.get_name().__contains__('arranjo'):
                self.preenche_arranjo(objeto)

            if tipo_evento == 'LT':
                if objeto.get_name().__contains__('religamento') and tipo_evento == 'LT':
                    objeto.append_text('MONO/TRI')
                    objeto.append_text('TRIPOLAR')
                if objeto.get_name().__contains__('f9') and tipo_evento == 'LT':
                    objeto.set_property('margin-start', 10)
                if objeto.get_name().__contains__('87l') and tipo_evento == 'LT':
                    objeto.set_property('margin-start', 20)
                    objeto.set_property('margin-end', 30)
            if tipo_evento == 'Trafo':
                if objeto.get_name().__contains__('equip'):
                    objeto.append_text('Banco Monof.')
                    objeto.append_text('Trifásico')
                if objeto.get_name().__contains__('combobox_relacao'):
                    objeto.append_text('500/230')
                    objeto.append_text('500/230/13,8')
                    objeto.append_text('230/138')
                    objeto.append_text('230/138/13,8')
                    objeto.append_text('230/69')
                    objeto.append_text('230/69/13,8')
                    objeto.append_text('138/69')
                    objeto.append_text('138/69/13,8')
                    objeto.append_text('230/6,9')
                    objeto.append_text('69/13,8')
                if objeto.get_name().__contains__('prot'):
                    objeto.append_text('PP/PA')
                    objeto.append_text('PU/PG')
                if objeto.get_name().__contains__('f9'):
                    objeto.set_property('margin-start', 10)
                    objeto.set_property('margin-end', 10)
            if tipo_evento == 'paisage':
                if objeto.get_name().__contains__('sagebastidor'):
                    objeto.append_text('SAGE')
                    objeto.append_text('BASTIDOR')
                if objeto.get_name().__contains__('nporta-fw'):
                    objeto.set_halign(Gtk.Align.CENTER)
                    objeto.set_valign(Gtk.Align.CENTER)
            if tipo_evento == 'reator':
                if objeto.get_name().__contains__('equip'):
                    objeto.append_text('Banco Monof.')
                    objeto.append_text('Trifásico')
            if tipo_evento == 'sistreg':
                if objeto.get_name().__contains__('nome'):
                    objeto.append_text('SAGE')
                    objeto.append_text('UTR-')
                    objeto.append_text('PCPG')
                    objeto.append_text('SART')
                if objeto.get_name().__contains__('tensao-reg'):
                    objeto.append_text('230kV')
                    objeto.append_text('138kV')
                    objeto.append_text('69kV')
                    objeto.append_text('13,8kV')

        self.dicionario_geral_objetos_dinamicos[tipo_evento] = dicionario_objetos
        self.dicionario_gera_num_de_linhas_ativas[tipo_evento] = Numero_linhas_ativas
        self.dicionario_geral_linhas_removidas[tipo_evento] = Linhas_Removidas
        return indice_a_adicionar

    # Função principal para realizar a cópia das linhas selecionadas
    def prepara_para_duplicar(self, Linhas_Removidas, Numero_linhas_ativas, tipo_evento, array_nomes_objetos,
                              dicionario_objetos, tabela_evento):
        for linha in Numero_linhas_ativas:  # Varre todas as linhas para achar os checkboxes selecionados
            try:  # Caso para os objetos que foram criados no botão adicionar (dinamicamente)
                check_select: Gtk.CheckButton = dicionario_objetos[
                    array_nomes_objetos[0] + str(linha)]  # Resgatando o objeto checkbutton da linha
                if check_select.get_active():
                    linhas_ocupadas, linha_a_duplicar = self.todas_linhas_preenchidas(linha, Numero_linhas_ativas,
                                                                                      array_nomes_objetos,
                                                                                      dicionario_objetos)
                    if linhas_ocupadas:
                        linha_a_duplicar = self.adicionar_linha(Linhas_Removidas, Numero_linhas_ativas, tipo_evento,
                                                                array_nomes_objetos, dicionario_objetos,
                                                                tabela_evento)
                        self.duplicar_linha(linha, linha_a_duplicar, True, array_nomes_objetos, dicionario_objetos)
                    else:
                        self.duplicar_linha(linha, linha_a_duplicar, True, array_nomes_objetos, dicionario_objetos)
            except:
                check_select: Gtk.CheckButton = builder.get_object(array_nomes_objetos[0] + str(1))
                if check_select.get_active():
                    linhas_ocupadas, linha_a_duplicar = self.todas_linhas_preenchidas(linha, Numero_linhas_ativas,
                                                                                      array_nomes_objetos,
                                                                                      dicionario_objetos)
                    if linhas_ocupadas:
                        linha_a_duplicar = self.adicionar_linha(Linhas_Removidas, Numero_linhas_ativas, tipo_evento,
                                                                array_nomes_objetos, dicionario_objetos,
                                                                tabela_evento)
                        self.duplicar_linha(linha, linha_a_duplicar, False, array_nomes_objetos, dicionario_objetos)
                    else:
                        self.duplicar_linha(linha, linha_a_duplicar, False, array_nomes_objetos, dicionario_objetos)

    # Função que transfere as propriedades da linha selecionada e para a(s) linha(s) inferior(es)
    def duplicar_linha(self, linha, linha_a_duplicar, widget_dinamico, array_nomes_objetos, dicionario_objetos):
        if widget_dinamico:
            for objeto in array_nomes_objetos:
                if objeto != array_nomes_objetos[0]:
                    if 'entry' in objeto:
                        objeto_selecionado: Gtk.Entry = dicionario_objetos[objeto + str(linha)]
                        objeto_duplicado: Gtk.Entry = dicionario_objetos[objeto + str(linha_a_duplicar)]
                        objeto_duplicado.set_text(objeto_selecionado.get_text())

                    elif 'combobox' in objeto:
                        objeto_selecionado: Gtk.ComboBoxText = dicionario_objetos[objeto + str(linha)]
                        objeto_duplicado: Gtk.ComboBoxText = dicionario_objetos[objeto + str(linha_a_duplicar)]
                        objeto_duplicado.set_active(objeto_selecionado.get_active())

                    elif 'checkbtt' in objeto:
                        objeto_selecionado: Gtk.CheckButton = dicionario_objetos[objeto + str(linha)]
                        objeto_duplicado: Gtk.CheckButton = dicionario_objetos[objeto + str(linha_a_duplicar)]
                        objeto_duplicado.set_active(objeto_selecionado.get_active())
        else:
            for objeto in array_nomes_objetos:
                if objeto != array_nomes_objetos[0]:
                    if 'entry' in objeto:
                        objeto_selecionado: Gtk.Entry = builder.get_object(objeto + str(1))
                        objeto_duplicado: Gtk.Entry = dicionario_objetos[objeto + str(linha_a_duplicar)]
                        objeto_duplicado.set_text(objeto_selecionado.get_text())

                    elif 'combobox' in objeto:
                        objeto_selecionado: Gtk.ComboBoxText = builder.get_object(objeto + str(1))
                        objeto_duplicado: Gtk.ComboBoxText = dicionario_objetos[objeto + str(linha_a_duplicar)]
                        objeto_duplicado.set_active(objeto_selecionado.get_active())

                    elif 'checkbtt' in objeto:
                        objeto_selecionado: Gtk.CheckButton = builder.get_object(objeto + str(1))
                        objeto_duplicado: Gtk.CheckButton = dicionario_objetos[objeto + str(linha_a_duplicar)]
                        objeto_duplicado.set_active(objeto_selecionado.get_active())

    # Função que limpa as configurações da(s) linha(s) selecionada(s)
    def limpar_linha(self, Numero_linhas_ativas, array_nomes_objetos, dicionario_objetos):
        for linha in Numero_linhas_ativas:  # Varre todas as linhas para achar os checkboxes selecionados
            try:  # Caso para os objetos que foram criados no botão adicionar (dinamicamente)
                check_select: Gtk.CheckButton = dicionario_objetos[
                    array_nomes_objetos[0] + str(linha)]  # Resgatando o objeto checkbutton da linha
                if check_select.get_active():
                    for objeto in array_nomes_objetos:
                        if objeto != array_nomes_objetos[0]:
                            if 'entry' in objeto:
                                objeto_selecionado: Gtk.Entry = dicionario_objetos[objeto + str(linha)]
                                objeto_selecionado.set_text('')
                            elif 'combobox' in objeto:
                                objeto_selecionado: Gtk.ComboBoxText = dicionario_objetos[objeto + str(linha)]
                                objeto_selecionado.set_active(-1)

                            elif 'checkbtt' in objeto:
                                objeto_selecionado: Gtk.CheckButton = dicionario_objetos[objeto + str(linha)]
                                objeto_selecionado.set_active(False)
            except:
                check_select: Gtk.CheckButton = builder.get_object(array_nomes_objetos[0] + str(linha))
                if check_select.get_active():
                    for objeto in array_nomes_objetos:
                        if objeto != array_nomes_objetos[0]:
                            if 'entry' in objeto:
                                objeto_selecionado: Gtk.Entry = builder.get_object(objeto + str(linha))
                                objeto_selecionado.set_text('')
                            elif 'combobox' in objeto:
                                objeto_selecionado: Gtk.ComboBoxText = builder.get_object(objeto + str(linha))
                                objeto_selecionado.set_active(-1)
                            elif 'checkbtt' in objeto:
                                objeto_selecionado: Gtk.CheckButton = builder.get_object(objeto + str(linha))
                                objeto_selecionado.set_active(False)

    # Função que exclui a(s) linha(s) selecionada(s)
    def exclui_linha(self, Linhas_Removidas, Numero_linhas_ativas, array_nomes_objetos, dicionario_objetos,
                     Nome_evento):
        linhas_removidas_agora = []
        for linha in Numero_linhas_ativas:  # Varre todas as linhas para achar os checkboxes selecionados
            try:  # Caso para os objetos que foram criados no botão adicionar (dinamicamente)
                check_select: Gtk.CheckButton = dicionario_objetos[
                    array_nomes_objetos[0] + str(linha)]  # Resgatando o objeto checkbutton da linha
                if check_select.get_active():
                    Linhas_Removidas.append(linha)
                    linhas_removidas_agora.append(linha)
                    for item in array_nomes_objetos:
                        objeto = dicionario_objetos[
                            item + str(linha)]  # Resgatando os objetos armazenados no dicionário
                        objeto.destroy()  # Comando para deletar os objetos
                        del dicionario_objetos[item + str(linha)]

            except:  # Caso para os objetos que foram criados no glade (primeira Linha)
                check_select: Gtk.CheckButton = builder.get_object(array_nomes_objetos[0] + str(linha))
                if check_select.get_active():
                    Linhas_Removidas.append(linha)
                    linhas_removidas_agora.append(linha)
                    for item in array_nomes_objetos:
                        objeto = builder.get_object(item + str(1))
                        objeto.destroy()

        for linha_remov in linhas_removidas_agora:  # Remove as linhas removidas do array de linhas ativas
            if linha_remov == 1:
                Numero_linhas_ativas[Numero_linhas_ativas.index(linha_remov)] = 0
            else:
                del Numero_linhas_ativas[Numero_linhas_ativas.index(linha_remov)]
        self.dicionario_geral_linhas_removidas[Nome_evento] = Linhas_Removidas
        self.dicionario_gera_num_de_linhas_ativas[Nome_evento] = Numero_linhas_ativas
        self.dicionario_geral_objetos_dinamicos[Nome_evento] = dicionario_objetos

    # Função que seleciona todas as linhas
    def selecionar_todas(self, Numero_linhas_ativas, array_nomes_objetos, dicionario_objetos):
        for linha in Numero_linhas_ativas:  # Varre todas as linhas para achar os checkboxes selecionados
            try:  # Caso para os objetos que foram criados no botão adicionar (dinamicamente)
                check_select: Gtk.CheckButton = dicionario_objetos[
                    array_nomes_objetos[0] + str(linha)]  # Resgatando o objeto checkbutton da linha
                if not check_select.get_active():
                    check_select.set_active(True)
            except:
                check_select: Gtk.CheckButton = builder.get_object(array_nomes_objetos[0] + str(linha))
                if not check_select.get_active():
                    check_select.set_active(True)

    # Função para retornar a tela ao estado inicial
    def restaurar_tela(self):  # Função que retorna os objetos para o estado inicial do programa (arquivo novo)
        for i in range(0, 14):
            nome_evento = self.dicionario_nomes_objetos[i][1].split('_')[0]
            array_num_linhas_ativas = self.dicionario_gera_num_de_linhas_ativas[nome_evento]
            if len(array_num_linhas_ativas) >= 1:
                self.selecionar_todas(array_num_linhas_ativas, self.dicionario_nomes_objetos[i],
                                      self.dicionario_geral_objetos_dinamicos[nome_evento])
                self.exclui_linha(self.dicionario_geral_linhas_removidas[nome_evento],
                                  self.dicionario_gera_num_de_linhas_ativas[nome_evento],
                                  self.dicionario_nomes_objetos[i],
                                  self.dicionario_geral_objetos_dinamicos[nome_evento], nome_evento)
            self.adicionar_linha(self.dicionario_geral_linhas_removidas[nome_evento],
                                 self.dicionario_gera_num_de_linhas_ativas[nome_evento],
                                 nome_evento, self.dicionario_nomes_objetos[i]
                                 , self.dicionario_geral_objetos_dinamicos[nome_evento],
                                 self.dicionario_geral_tabelas[nome_evento])
        for nome in ['saux_entry_nome-painel-ua_', 'saux_combobox_tensao-ca_', 'saux_combobox_tensao-cc_',
                     'paisage_entry_rdp-central-de_', 'paisage_entry_rdp-central-ate_']:
            objeto = builder.get_object(nome + '1')
            try:
                objeto.set_text('')
            except:
                objeto.set_active(-1)
        self.codigo_se.set_text('')
        self.usuario.set_text('')
        self.fornecedor.set_text('')
        self.descricao.set_text('')
        self.nome_arq_saida = 'Arqconf-novo'  # Nome do arquivo de saída
        seq_arq = 0  # Sequência do número de arquivo
        while os.path.exists(
                self.nome_arq_saida + '.fas'):  # Enquanto existir na pasta um arquivo com o nome definido
            seq_arq += 1  # Adicionar um a sequência do número do arquivo
            self.nome_arq_saida = self.nome_arq_saida.split('_')[0] + '_' + str(seq_arq)  #
        self.nome_arq_saida = self.nome_arq_saida + '.fas'
        self.window.set_title(self.nome_arq_saida)

        # Lista elementos descendentes: atributos

    # Funções Adicionais Auxiliares

    def preenche_arranjo(self, objeto):

        for arranjo in self.Arranjos:
            objeto.append_text(arranjo)

    # Função que retorna o documento xml em forma organizada para leitura humana
    def prettify(self, elem):
        """Return a pretty-printed XML string for the Element.
        """
        rough_string = tostring(elem, 'utf-8')
        reparsed = minidom.parseString(rough_string)
        return reparsed.toprettyxml(indent="  ")

    # Função que retorna o índice do texto dentro do ComboboxText
    def indice_pela_string(self, combobox, string):
        for i in range(0, 10):
            try:
                combobox.set_active(i)
                if str(combobox.get_active_text()).lower() == string.lower():
                    return i
                    break
            except:
                break

    # Função que identifica se todas as linhas já estão preenchidas (Usada na função de prapara_para_duplicar)
    def todas_linhas_preenchidas(self, linha, Numero_linhas_ativas, array_nomes_objetos, dicionario_objetos):

        preenchidas = True
        for linha_a_duplicar in Numero_linhas_ativas:
            if linha_a_duplicar != linha:
                if (array_nomes_objetos[0] + str(linha_a_duplicar)) in dicionario_objetos:
                    codigo_evento: Gtk.Entry = dicionario_objetos[array_nomes_objetos[1] + str(linha_a_duplicar)]
                    if codigo_evento.get_name().__contains__('entry'):
                        if codigo_evento.get_text() == '':
                            preenchidas = False
                            break
                        else:
                            preenchidas = True
                    elif codigo_evento.get_name().__contains__('combobox'):
                        if codigo_evento.get_active() == -1:
                            preenchidas = False
                            break
                        else:
                            preenchidas = True
        if preenchidas:
            return [preenchidas, 0]
        else:
            return [preenchidas, linha_a_duplicar]

# Eventos ligados ao salvamento do arquivo de configuração

    # Evento de clique do botão salvar do diálogo de salvamento
    def on_arqconf_button_salvar_clicked(self, button):
        self.nome_arq_saida = str(
            self.arqconf_salvar_dialogo.get_current_folder() + '\\' + self.nome_arqconf.get_text())
        if not self.nome_arq_saida.endswith('.fas'):
            self.nome_arq_saida = self.nome_arq_saida + '.fas'
        if os.path.exists(self.nome_arq_saida):
            sobrescrever = pergunta_sim_nao('Aviso', 'O nome do arquivo especificado já existe, deseja sobrescrever?')
            if sobrescrever:
                self.salvar_arqconf()
                if self.gerar_lp:
                    self.gerarlp()
                    self.gerar_lp = False
                if self.checar_lp:
                    self.checarlp()
                    self.checar_lp = False
        else:
            self.salvar_arqconf()
            if self.gerar_lp:
                self.gerarlp()
                self.gerar_lp = False
            if self.checar_lp:
                self.checarlp()
                self.checar_lp = False

    # Evento de clique do botão cancelar do diálogo de salvamento
    def on_arqconf_button_cancelar_clicked(self, button):
        self.arqconf_salvar_dialogo.hide()

    # Evento da ativação do widget ou menu de salvar como
    def on_arqconf_salvar_como_activate(self, button):
        self.arqconf_salvar_dialogo.set_current_folder()
        self.arqconf_salvar_dialogo.show()
        self.nome_arqconf.set_text('')

    # Função que realiza o salvamento
    def salvar_arqconf(self):
        root = Element('Arqconf', data='{}'.format(date.today()),
                       fornecedor=self.fornecedor.get_text(),
                       usuário=self.usuario.get_text(),
                       versão=self.versao)
        eventos = SubElement(root, 'Eventos', codigo_se=str(self.codigo_se.get_text().upper()),
                             lppadrao=str(self.Lppadrao.get_filename()).rsplit('\\', 1)[1],
                             descrição=str(self.descricao.get_text()))
        for i in range(0, 14):
            Nome_Aba = self.dicionario_nomes_objetos[i][1].split('_')[0]  # Pega o nome do evento no dicionario de nomes
            self.recolhe_dados(self.dicionario_gera_num_de_linhas_ativas[Nome_Aba],
                               self.dicionario_nomes_objetos[i],
                               self.dicionario_geral_objetos_dinamicos[Nome_Aba], eventos)
        self.recolhe_dados([1], self.dicionario_nomes_objetos[14], {}, eventos)  # Recolhe dados do serviço auxilar
        rdp_de = builder.get_object('paisage_entry_rdp-central-de_1')
        if rdp_de.get_text() != "":
            rdp_ate = builder.get_object('paisage_entry_rdp-central-ate_1')
            Rdp_central = SubElement(eventos, 'rdp_central', rdpde=str(rdp_de.get_text()),
                                     rdpate=str(rdp_ate.get_text()))

        ElementTree(root).write(self.nome_arq_saida, 'UTF-8')
        if str(self.descricao.get_text()):
            self.window.set_title(self.nome_arqconf.get_text() + '-' + str(self.descricao.get_text()))
        else:
            self.window.set_title(self.nome_arqconf.get_text())
        self.arqconf_salvar_dialogo.hide()
        self.arqconf_novo = False

    # Evento da ativação do widget ou menu de salvar
    def on_arqconf_salvar_activate(self, button):
        if self.arqconf_novo:
            self.arqconf_salvar_dialogo.show()
            self.nome_arqconf.set_text(self.nome_arq_saida.split('.')[0])
        else:
            self.salvar_arqconf()

# Eventos ligados ao abrir o arquivo de configuração

    def on_arqconf_abrir_activate(self, button):
        self.arqconf_abrir_dialogo.show()

    def on_arqconf_button_abrir_clicked(self, button):
        self.abrir_arquivo(self.arqconf_abrir_dialogo.get_filename())
        self.arqconf_abrir_dialogo.hide()

    def on_arqconf_abrir_button_cancelar_clicked(self, button):
        self.arqconf_abrir_dialogo.hide()

    # Função que faz efetivamente o carregamento do arquivo
    def abrir_arquivo(self, nome_arquivo):
        root = BeautifulSoup(open(nome_arquivo, 'r', encoding='utf-8'), 'html.parser')
        self.pathchecar = path.dirname(nome_arquivo)
        self.restaurar_tela()  # Retorna a tela para o estado inicial
        for i in range(0, 15):
            nome_evento = self.dicionario_nomes_objetos[i][1].split('_')[0]
            array_nomes_objetos = self.dicionario_nomes_objetos[i]
            if nome_evento.lower() != 'saux':
                dicionario_objetos = self.dicionario_geral_objetos_dinamicos[nome_evento]
            else:
                dicionario_objetos = []
            nome_evento_tag = nome_evento.lower()  # o beautifulsoap reconhece os campos apenas em minúsculo
            eventos = root.find_all(
                nome_evento_tag)  # captura todos os eventos com aquele nome, ex: lt ou trafo ou saux...
            if eventos:
                self.notebook.set_current_page(i)  # Deixa ativa a ultima aba a ser alterada
                for j in range(0, len(eventos)):
                    if j < 1:
                        for objeto in array_nomes_objetos:
                            if objeto != array_nomes_objetos[0]:
                                try:
                                    if objeto == array_nomes_objetos[1]:
                                        if 'entry' in objeto:
                                            objeto_duplicado: Gtk.Entry = dicionario_objetos[objeto + str(j + 1)]
                                            objeto_duplicado.set_text(str(eventos[j].string))
                                        elif 'combobox' in objeto:
                                            objeto_duplicado: Gtk.ComboBoxText = dicionario_objetos[
                                                objeto + str(j + 1)]
                                            objeto_duplicado.set_active(
                                                self.indice_pela_string(objeto_duplicado, str(eventos[j].string)))
                                    else:
                                        if 'entry' in objeto:
                                            objeto_duplicado: Gtk.Entry = dicionario_objetos[objeto + str(j + 1)]
                                            objeto_duplicado.set_text(eventos[j][objeto.split('_')[2].lower()])

                                        elif 'combobox' in objeto:
                                            objeto_duplicado: Gtk.ComboBoxText = dicionario_objetos[
                                                objeto + str(j + 1)]
                                            objeto_duplicado.set_active(self.indice_pela_string(objeto_duplicado,
                                                                                                eventos[j][
                                                                                                    objeto.split(
                                                                                                        '_')[
                                                                                                        2].lower()]))
                                        elif 'checkbtt' in objeto:
                                            objeto_duplicado: Gtk.CheckButton = dicionario_objetos[
                                                objeto + str(j + 1)]
                                            if eventos[j][objeto.split('_')[2].lower()] == 'True':
                                                objeto_duplicado.set_active(True)
                                            else:
                                                objeto_duplicado.set_active(False)
                                except:
                                    if objeto == array_nomes_objetos[1]:
                                        if 'entry' in objeto:
                                            objeto_duplicado: Gtk.Entry = builder.get_object(objeto + str(j + 1))
                                            objeto_duplicado.set_text(str(eventos[j].string))
                                        elif 'combobox' in objeto:
                                            objeto_duplicado: Gtk.ComboBoxText = builder.get_object(
                                                objeto + str(j + 1))
                                            objeto_duplicado.set_active(
                                                self.indice_pela_string(objeto_duplicado, str(eventos[j].string)))
                                    else:
                                        if 'entry' in objeto:
                                            objeto_duplicado: Gtk.Entry = builder.get_object(objeto + str(j + 1))
                                            objeto_duplicado.set_text(eventos[j][objeto.split('_')[2].lower()])

                                        elif 'combobox' in objeto:
                                            objeto_duplicado: Gtk.ComboBoxText = builder.get_object(
                                                objeto + str(j + 1))
                                            objeto_duplicado.set_active(self.indice_pela_string(objeto_duplicado,
                                                                                                str(eventos[j][
                                                                                                        objeto.split(
                                                                                                            '_')[
                                                                                                            2]])))
                    else:
                        self.adicionar_linha(self.dicionario_geral_linhas_removidas[nome_evento],
                                             self.dicionario_gera_num_de_linhas_ativas[nome_evento],
                                             nome_evento, array_nomes_objetos, dicionario_objetos,
                                             self.dicionario_geral_tabelas[nome_evento])
                        for objeto in array_nomes_objetos:
                            if objeto != array_nomes_objetos[0]:
                                if objeto == array_nomes_objetos[1]:
                                    if 'entry' in objeto:
                                        objeto_duplicado: Gtk.Entry = dicionario_objetos[objeto + str(j + 1)]
                                        objeto_duplicado.set_text(str(eventos[j].string))
                                    elif 'combobox' in objeto:
                                        objeto_duplicado: Gtk.ComboBoxText = dicionario_objetos[
                                            objeto + str(j + 1)]
                                        objeto_duplicado.set_active(
                                            self.indice_pela_string(objeto_duplicado, str(eventos[j].string)))
                                else:
                                    if 'entry' in objeto:
                                        objeto_duplicado: Gtk.Entry = dicionario_objetos[objeto + str(j + 1)]
                                        objeto_duplicado.set_text(eventos[j][objeto.split('_')[2].lower()])

                                    elif 'combobox' in objeto:
                                        objeto_duplicado: Gtk.ComboBoxText = dicionario_objetos[
                                            objeto + str(j + 1)]
                                        objeto_duplicado.set_active(self.indice_pela_string(objeto_duplicado,
                                                                                            eventos[j][
                                                                                                objeto.split('_')[
                                                                                                    2].lower()]))
                                    elif 'checkbtt' in objeto:
                                        objeto_duplicado: Gtk.CheckButton = dicionario_objetos[
                                            objeto + str(j + 1)]
                                        if eventos[j][objeto.split('_')[2].lower()] == 'True':
                                            objeto_duplicado.set_active(True)
                                        else:
                                            objeto_duplicado.set_active(False)
        rdpcentral = root.rdp_central
        if rdpcentral:
            rdp_de = builder.get_object('paisage_entry_rdp-central-de_1')
            rdp_ate = builder.get_object('paisage_entry_rdp-central-ate_1')
            rdp_de.set_text(str(rdpcentral['rdpde']))
            rdp_ate.set_text(str(rdpcentral['rdpate']))
        self.usuario.set_text(root.arqconf['usuário'])
        self.fornecedor.set_text(root.arqconf['fornecedor'])
        self.codigo_se.set_text(root.eventos['codigo_se'])
        self.descricao.set_text(root.eventos['descrição'])
        self.nome_arqconf.set_text(nome_arquivo.split('\\')[len(nome_arquivo.split('\\')) - 1])
        self.window.set_title(self.nome_arqconf.get_text())
        self.arqconf_novo = False
        self.arqconf_abrir_dialogo.hide()
        self.nome_arq_saida = nome_arquivo
        caminho_lp_padrao = os.getcwd() + '\\' + 'static\\Planilhas Padrao' + '\\' + root.eventos['lppadrao']
        if os.path.exists(caminho_lp_padrao):
            self.Lppadrao.set_filename(os.getcwd() + '\\' + 'static\\Planilhas Padrao' + '\\' + root.eventos['lppadrao'])
        else:
            mensagem_erro('Erro', 'Arquivo {} não encontrado na pasta Planilhas Padrao'.format(
                caminho_lp_padrao.split('\\')[len(caminho_lp_padrao.split('\\')) - 1]))
            self.Lppadrao.set_filename('')

# Evento da tela principal
    def on_arqconf_menubar_arquivo_sair_activate(self, button):
        Gtk.main_quit()

    def on_arqconf_menubar_ferramentas_cepel2excel_activate(self, menubar):
        mensagem_aviso('Aviso', 'você foi avisado')

    def on_janela_principal_destroy(self, window):
        Gtk.main_quit()  # Encerra a aplicação quando fechar a janela no X vermelho

# Sinais das funcionalidades

    def on_gerarlp_clicked(self, button):
        if self.arqconf_novo:
            self.arqconf_salvar_dialogo.show()
            self.nome_arqconf.set_text(self.nome_arq_saida.split('.')[0])
            self.gerar_lp = True
        else:
            self.gerarlp()

    def on_checarlp_clicked(self, button):
        if not self.Lp_a_checar.get_filename():
            mensagem_erro('Erro', 'Nenhuma LP a ser Checada foi selecionada')
        else:
            self.PlanilhaArquivoLPEditado = self.comboplan.get_active_text()
            if self.arqconf_novo:
                self.arqconf_salvar_dialogo.show()
                self.nome_arqconf.set_text(self.nome_arq_saida.split('.')[0])
                self.checar_lp = True
            else:
                self.checarlp()

    def checarlp(self):
        try:
            arq_conf = BeautifulSoup(open(self.nome_arq_saida, 'r', encoding='utf-8'),
                                     'html.parser')  # Abrir arquivo de cofiguração
        except:
            aviso = 'Arquivo \"' + self.nome_arq_saida + u'\" não encontrado'
            mensagem_erro('Erro', aviso)
        try:
            vers = arq_conf.arqconf['versão'].split('.')
            vers = list(map(int, vers))  # Transformar array de string em array de inteiro
            if vers < [2, 1, 0]:
                mensagem_erro('Erro', 'Deve ser usado arquivo de configuração com versão igual ou maior a 2.1.0')
            else:
                try:
                    processing(Checar_LP.checar,
                               {'LP_Padrao': self.Lppadrao.get_filename(), 'LP_Editado': self.caminhoArquivoLPEditado,
                                'planilha': self.PlanilhaArquivoLPEditado,
                                'Arq_Conf': self.nome_arq_saida,
                                'Diretorio_Padrao': self.Diretorio_de_salvamento})
                except:
                    print_exc(file=stdout)
                    mensagem_erro('Erro', 'Erro inesperado ao tentar checar lista de pontos.')
        except:
            mensagem_erro('Erro', 'Arquivo indicado não corresponde a arquivo de parametrização válido')

    def on_lp_a_checar_file_set(self, button):
        temp = self.Lp_a_checar.get_filename()
        self.caminhoArquivoLPEditado = temp
        try:
            book = load_workbook(temp)  # Abrir arquivo de a ser checado
        except:
            aviso = 'Arquivo \"' + temp + '\" não encontrado'
            self.button_checarlp.set_sensitive(False)
            self.menubar_checar.set_sensitive(False)
            self.toolbar_checar.set_sensitive(False)
            mensagem_erro('Erro', aviso)
        self.comboplan.remove_all()
        for nome_aba in book.sheetnames:
            self.comboplan.append_text(nome_aba)
        self.comboplan.set_active(0)
        self.button_checarlp.set_sensitive(True)
        self.menubar_checar.set_sensitive(True)
        self.toolbar_checar.set_sensitive(True)

    def gerarlp(self):
        try:
            arq_conf = BeautifulSoup(open(self.nome_arq_saida, 'r', encoding='utf-8'),
                                     'html.parser')  # Abrir arquivo de cofiguração
        except:
            aviso = 'Arquivo \"' + self.nome_arq_saida + u'\" não encontrado'
            mensagem_erro('Erro', aviso)
        try:
            vers = arq_conf.arqconf['versão'].split('.')
            vers = list(map(int, vers))  # Transformar array de string em array de inteiro
            if vers < [2, 1, 0]:
                mensagem_erro('Erro', 'Deve ser usado arquivo de configuração com versão igual ou maior a 2.1.0')
            else:
                try:
                    if self.codigo_se.get_text() != '' and arq_conf.eventos['codigo_se'] != '':
                        Diretorio_de_salvamento = self.Diretorio_de_salvamento
                        processing(Gerar_LP.gerar, {'LP_Padrao': self.Lppadrao.get_filename(),
                                                'Arq_Conf': self.nome_arq_saida,
                                                'Diretorio_Padrao': self.Diretorio_de_salvamento})
                    else:
                        mensagem_erro('Erro', 'Código da SE não especificado, se o campo estiver preenchido, salve antes de gerar')
                except:
                    print_exc(file=stdout)
                    mensagem_erro('Erro', 'Erro inesperado ao tentar gerar lista de pontos.')
        except:
            mensagem_erro('Erro', 'Arquivo indicado não corresponde a arquivo de parametrização válido')

    def set_nome_arq_saida(self, nome):
        self.lp_de_saida = nome

    def set_diretorio_salvamento(self, nome):
        self.Diretorio_de_salvamento = nome

    # Eventos ligados a função base SAGE para LP excel

    def on_Base_SAGE_para_LP_Excel_activate(self, menubar):
        self.base_para_lp = True
        self.dialogo_diretorio.set_title('Selecione o diretório onde estão os arquivos .dat (Pasta dados ou include)')
        self.dialogo_diretorio.show()


    def on_diretorio_dialogo_pasta_button_cancelar_clicked(self, button):
        self.dialogo_diretorio.hide()

    def on_diretorio_dialogo_selection_changed(self, selection):
        if self.dialogo_diretorio.get_filename():
            self.diretorio_dialogo_pasta_entry.set_text(self.dialogo_diretorio.get_filename())

    def on_diretorio_dialogo_pasta_button_selecionar_clicked(self, button):
        if self.base_para_lp:
            self.Base2Lp()
        else:
            self.Diretorio_de_salvamento = self.diretorio_dialogo_pasta_entry.get_text()
            root = BeautifulSoup(open(os.getcwd() + '\\' + 'ini.xml', 'r', encoding='utf-8'), 'html.parser')
            lppadrao = root.ini['arqlppadrao']
            root = Element('ini', data='{}'.format(date.today()),
                           versão=self.versao,
                           arqlppadrao=str(lppadrao),
                           diretorio_padrao=self.Diretorio_de_salvamento)
            ElementTree(root).write(os.getcwd() + '\\' + 'ini.xml', 'UTF-8')
            self.dialogo_diretorio.hide()

    def Base2Lp(self):
        try:
            from lp_lib.base2lp_sagon import base2xls
        except:
            mensagem_erro('Erro', 'Módulo base2lp não instalado')
            return 0
        diretorio = self.diretorio_dialogo_pasta_entry.get_text()
        self.dialogo_diretorio.hide()
        self.diretorio_dialogo_pasta_entry.set_text("")
        if diretorio:
            try:
                telas, include_indice = self.analisa_vtelas(diretorio)
                if telas != None:
                    processing(base2xls,{'base_path': diretorio, 'Diretorio_Padrao' : self.Diretorio_de_salvamento,
                                         'label_progressbar':self.label_progressbar,
                                         'progress_bar':self.progress_bar,
                                         'janela_carregando':self.janela_carregando,
                                         'telas':telas,
                                         'include_indice':include_indice})
                    self.janela_carregando.show_all()
                    self.base_para_lp = False
            except:
                print_exc(file=stdout)
                mensagem_erro('Erro', 'Erro inesperado ao tentar gerar lista de pontos.')

    def analisa_vtelas(self, diretorio):
        include_indice = 2  # Variável para identificar se foi selecionado um include
        try:
            nome_arq_tela = ''
            if os.path.exists('{}\\ihm\\VTelasBotoes.led'.format(diretorio.rsplit('\\', 2)[0])):
                nome_arq_tela = '{}\\ihm\\VTelasBotoes.led'.format(diretorio.rsplit('\\', 2)[0])
                include_indice = 2
            elif os.path.exists('{}\\ihm\\VTelasBotoes.led'.format(diretorio.rsplit('\\', 3)[0])):
                nome_arq_tela = '{}\\ihm\\VTelasBotoes.led'.format(diretorio.rsplit('\\', 3)[0])
                include_indice = 3
            arq_telas = open(nome_arq_tela)
            telas = []
            for linha in arq_telas.readlines():
                if 'TELA' in linha:
                    telas.append(linha.split('\"')[1].split()[1])
            arq_telas.close()
            return telas, include_indice
        except:
            if diretorio.rsplit('\\',3)[2] == 'dados':
                include_indice = 3
            elif diretorio.rsplit('\\',3)[2] == 'bd':
                include_indice = 2
            else:
                mensagem_erro('Erro', 'A Base não pôde ser carregada, pasta dados ou de include não foram selecionados')
                return None, None
            vtelas_manualmente = pergunta_sim_nao('Aviso', 'O arquivo {}\\ihm\\VTelasBotoes.led não pode ser carregado. Deseja selecionar manualmente o arquivo?  \
                            \n OBS:A coluna "TELA" da tabela gerada não será preenchida caso não selecione o arquivo.'.format(diretorio.rsplit('\\', include_indice)[0]))

            if vtelas_manualmente:
                self.dialogo_vtelas.set_current_folder(diretorio)
                self.dialogo_vtelas.run()
                if 'VTelasBotoes.led' in self.nome_arq_tela:
                    arq_telas = open(self.nome_arq_tela)
                    telas = []
                    for linha in arq_telas.readlines():
                        if 'TELA' in linha:
                            telas.append(linha.split('\"')[1].split()[1])
                    arq_telas.close()
                elif 'cancelar' in self.nome_arq_tela:
                    mensagem_aviso('Aviso', 'Operação cancelada.')
                    return None, None
                else:
                    mensagem_erro('Erro', 'Arquivo inválido.')
                    return None, None
            else:
                telas = []
            return telas,include_indice

    # Comparar Listas

    def on_arqconf_ferramentas_comparar_activate(self, button):

        self.filechooser_comparar_arquivo_lp_base: Gtk.FileChooserButton = builder.get_object(
            'filechooser_comparar_arquivo_lp_base')
        self.filechooser_comparar_lp_a_ser_checado: Gtk.FileChooserButton = builder.get_object(
            'filechooser_comparar_lp_a_ser_checado')
        self.arqlpbase_comparar_dialog: Gtk.FileChooserDialog = builder.get_object('arqlpbase_comparar_dialog')
        self.arqlpachecar_comparar_dialog: Gtk.FileChooserDialog = builder.get_object('arqlpachecar_comparar_dialog')
        self.combobox_arq_lp_base: Gtk.ComboBoxText = builder.get_object('combobox_arq_lp_base')
        self.combobox_arq_lp_a_checar: Gtk.ComboBoxText = builder.get_object('combobox_arq_lp_a_checar')

        self.janela_comparar.show_all()

    def on_arqlpbase_button_abrir_clicked(self, button):
        self.LPBase = self.arqlpbase_comparar_dialog.get_filename()
        self.selecionando_planilha_para_checar(self.LPBase, self.filechooser_comparar_arquivo_lp_base,
                                               self.combobox_arq_lp_base)
        self.arqlpbase_comparar_dialog.hide()
        if self.arqlpbase_comparar_dialog.get_filename() and self.arqlpachecar_comparar_dialog.get_filename():
            self.button_comparar.set_sensitive(True)

    def on_arqlpbase_button_fechar_clicked(self, button):
        self.arqlpbase_comparar_dialog.hide()

    def on_arqlpachecar_button_abrir_clicked(self, button):
        self.Checar = self.arqlpachecar_comparar_dialog.get_filename()
        self.selecionando_planilha_para_checar(self.Checar, self.filechooser_comparar_lp_a_ser_checado,
                                               self.combobox_arq_lp_a_checar)
        self.arqlpachecar_comparar_dialog.hide()
        if self.arqlpbase_comparar_dialog.get_filename() and self.arqlpachecar_comparar_dialog.get_filename():
            self.button_comparar.set_sensitive(True)

    def on_arqlpachecar_button_fechar_clicked(self, button):
        self.arqlpachecar_comparar_dialog.hide()

    def selecionando_planilha_para_checar(self, nome_arq, file_chooser_button, combobox):
        file_chooser_button.set_filename(nome_arq)
        try:
            book = load_workbook(nome_arq)  # Abrir arquivo de a ser checado
        except:
            aviso = 'Arquivo \"' + nome_arq + '\" não encontrado'
            mensagem_erro('Erro', aviso)
        combobox.remove_all()
        for nome_aba in book.sheetnames:
            combobox.append_text(nome_aba)
        combobox.set_active(0)

    def on_button_fechar_compararlp_clicked(self, button):
        self.filechooser_comparar_arquivo_lp_base.set_filename('')
        self.filechooser_comparar_lp_a_ser_checado.set_filename('')
        self.janela_comparar.hide()

    def on_button_compararlp_clicked(self, button):
        if not self.LPBase:
            mensagem_erro('Erro', 'Arquivo LP Base não selecionado')
        elif not self.Checar:
            mensagem_erro('Erro', 'Arquivo LP a ser Checado não selecionado')
        else:
            LP_Comparar.Comparar(self.LPBase, self.combobox_arq_lp_base, self.Checar, self.combobox_arq_lp_a_checar, self.Diretorio_de_salvamento)

    # Cepel2Excel
    def on_arqconf_cepel2excel_activate(self, button):

        self.filechooser_dialog_cepel2excel: Gtk.FileChooserDialog = builder.get_object(
            'filechooser_dialog_cepel2excel')
        self.filechooser_dialog_cepel2excel.show_all()

    def on_button_cepel2excel_abrir_clicked(self, button):
        self.arqcepel = self.filechooser_dialog_cepel2excel.get_filename()
        self.filechooser_dialog_cepel2excel.hide()
        if self.arqcepel:
            try:
                cepel2lp(self.arqcepel, self.Diretorio_de_salvamento)
            except:
                print_exc(file=stdout)
                mensagem_erro('Erro', 'Erro inesperado ao tentar checar lista de pontos.')

    def on_button_cepel2excel_cancelar_clicked(self, button):
        self.filechooser_dialog_cepel2excel.hide()

    # Funções de configuração do FAS
    def on_arqconf_menubar_config_seleclppadrao_activate(self, button):
        self.dialogo_lppadrao.show_all()

    def on_button_fechar_lppadrao_dialog_clicked(self, button):
        self.dialogo_lppadrao.hide()

    def on_button_abrir_lppadrao_dialog_clicked(self, button):
        self.Lppadrao.set_filename(self.dialogo_lppadrao.get_filename())
        lppadrao = str(self.Lppadrao.get_filename()).rsplit('\\', 1)[1]
        root = Element('ini', data='{}'.format(date.today()),
                       versão=self.versao,
                       arqlppadrao=str(lppadrao),
                       diretorio_padrao=self.Diretorio_de_salvamento)
        ElementTree(root).write(os.getcwd() + '\\' + 'ini.xml', 'UTF-8')
        self.dialogo_lppadrao.hide()

    def on_arqconf_menubar_config_selecdir_salvamento_activate(self, button):
        self.dialogo_diretorio.set_title('Selecione o diretório de salvamento padrão')
        self.dialogo_diretorio.show()

    def on_Vtelasbotoes_button_abrir_clicked(self, button):
        self.nome_arq_tela = self.dialogo_vtelas.get_filename()
        self.dialogo_vtelas.hide()

    def on_Vtelasbotoes_button_cancelar_clicked(self, button):
        self.nome_arq_tela = 'cancelar'
        self.dialogo_vtelas.hide()


def dialogo_abrir_arquivo_gerado(lp_de_saida, Diretorio_de_salvamento):
    abrirarquivo = pergunta_sim_nao('Aviso', 'Arquivo \"' + lp_de_saida
                                    + '\" gerado em ' + Diretorio_de_salvamento + '\n\n Deseja abrir o arquivo gerado agora?')
    if abrirarquivo:
        os.startfile(Diretorio_de_salvamento + '\\' + lp_de_saida)

if __name__ == '__main__':
    builder.connect_signals(Manipulador())  # Conecta os sinais da interface com a classe manipuladora "manipulador"
    Gtk.main()
